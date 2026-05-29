"""sync_tc_results.py — Map Playwright JSON report to TC Excel"""

import json, re, openpyxl, sys
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

JSON_PATH   = r"E:\AI\Ecom\automation-framework\test-results\report.json"
EXCEL_PATH  = r"E:\AI\Ecom\ecom-pdh\03_test-cases\AI_ISC_ecom-pdh_v1.1_TC_v1.0.xlsx"
OUT_PATH    = r"E:\AI\Ecom\ecom-pdh\03_test-cases\AI_ISC_ecom-pdh_v1.1_TC_v1.0_results_20260529.xlsx"
SHEET_NAME  = "DangKy_UltraFast"
ACTUAL_COL  = 9   # I
EXEC_COL    = 10  # J
ROUND_ROW   = 7

FILLS = {
    "Pass":  PatternFill("solid", fgColor="C6EFCE"),
    "Fail":  PatternFill("solid", fgColor="FFC7CE"),
    "Block": PatternFill("solid", fgColor="FFEB9C"),
}
FONTS = {
    "Pass":  Font(color="276221", bold=True, name="Calibri"),
    "Fail":  Font(color="9C0006", bold=True, name="Calibri"),
    "Block": Font(color="7D6608", bold=True, name="Calibri"),
}

# Bước 1: Parse JSON
with open(JSON_PATH, encoding="utf-8-sig") as f:
    report = json.load(f)

stats = report.get("stats", {})
start_dt = datetime.fromisoformat(
    stats["startTime"].replace("Z", "+00:00")
).astimezone().strftime("%Y-%m-%d %H:%M")

results = {}

def walk(nodes):
    for n in nodes:
        for spec in n.get("specs", []):
            m = re.search(r"TC_[A-Z0-9_]+\.\d+", spec.get("title", ""))
            if m:
                tc_id = m.group()
                ok = spec.get("ok", False)
                tests = spec.get("tests", [])
                raw = tests[0].get("results", [{}])[0].get("status", "") if tests else ""
                if ok or raw == "passed":
                    status = "Pass"
                elif raw == "skipped":
                    status = "Block"
                else:
                    status = "Fail"
                results[tc_id] = status
        walk(n.get("suites", []))

walk(report.get("suites", []))
print(f"[B1] Report parsed: {len(results)} TCs | Time: {start_dt}")

# Bước 2: Build TC map từ Excel
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb[SHEET_NAME]

tc_map = {}
for r in range(9, ws.max_row + 1):
    v = ws.cell(r, 2).value
    if v and re.match(r"TC_|API_", str(v)):
        tc_map[str(v).strip()] = r

print(f"[B2] Excel mapped: {len(tc_map)} TCs in sheet [{SHEET_NAME}]")

# Update Round header
rc = ws.cell(ROUND_ROW, ACTUAL_COL)
if rc.value and str(rc.value).strip() == "Round 1":
    rc.value = f"Round 1 - {start_dt}"

# Bước 3: Điền kết quả
filled_ok, filled_fail = [], []
not_in_excel, not_in_report = [], []

for tc_id, status in sorted(results.items()):
    if tc_id in tc_map:
        row = tc_map[tc_id]
        ca = ws.cell(row, ACTUAL_COL)
        ce = ws.cell(row, EXEC_COL)
        ca.value = status
        ca.fill  = FILLS[status]
        ca.font  = FONTS[status]
        ca.alignment = Alignment(horizontal="center")
        ce.value = "Auto"
        (filled_ok if status == "Pass" else filled_fail).append((tc_id, row))
    else:
        not_in_excel.append(tc_id)

for tc_id in sorted(tc_map.keys()):
    if tc_id not in results:
        not_in_report.append(tc_id)

# Bước 4: Lưu backup
wb.save(OUT_PATH)
print(f"[B4] Saved: {OUT_PATH}")

# Bước 5: Summary
print()
print("DIEN THANH CONG:")
for t, r in filled_ok:
    print(f"  [Pass] {t} -> row {r}")
for t, r in filled_fail:
    print(f"  [FAIL] {t} -> row {r}")

if not_in_report:
    print()
    print("TRONG EXCEL NHUNG KHONG CO TRONG REPORT (giu nguyen):")
    for t in not_in_report:
        print(f"  [INFO] {t}")

if not_in_excel:
    print()
    print("TRONG REPORT NHUNG KHONG CO TRONG EXCEL:")
    for t in not_in_excel:
        print(f"  [SKIP] {t}")

print()
print(f"SUMMARY => Pass:{len(filled_ok)} | Fail:{len(filled_fail)} | "
      f"NotInExcel:{len(not_in_excel)} | NotInReport:{len(not_in_report)}")
