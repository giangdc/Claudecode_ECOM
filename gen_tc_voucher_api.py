# -*- coding: utf-8 -*-
"""
gen-testcase-api-v3  —  Voucher API, ecom-pdh v1.1
Output: ecom-pdh/03_test-cases/api/AI_ISC_ecom-pdh_v1.1_TC_API_v1.1.xlsx

Spec: api doc v1.xlsx (updated 2026-05-27)
  API_01  POST /public/v1/voucher/list        -> data[] 8 fields
  API_02  POST /public/v1/voucher/content     -> voucher_code + content1-content6  [BA updated]
  API_03  POST /public/v1/voucher/apply       -> 17 top-level + applies[] 10 sub-fields

Resolved CLAs applied:
  VOUCHER-002  : HTTP 200 ALL cases (even failures)
  APISPEC-001  : Accept-Language removed from spec
  APISPEC-003  : API_02 != API_03 output (different schemas)
  VOUCHER-001  : API_02 = content1-6 (partial)

Open CLAs noted in TCs:
  APISPEC-002  : voucher_type "General" string mapping
  APISPEC-004  : error response body format
  VOUCHER-005  : Client-Id effect on result
  VOUCHER-006  : vouchers=[] behavior

v3 Security rules applied:
  3 auth headers -> min 12 auth TCs/endpoint
  SECURITY-INFERRED TCs for cross-auth, replay, IDOR, injection
"""

import os
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Paths ───────────────────────────────────────────────────────────────────
OUT_DIR  = r"E:\AI\Ecom\ecom-pdh\03_test-cases\api"
OUT_FILE = os.path.join(OUT_DIR, "AI_ISC_ecom-pdh_v1.1_TC_API_v1.1.xlsx")

# ─── Palette ─────────────────────────────────────────────────────────────────
C_HDR_BG  = "1F4E79"   # dark navy  — info block
C_HDR_FG  = "FFFFFF"
C_COL_BG  = "2E75B6"   # medium blue — column header row
C_GRP_BG  = "D6E4F0"   # light blue  — group rows
C_GRP_FG  = "1F4E79"
C_SEC_BG  = "FFF2CC"   # amber       — SECURITY-INFERRED rows
C_ODD     = "FFFFFF"
C_EVEN    = "EBF5FB"
C_PRI_H_BG = "FADBD8"; C_PRI_H_FG = "7B241C"
C_PRI_M_BG = "FEF9E7"; C_PRI_M_FG = "7D6608"
C_PRI_L_BG = "D6EAF8"; C_PRI_L_FG = "1A5276"

# ─── Helpers ─────────────────────────────────────────────────────────────────
def s(style="thin"):   return Side(style=style, color="B0B0B0")
def bdr():             return Border(left=s(), right=s(), top=s(), bottom=s())
def fill(h):           return PatternFill("solid", fgColor=h)
def font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)
def align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

COL_NAMES  = [
    "QC/AI", "Testcase ID", "Priority", "Nội Dung Test",
    "Điều Kiện / Dữ Liệu Test", "Các Bước Thực Hiện",
    "Kết Quả Mong Đợi (Expected Response)",
    "Round 1\nResult", "Round 1\nExecuted By", "Round 1\nBug ID", "Round 1\nRemark",
]
COL_W = [6, 12, 10, 45, 45, 42, 68, 12, 15, 12, 18]

def pri_style(p):
    return {
        "High":   (C_PRI_H_BG, C_PRI_H_FG),
        "Medium": (C_PRI_M_BG, C_PRI_M_FG),
        "Low":    (C_PRI_L_BG, C_PRI_L_FG),
    }.get(p, (C_ODD, "000000"))


# ─── Sheet builder ────────────────────────────────────────────────────────────
def build_sheet(wb, api):
    ws = wb.create_sheet(title=api["sheet"])

    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    N = len(COL_NAMES)

    # ── Info block rows 1-4 ──
    for r, (lbl, val) in enumerate(
        [("Mã API", api["code"]),
         ("Tên API", api["name"]),
         ("Phương thức", api["method"]),
         ("URL", api["url"])],
        start=1,
    ):
        ws.row_dimensions[r].height = 18
        lc = ws.cell(r, 1, lbl)
        lc.font, lc.fill, lc.alignment = font(bold=True, color=C_HDR_FG), fill(C_HDR_BG), align("left","center",False)
        vc = ws.cell(r, 2, val)
        vc.font, vc.fill, vc.alignment = font(), fill("DEEAF1"), align("left","center",False)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=N)

    # ── Column header row 5 ──
    ws.row_dimensions[5].height = 30
    for i, nm in enumerate(COL_NAMES, 1):
        c = ws.cell(5, i, nm)
        c.font, c.fill, c.alignment, c.border = font(bold=True,color=C_HDR_FG), fill(C_COL_BG), align("center","center"), bdr()

    # ── TC rows from row 6 ──
    row = 6
    tc_n = 0

    for grp in api["groups"]:
        ws.row_dimensions[row].height = 17
        gc = ws.cell(row, 1, grp["name"])
        gc.font, gc.fill, gc.alignment, gc.border = font(bold=True,color=C_GRP_FG), fill(C_GRP_BG), align("left","center",False), bdr()
        for col in range(2, N+1):
            c = ws.cell(row, col, "")
            c.fill, c.border = fill(C_GRP_BG), bdr()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        row += 1

        for tc in grp["tcs"]:
            tc_n += 1
            ws.row_dimensions[row].height = 62
            is_sec = "[SECURITY-INFERRED]" in tc.get("title","") or tc.get("security", False)
            bg = C_SEC_BG if is_sec else (C_ODD if tc_n % 2 else C_EVEN)
            pbg, pfg = pri_style(tc.get("priority","Medium"))

            vals = [
                "AI",
                f"{api['code']}.{tc_n}",
                tc.get("priority","Medium"),
                tc.get("title",""),
                tc.get("condition",""),
                tc.get("steps",""),
                tc.get("expected",""),
                "","","",""
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row, col, val)
                c.border = bdr()
                if col == 3:  # priority cell
                    c.font, c.fill, c.alignment = font(bold=True,color=pfg,size=9), fill(pbg), align("center","center",False)
                else:
                    c.font  = font(bold=(col==2), size=9, color="000000")
                    c.fill  = fill(bg)
                    c.alignment = align("center" if col<=2 else "left", "top")
            row += 1

    ws.freeze_panes = f"A6"
    return tc_n


# ═══════════════════════════════════════════════════════════════════════════════
# TC DATA
# ═══════════════════════════════════════════════════════════════════════════════

# ── Shared auth block (14 TCs, 3 groups) used by all 3 APIs ──────────────────
# Resolves: CLA-VOUCHER-002 (HTTP 200 for ALL cases, even failures)
# Header spec: X-Checkout-Token (required), Authorization (required), Client-Id (optional)

def auth_groups(ep):
    return [
        {
            "name": "Authentication / Authorization  —  X-Checkout-Token",
            "tcs": [
                {
                    "priority": "High",
                    "title":    f"Thiếu header X-Checkout-Token  [{ep}]",
                    "condition": "Authorization Bearer hợp lệ; X-Checkout-Token: không truyền",
                    "steps":    f"POST {ep}\nHeaders: Authorization: Bearer <valid_token>  (không có X-Checkout-Token)",
                    "expected": "HTTP 200\nsuccess = false (hoặc tương đương)\nResponse báo lỗi cho người dùng, không hiển thị voucher\n[Rule: \"Không nhập báo lỗi cho người dùng, không hiển thị voucher\"]\n[CLA-APISPEC-004: format error body TBD]",
                },
                {
                    "priority": "High",
                    "title":    f"X-Checkout-Token = \"\" (rỗng)  [{ep}]",
                    "condition": "Authorization Bearer hợp lệ; X-Checkout-Token: \"\" (empty string)",
                    "steps":    f"POST {ep}\nHeaders: X-Checkout-Token: \"\"  |  Authorization: Bearer <valid_token>",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi token không hợp lệ\n[CLA-APISPEC-004: format TBD]",
                },
                {
                    "priority": "High",
                    "title":    f"X-Checkout-Token sai giá trị (invalid)  [{ep}]",
                    "condition": "X-Checkout-Token = \"invalid_token_xyz_000\"; Authorization hợp lệ",
                    "steps":    f"POST {ep}\nHeaders: X-Checkout-Token: invalid_token_xyz_000  |  Authorization: Bearer <valid_token>",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi \"Không lấy được voucher\"\n[Rule: \"Nhập sai báo lỗi không lấy được voucher\"]\n[CLA-APISPEC-004: format TBD]",
                },
                {
                    "priority": "High",
                    "title":    f"X-Checkout-Token đã hết hạn  [{ep}]",
                    "condition": "X-Checkout-Token = token hợp lệ nhưng đã expired; Authorization hợp lệ",
                    "steps":    f"POST {ep}\nHeaders: X-Checkout-Token: <expired_token>  |  Authorization: Bearer <valid_token>",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi hết hạn, không hiển thị voucher\n[Rule: \"Hết hạn báo lỗi, không hiển thị voucher\"]\n[CLA-APISPEC-004: format TBD]",
                },
                {
                    "priority": "High",
                    "title":    f"[SECURITY-INFERRED] X-Checkout-Token bị tampered (payload chỉnh sửa)  [{ep}]",
                    "condition": "Decode token hợp lệ -> chỉnh sửa payload -> encode lại; Authorization hợp lệ",
                    "steps":    f"1. Tamper payload của X-Checkout-Token\n2. POST {ep} với token đã sửa",
                    "expected": "HTTP 200\nsuccess = false\nResponse từ chối token (chữ ký không khớp) — không trả data voucher",
                    "security": True,
                },
                {
                    "priority": "High",
                    "title":    f"[SECURITY-INFERRED] Replay X-Checkout-Token từ phiên checkout cũ  [{ep}]",
                    "condition": "X-Checkout-Token từ session checkout đã completed/expired",
                    "steps":    f"1. Lấy X-Checkout-Token từ phiên cũ\n2. POST {ep} với token đó",
                    "expected": "HTTP 200\nsuccess = false\nResponse từ chối token replay — không trả data voucher",
                    "security": True,
                },
            ],
        },
        {
            "name": "Authentication / Authorization  —  Authorization Bearer",
            "tcs": [
                {
                    "priority": "High",
                    "title":    f"Thiếu header Authorization  [{ep}]",
                    "condition": "X-Checkout-Token hợp lệ; Authorization: không truyền",
                    "steps":    f"POST {ep}\nHeaders: X-Checkout-Token: <valid>  (không có Authorization)",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi, không hiển thị voucher\n[Rule: \"Không nhập báo lỗi cho người dùng, không hiển thị voucher\"]",
                },
                {
                    "priority": "High",
                    "title":    f"Authorization Bearer rỗng (\"Bearer \")  [{ep}]",
                    "condition": "X-Checkout-Token hợp lệ; Authorization: Bearer \"\" (empty token)",
                    "steps":    f"POST {ep}\nHeaders: X-Checkout-Token: <valid>  |  Authorization: Bearer ",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi token không hợp lệ",
                },
                {
                    "priority": "High",
                    "title":    f"Authorization Bearer sai giá trị  [{ep}]",
                    "condition": "X-Checkout-Token hợp lệ; Authorization: Bearer wrong_token_xyz",
                    "steps":    f"POST {ep}\nHeaders: X-Checkout-Token: <valid>  |  Authorization: Bearer wrong_token_xyz",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi \"Không lấy được voucher\"\n[Rule: \"Nhập sai báo lỗi không lấy được voucher\"]",
                },
                {
                    "priority": "High",
                    "title":    f"Authorization Bearer đã hết hạn  [{ep}]",
                    "condition": "X-Checkout-Token hợp lệ; Authorization: Bearer <expired_token>",
                    "steps":    f"POST {ep}\nHeaders: X-Checkout-Token: <valid>  |  Authorization: Bearer <expired_token>",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi hết hạn, không xử lý voucher\n[Rule: \"Hết hạn báo lỗi, không hiển thị voucher\"]",
                },
                {
                    "priority": "High",
                    "title":    f"[SECURITY-INFERRED] X-Checkout-Token valid + Authorization Bearer invalid  [{ep}]",
                    "condition": "X-Checkout-Token hợp lệ; Authorization: Bearer invalid_token",
                    "steps":    f"POST {ep}\nHeaders: X-Checkout-Token: <valid>  |  Authorization: Bearer invalid_token",
                    "expected": "HTTP 200\nsuccess = false\nResponse từ chối vì Authorization sai — không trả data",
                    "security": True,
                },
                {
                    "priority": "High",
                    "title":    f"[SECURITY-INFERRED] Bearer user A + X-Checkout-Token checkout user B (mismatch)  [{ep}]",
                    "condition": "Authorization Bearer token của user A; X-Checkout-Token của checkout user B",
                    "steps":    f"1. Lấy Bearer token của user A\n2. Lấy X-Checkout-Token checkout user B\n3. POST {ep} với cặp không khớp",
                    "expected": "HTTP 200\nsuccess = false\nResponse từ chối (token ownership mismatch) — không trả data voucher của user B",
                    "security": True,
                },
            ],
        },
        {
            "name": "Authentication / Authorization  —  Client-Id (Header Tùy Chọn)",
            "tcs": [
                {
                    "priority": "Medium",
                    "title":    f"Client-Id không truyền — API vẫn hoạt động bình thường  [{ep}]",
                    "condition": "X-Checkout-Token hợp lệ; Authorization hợp lệ; Client-Id: không truyền",
                    "steps":    f"POST {ep}\nHeaders: X-Checkout-Token: <valid>  |  Authorization: Bearer <valid>  (không có Client-Id)",
                    "expected": "HTTP 200\nAPI xử lý và trả kết quả bình thường\nClient-Id là header KHÔNG bắt buộc",
                },
                {
                    "priority": "Low",
                    "title":    f"Client-Id có giá trị — không ảnh hưởng kết quả business  [{ep}]",
                    "condition": "X-Checkout-Token, Authorization hợp lệ; Client-Id = \"test_client_001\"",
                    "steps":    f"POST {ep} 2 lần: lần 1 có Client-Id, lần 2 không có\nSo sánh kết quả",
                    "expected": "HTTP 200\nKết quả business giống nhau ở 2 lần gọi\n[CLA-VOUCHER-005 pending: Client-Id có ảnh hưởng kết quả API không?]",
                },
            ],
        },
    ]


# ──────────────────────────────────────────────────────────────────────────────
# API_01  POST /public/v1/voucher/list
# ──────────────────────────────────────────────────────────────────────────────
API_01 = {
    "code":   "API_01",
    "name":   "Lấy danh sách EVC khả dụng",
    "method": "POST",
    "url":    "/public/v1/voucher/list",
    "sheet":  "Danh sách Voucher",
    "groups": auth_groups("/public/v1/voucher/list") + [
        {
            "name": "Business Flow — Happy Path",
            "tcs": [
                {
                    "priority": "High",
                    "title":    "Lấy danh sách EVC thành công — context có ≥1 EVC phù hợp",
                    "condition": "X-Checkout-Token: encode context checkout hợp lệ (gói + PTTT + địa chỉ có EVC phù hợp); Authorization hợp lệ",
                    "steps":    "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: <valid_with_context>  |  Authorization: Bearer <valid>\n(Không có request body)",
                    "expected": "HTTP 200\nsuccess = true\ndata[] không rỗng (≥1 item)\nMỗi item có: voucherCode (non-null, string), voucherType (integer, 1 hoặc 2)",
                },
                {
                    "priority": "High",
                    "title":    "Context không có EVC phù hợp → data = [] (mảng rỗng)",
                    "condition": "X-Checkout-Token: context checkout không có EVC nào khớp",
                    "steps":    "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: <valid_no_voucher>  |  Authorization: Bearer <valid>",
                    "expected": "HTTP 200\nsuccess = true\ndata = [] hoặc null\nKhông báo lỗi nghiệp vụ",
                },
                {
                    "priority": "High",
                    "title":    "Mỗi item data[] có đủ 2 fields bắt buộc: voucherCode và voucherType",
                    "condition": "X-Checkout-Token với context có ≥1 EVC",
                    "steps":    "POST /public/v1/voucher/list\nDuyệt từng item trong data[]",
                    "expected": "HTTP 200\nMọi item: voucherCode ≠ null, ≠ \"\"  |  voucherType ∈ {1, 2} (integer)",
                },
                {
                    "priority": "High",
                    "title":    "Mỗi item data[] có đủ 6 optional fields (key tồn tại, giá trị có thể null)",
                    "condition": "X-Checkout-Token hợp lệ với context có EVC",
                    "steps":    "POST /public/v1/voucher/list\nKiểm tra structure từng item trong data[]",
                    "expected": "HTTP 200\nMỗi item có key: description, note, expiredDate, applyTypeId, promotionTypeId, policyGroupId\n(Giá trị có thể null/rỗng — key phải tồn tại trong object)",
                },
                {
                    "priority": "High",
                    "title":    "expiredDate trả đúng format dd/MM/yyyy",
                    "condition": "Context có EVC sắp hết hạn (expiredDate có giá trị)",
                    "steps":    "POST /public/v1/voucher/list\nKiểm tra format expiredDate của item có ngày hết hạn",
                    "expected": "HTTP 200\nexpiredDate: string, đúng format \"dd/MM/yyyy\" (vd: \"31/12/2026\")\nKhông phải ISO 8601, không phải timestamp",
                },
                {
                    "priority": "High",
                    "title":    "voucherType boundary — chỉ nhận giá trị 1 hoặc 2",
                    "condition": "X-Checkout-Token với context có EVC đa dạng loại",
                    "steps":    "POST /public/v1/voucher/list\nCollect tất cả voucherType từ data[]",
                    "expected": "HTTP 200\nMọi voucherType ∈ {1, 2}  —  Không có 0, 3, hoặc null",
                },
            ],
        },
        {
            "name": "Business Flow — Context & Isolation",
            "tcs": [
                {
                    "priority": "High",
                    "title":    "Kết quả thay đổi khi context checkout thay đổi (gói khác nhau)",
                    "condition": "X-Checkout-Token-A (gói A) và X-Checkout-Token-B (gói B) — 2 gói có EVC khác nhau",
                    "steps":    "Call 1: X-Checkout-Token-A\nCall 2: X-Checkout-Token-B\nSo sánh data[]",
                    "expected": "HTTP 200 x2\ndata[] của 2 call khác nhau tương ứng context",
                },
                {
                    "priority": "High",
                    "title":    "Không lẫn EVC của kênh bán khác vào kết quả",
                    "condition": "X-Checkout-Token của kênh PDH",
                    "steps":    "POST /public/v1/voucher/list\nKiểm tra tất cả EVC trong data[]",
                    "expected": "HTTP 200\ndata[] chỉ chứa EVC áp dụng cho kênh PDH — không có EVC kênh khác",
                },
                {
                    "priority": "Medium",
                    "title":    "Gọi API 2 lần cùng X-Checkout-Token — kết quả nhất quán (idempotent)",
                    "condition": "Cùng X-Checkout-Token; không thay đổi context",
                    "steps":    "Gọi API lần 1 và lần 2 với cùng token\nSo sánh data[]",
                    "expected": "HTTP 200 x2\ndata[] giống nhau ở cả 2 lần — read-only, no side-effect",
                },
            ],
        },
        {
            "name": "Validation — Output Fields",
            "tcs": [
                {
                    "priority": "Medium",
                    "title":    "voucherCode trong data[] là unique (không trùng)",
                    "condition": "Context có ≥2 EVC",
                    "steps":    "POST /public/v1/voucher/list\nCollect tất cả voucherCode trong data[]",
                    "expected": "HTTP 200\nKhông có 2 item nào có cùng voucherCode",
                },
                {
                    "priority": "Medium",
                    "title":    "note field null/rỗng — response không crash",
                    "condition": "Context có EVC không có ghi chú",
                    "steps":    "POST /public/v1/voucher/list\nKiểm tra item.note",
                    "expected": "HTTP 200\nitem.note = null hoặc \"\" — không crash, JSON parse bình thường",
                },
                {
                    "priority": "Low",
                    "title":    "voucherCode và voucherType phải đọc được để dùng khi apply",
                    "condition": "X-Checkout-Token hợp lệ có data",
                    "steps":    "POST /public/v1/voucher/list\nXác nhận đọc được 2 fields bắt buộc",
                    "expected": "HTTP 200\nvoucherCode: parseable string  |  voucherType: integer 1 hoặc 2\n(Client phải lưu lại để truyền vào /apply — per BA spec)",
                },
            ],
        },
        {
            "name": "Security [SECURITY-INFERRED]",
            "tcs": [
                {
                    "priority": "High",
                    "title":    "[SECURITY-INFERRED] Không lộ EVC của checkout/user khác (cross-tenant isolation)",
                    "condition": "X-Checkout-Token của checkout A; biết trước EVC của checkout B (khác user)",
                    "steps":    "1. Gọi API với token checkout A\n2. Verify data[] không chứa EVC của checkout B",
                    "expected": "HTTP 200\ndata[] chỉ chứa EVC đúng context token — không lộ EVC user/checkout khác",
                    "security": True,
                },
                {
                    "priority": "Medium",
                    "title":    "[SECURITY-INFERRED] X-Checkout-Token chứa ký tự injection — API xử lý an toàn",
                    "condition": "X-Checkout-Token = \"<script>alert(1)</script>\" hoặc SQL injection payload",
                    "steps":    "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: <injection_payload>",
                    "expected": "HTTP 200\nResponse báo lỗi token không hợp lệ — không execute script, không SQL error",
                    "security": True,
                },
                {
                    "priority": "Low",
                    "title":    "[SECURITY-INFERRED] Header bổ sung không trong spec — API ignore graceful",
                    "condition": "Thêm header \"X-Debug-Mode: true\" hoặc \"X-Internal: bypass\"",
                    "steps":    "POST /public/v1/voucher/list\nThêm header không trong spec\nGọi API",
                    "expected": "HTTP 200\nAPI hoạt động bình thường, bỏ qua header không rõ nguồn gốc",
                    "security": True,
                },
            ],
        },
    ],
}


# ──────────────────────────────────────────────────────────────────────────────
# API_02  POST /public/v1/voucher/content
# Output: voucher_code + content1..content6  [BA updated 2026-05-27]
# ──────────────────────────────────────────────────────────────────────────────
API_02 = {
    "code":   "API_02",
    "name":   "Lấy nội dung hiển thị EVC",
    "method": "POST",
    "url":    "/public/v1/voucher/content",
    "sheet":  "Noi dung Voucher",
    "groups": auth_groups("/public/v1/voucher/content") + [
        {
            "name": "Validation — Required Fields (Request Body)",
            "tcs": [
                {
                    "priority": "High",
                    "title":    "Thiếu field voucher_code trong body",
                    "condition": "X-Checkout-Token, Authorization hợp lệ; body: {}",
                    "steps":    "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <valid>  |  Authorization: Bearer <valid>\nBody: {}",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi thiếu voucher_code bắt buộc\n[CLA-APISPEC-004: format error body TBD]",
                },
                {
                    "priority": "High",
                    "title":    "voucher_code = null",
                    "condition": "body: {\"voucher_code\": null}",
                    "steps":    "POST /public/v1/voucher/content\nBody: {\"voucher_code\": null}",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi voucher_code không hợp lệ",
                },
                {
                    "priority": "High",
                    "title":    "voucher_code = \"\" (chuỗi rỗng)",
                    "condition": "body: {\"voucher_code\": \"\"}",
                    "steps":    "POST /public/v1/voucher/content\nBody: {\"voucher_code\": \"\"}",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi voucher_code không được rỗng",
                },
                {
                    "priority": "High",
                    "title":    "voucher_code không tồn tại trong hệ thống",
                    "condition": "body: {\"voucher_code\": \"NONEXISTENT_XYZ999\"}",
                    "steps":    "POST /public/v1/voucher/content\nBody: {\"voucher_code\": \"NONEXISTENT_XYZ999\"}",
                    "expected": "HTTP 200\nsuccess = false (hoặc tương đương)\nResponse kèm thông báo lỗi tiếng Việt\nKhông lộ cấu trúc nội bộ\n[CLA-APISPEC-004: format TBD]",
                },
                {
                    "priority": "Medium",
                    "title":    "[SECURITY-INFERRED] voucher_code quá dài (>100 ký tự) — server không crash",
                    "condition": "body: {\"voucher_code\": \"A\" * 200}",
                    "steps":    "POST /public/v1/voucher/content\nBody: {\"voucher_code\": <200 ký tự A>}",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi input không hợp lệ — không crash, không HTTP 500",
                    "security": True,
                },
                {
                    "priority": "High",
                    "title":    "[SECURITY-INFERRED] voucher_code chứa XSS payload",
                    "condition": "body: {\"voucher_code\": \"<script>alert(document.cookie)</script>\"}",
                    "steps":    "POST /public/v1/voucher/content\nBody với XSS payload trong voucher_code\nKiểm tra response không reflect script",
                    "expected": "HTTP 200\nsuccess = false\nXSS payload không được execute/reflect trong response body",
                    "security": True,
                },
                {
                    "priority": "High",
                    "title":    "[SECURITY-INFERRED] voucher_code chứa SQL Injection attempt",
                    "condition": "body: {\"voucher_code\": \"'; DROP TABLE vouchers; --\"}",
                    "steps":    "POST /public/v1/voucher/content\nBody với SQL injection payload",
                    "expected": "HTTP 200\nsuccess = false\nKhông có SQL error, không crash DB",
                    "security": True,
                },
            ],
        },
        {
            "name": "Business Flow — Happy Path",
            "tcs": [
                {
                    "priority": "High",
                    "title":    "Lấy nội dung EVC thành công — voucher_code hợp lệ",
                    "condition": "voucher_code hợp lệ tồn tại trong hệ thống; X-Checkout-Token, Authorization hợp lệ",
                    "steps":    "POST /public/v1/voucher/content\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
                    "expected": "HTTP 200\nsuccess = true\ndata.voucher_code = \"CA21060100KTHIETBIKHOFG039\"\nCó ít nhất 1 content field không null",
                },
                {
                    "priority": "High",
                    "title":    "Response có đủ 7 output fields: voucher_code + content1-content6",
                    "condition": "voucher_code hợp lệ đầy đủ thông tin",
                    "steps":    "POST /public/v1/voucher/content\nKiểm tra structure response",
                    "expected": "HTTP 200\ndata có: voucher_code (Y, non-null) + content1, content2, content3, content4, content5, content6\n(content2-6 có thể null nếu voucher chưa cấu hình đủ nội dung)\n[CLA-VOUCHER-001 partial: tên field JSON inferred từ spec — confirm với dev]",
                },
                {
                    "priority": "High",
                    "title":    "voucher_code trong response khớp với input",
                    "condition": "voucher_code = \"CA21060100KTHIETBIKHOFG039\"",
                    "steps":    "POST /public/v1/voucher/content\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}\nSo sánh response.data.voucher_code với input",
                    "expected": "HTTP 200\ndata.voucher_code == \"CA21060100KTHIETBIKHOFG039\"",
                },
                {
                    "priority": "Medium",
                    "title":    "Voucher chưa cấu hình content — content1-6 đều null",
                    "condition": "voucher_code tồn tại nhưng chưa có content",
                    "steps":    "POST /public/v1/voucher/content\nBody: {\"voucher_code\": \"<voucher_no_content>\"}",
                    "expected": "HTTP 200\nsuccess = true\ndata.voucher_code khớp input\ndata.content1 = null ... data.content6 = null",
                },
                {
                    "priority": "Medium",
                    "title":    "Gọi /content nhiều lần cùng voucher_code — kết quả nhất quán (idempotent)",
                    "condition": "Cùng voucher_code; không có thay đổi dữ liệu",
                    "steps":    "Gọi API 2 lần với cùng voucher_code\nSo sánh response",
                    "expected": "HTTP 200 x2\nContent giống nhau ở cả 2 lần — read-only, idempotent",
                },
                {
                    "priority": "Medium",
                    "title":    "Gọi /content không cần gọi /list trước (on-demand API)",
                    "condition": "Có voucher_code hợp lệ — bỏ qua API_01",
                    "steps":    "Trực tiếp gọi POST /voucher/content với voucher_code đã biết (không call /list trước)",
                    "expected": "HTTP 200\nAPI trả content bình thường — không yêu cầu phải call /list trước",
                },
            ],
        },
        {
            "name": "Business Flow — Edge Cases",
            "tcs": [
                {
                    "priority": "High",
                    "title":    "voucher_code đã hết hạn → response lỗi nghiệp vụ",
                    "condition": "voucher_code của EVC đã expired",
                    "steps":    "POST /public/v1/voucher/content\nBody: {\"voucher_code\": \"<expired_code>\"}",
                    "expected": "HTTP 200\nsuccess = false\nThông báo voucher đã hết hạn",
                },
                {
                    "priority": "Medium",
                    "title":    "voucher_code của kênh bán khác — không lấy được content",
                    "condition": "voucher_code thuộc EVC kênh khác (không phải PDH)",
                    "steps":    "POST /public/v1/voucher/content\nBody: {\"voucher_code\": \"<other_channel_code>\"}",
                    "expected": "HTTP 200\nsuccess = false\nResponse lỗi (không trả content EVC kênh khác)",
                },
                {
                    "priority": "Medium",
                    "title":    "[SECURITY-INFERRED] IDOR — không lấy content EVC của user khác",
                    "condition": "voucher_code thuộc checkout user B; X-Checkout-Token của user A",
                    "steps":    "1. Chuẩn bị voucher_code của user B\n2. POST với X-Checkout-Token của user A",
                    "expected": "HTTP 200\nsuccess = false\nKhông lộ thông tin EVC của user khác",
                    "security": True,
                },
            ],
        },
    ],
}


# ──────────────────────────────────────────────────────────────────────────────
# API_03  POST /public/v1/voucher/apply
# ──────────────────────────────────────────────────────────────────────────────
API_03 = {
    "code":   "API_03",
    "name":   "Ap dung eVoucher vao checkout",
    "method": "POST",
    "url":    "/public/v1/voucher/apply",
    "sheet":  "Ap dung Voucher",
    "groups": auth_groups("/public/v1/voucher/apply") + [
        {
            "name": "Validation — Request Body",
            "tcs": [
                {
                    "priority": "High",
                    "title":    "Thiếu field vouchers trong body",
                    "condition": "X-Checkout-Token, Authorization hợp lệ; body: {}",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {}",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi thiếu field bắt buộc\n[CLA-APISPEC-004: format TBD]",
                },
                {
                    "priority": "High",
                    "title":    "vouchers = [] (array rỗng) khi checkout chưa có voucher",
                    "condition": "body: {\"vouchers\": []}; checkout chưa có voucher nào",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {\"vouchers\": []}",
                    "expected": "HTTP 200\n[CLA-VOUCHER-006 OPEN — Behavior TBD: HTTP 200 graceful OR lỗi]\nGhi nhận thực tế và verify với BA/Dev",
                },
                {
                    "priority": "High",
                    "title":    "Thiếu voucher_code trong một item của vouchers[]",
                    "condition": "body: {\"vouchers\": [{\"voucher_type\": \"General\"}]}",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {\"vouchers\": [{\"voucher_type\": \"General\"}]}",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi thiếu voucher_code trong item",
                },
                {
                    "priority": "High",
                    "title":    "Thiếu voucher_type trong một item",
                    "condition": "body: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}]}",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}]}",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi thiếu voucher_type\n[CLA-APISPEC-002 pending: \"General\" string mapping chưa rõ]",
                },
                {
                    "priority": "High",
                    "title":    "voucher_code = \"\" (rỗng) trong item",
                    "condition": "body: {\"vouchers\": [{\"voucher_code\": \"\", \"voucher_type\": \"General\"}]}",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {\"vouchers\": [{\"voucher_code\": \"\", \"voucher_type\": \"General\"}]}",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi voucher_code không được rỗng",
                },
                {
                    "priority": "Medium",
                    "title":    "voucher_type = \"\" (rỗng) trong item",
                    "condition": "body item với voucher_type = \"\"",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {\"vouchers\": [{\"voucher_code\": \"valid_code\", \"voucher_type\": \"\"}]}",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi voucher_type không hợp lệ\n[CLA-APISPEC-002 pending]",
                },
                {
                    "priority": "Medium",
                    "title":    "voucher_type = giá trị không hợp lệ (\"InvalidType\")",
                    "condition": "voucher_type = \"InvalidType\" (không phải giá trị spec cho phép)",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {\"vouchers\": [{\"voucher_code\": \"valid\", \"voucher_type\": \"InvalidType\"}]}",
                    "expected": "HTTP 200\nsuccess = false\nResponse báo lỗi voucher_type không hợp lệ\n[CLA-APISPEC-002: mapping string->integer chưa xác định]",
                },
                {
                    "priority": "Medium",
                    "title":    "[SECURITY-INFERRED] voucher_code chứa XSS/SQL injection trong array",
                    "condition": "vouchers[0].voucher_code = \"<script>alert(1)</script>\"",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {\"vouchers\": [{\"voucher_code\": \"<script>alert(1)</script>\", \"voucher_type\": \"General\"}]}",
                    "expected": "HTTP 200\nsuccess = false\nPayload không execute/reflect trong response",
                    "security": True,
                },
            ],
        },
        {
            "name": "Business Flow — Happy Path (QLCS result = 1)",
            "tcs": [
                {
                    "priority": "High",
                    "title":    "Apply 1 voucher thành công — QLCS result=1",
                    "condition": "voucher_code hợp lệ, voucher_type = \"General\"; context checkout phù hợp",
                    "steps":    "POST /public/v1/voucher/apply\nBody:\n{\n  \"vouchers\": [\n    {\n      \"voucher_code\": \"CA21060100KTHIETBIKHOFG039\",\n      \"voucher_type\": \"General\"\n    }\n  ]\n}",
                    "expected": "HTTP 200\nsuccess = true\nResponse có promotion_id (non-null)\nvoucher_code khớp input\ndiscount_value > 0",
                },
                {
                    "priority": "High",
                    "title":    "Response có đủ 17 top-level fields",
                    "condition": "Apply 1 voucher thành công (QLCS result=1)",
                    "steps":    "Apply voucher thành công\nKiểm tra toàn bộ top-level keys trong response",
                    "expected": "HTTP 200\nResponse có đủ 17 fields:\npromotion_id, promotion_title, voucher_code, referrer_code, discount_type,\ndiscount_value, discount_ex_vat_value, discount_rate, apply_type, apply_from, apply_to,\noriginal_discount_value, original_discount_ex_vat, voucher_type, voucher_type_l2, type_id, applies[]",
                },
                {
                    "priority": "High",
                    "title":    "applies[] có đủ 10 sub-fields cho mỗi service",
                    "condition": "Voucher áp dụng cho ≥1 service",
                    "steps":    "Apply voucher thành công\nDuyệt từng item trong applies[]\nKiểm tra sub-fields",
                    "expected": "HTTP 200\nMỗi applies[i] có đủ:\nservice_id, sub_service_type_id, sub_service_id, service_code,\ndiscount_ex_vat, discount, dismonth, is_deduct_order,\noriginal_discount_value, original_discount_ex_vat",
                },
                {
                    "priority": "High",
                    "title":    "Apply 2 vouchers cùng lúc — response đủ data",
                    "condition": "vouchers = [A, B]; cả 2 hợp lệ với context",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {\"vouchers\": [voucher_A, voucher_B]}",
                    "expected": "HTTP 200\nsuccess = true\ndiscount_value tổng hợp từ 2 vouchers",
                },
                {
                    "priority": "High",
                    "title":    "Apply thêm voucher B khi đã có A — CO gửi [A+B] lên QLCS",
                    "condition": "Checkout đã có voucher A; gọi /apply với [A, B]",
                    "steps":    "1. Đảm bảo checkout có voucher A\n2. POST /apply với vouchers=[A, B]\n3. Verify CO gửi đủ [A, B] lên QLCS (không chỉ gửi B)",
                    "expected": "HTTP 200\nCO gửi toàn bộ [A, B] đến QLCS\nResponse có data cho cả 2 vouchers",
                },
            ],
        },
        {
            "name": "Business Flow — Failure Cases (QLCS result != 1)",
            "tcs": [
                {
                    "priority": "High",
                    "title":    "QLCS result != 1 → apply thất bại, checkout không thay đổi",
                    "condition": "voucher_code không phù hợp context hiện tại (QLCS reject)",
                    "steps":    "POST /public/v1/voucher/apply với voucher không phù hợp\nKiểm tra response và checkout state",
                    "expected": "HTTP 200\nsuccess = false\nErrorMessage tiếng Việt\nCheckout price không thay đổi sau call",
                },
                {
                    "priority": "High",
                    "title":    "Apply thêm thất bại → voucher A vẫn giữ nguyên",
                    "condition": "Checkout có voucher A; apply thêm B thất bại (QLCS reject B)",
                    "steps":    "1. Checkout đang có voucher A\n2. POST /apply với [A, B] — voucher B không hợp lệ\n3. Verify checkout state sau fail",
                    "expected": "HTTP 200\nsuccess = false\nCheckout vẫn có voucher A với discount gốc — voucher A không bị xóa",
                },
                {
                    "priority": "High",
                    "title":    "voucher_code không tồn tại → QLCS reject → apply thất bại",
                    "condition": "vouchers = [{voucher_code: \"NONEXISTENT\", voucher_type: \"General\"}]",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {\"vouchers\": [{\"voucher_code\": \"NONEXISTENT\", \"voucher_type\": \"General\"}]}",
                    "expected": "HTTP 200\nsuccess = false\nThông báo lỗi tiếng Việt\nCheckout không thay đổi",
                },
                {
                    "priority": "High",
                    "title":    "Voucher hết hạn → QLCS reject → apply thất bại",
                    "condition": "voucher_code của EVC đã expired",
                    "steps":    "POST /public/v1/voucher/apply\nBody: {\"vouchers\": [{\"voucher_code\": \"<expired>\", \"voucher_type\": \"General\"}]}",
                    "expected": "HTTP 200\nsuccess = false\nThông báo voucher hết hạn",
                },
            ],
        },
        {
            "name": "Validation — Output Fields Boundary",
            "tcs": [
                {
                    "priority": "Medium",
                    "title":    "discount_value >= 0 (không âm)",
                    "condition": "Apply voucher thành công",
                    "steps":    "Apply voucher\nKiểm tra discount_value trong response",
                    "expected": "HTTP 200\ndiscount_value: number >= 0",
                },
                {
                    "priority": "Medium",
                    "title":    "dismonth = 0 → áp dụng 1 lần (semantic boundary)",
                    "condition": "Voucher có applies[].dismonth = 0",
                    "steps":    "Apply voucher có service với dismonth=0\nKiểm tra applies[].dismonth",
                    "expected": "HTTP 200\napplies[].dismonth = 0 có nghĩa áp dụng 1 lần (không phải 0 lần)",
                },
                {
                    "priority": "Medium",
                    "title":    "is_deduct_order = 1 → discount khấu trừ thẳng vào tổng đơn",
                    "condition": "Voucher có applies[].is_deduct_order = 1",
                    "steps":    "Apply voucher có is_deduct_order=1\nVerify logic tính discount",
                    "expected": "HTTP 200\napplies[].is_deduct_order = 1 → discount trừ vào tổng đơn hàng",
                },
                {
                    "priority": "Medium",
                    "title":    "original_discount_value >= discount_value (logic consistency)",
                    "condition": "Apply voucher có original và actual discount",
                    "steps":    "Apply voucher\nSo sánh original_discount_value với discount_value",
                    "expected": "HTTP 200\noriginal_discount_value >= discount_value",
                },
                {
                    "priority": "Medium",
                    "title":    "applies[] = [] khi không có service nào được chiết khấu",
                    "condition": "Voucher apply thành công nhưng không áp cho service cụ thể",
                    "steps":    "Apply voucher\nKiểm tra applies[] trong response",
                    "expected": "HTTP 200\napplies[] = [] hoặc null — không crash response",
                },
                {
                    "priority": "Low",
                    "title":    "voucher_code trong response khớp với input",
                    "condition": "Apply voucher_code = \"CA21060100KTHIETBIKHOFG039\"",
                    "steps":    "Apply voucher\nKiểm tra response.voucher_code",
                    "expected": "HTTP 200\nresponse.voucher_code == \"CA21060100KTHIETBIKHOFG039\"",
                },
            ],
        },
        {
            "name": "Security [SECURITY-INFERRED]",
            "tcs": [
                {
                    "priority": "High",
                    "title":    "[SECURITY-INFERRED] IDOR — không apply được voucher của checkout/user khác",
                    "condition": "voucher_code của checkout user B; X-Checkout-Token của user A",
                    "steps":    "1. Lấy voucher_code checkout user B\n2. POST /apply với X-Checkout-Token của user A",
                    "expected": "HTTP 200\nsuccess = false\nKhông áp EVC của user khác vào checkout hiện tại",
                    "security": True,
                },
                {
                    "priority": "High",
                    "title":    "[SECURITY-INFERRED] Double-apply cùng voucher — không double-discount",
                    "condition": "Gọi /apply 2 lần liên tiếp với cùng voucher_code",
                    "steps":    "1. Call /apply lần 1 với voucher A (thành công)\n2. Call /apply lần 2 với cùng voucher A",
                    "expected": "HTTP 200\nLần 2: checkout không bị double-discount — idempotent hoặc báo \"đã apply\"",
                    "security": True,
                },
                {
                    "priority": "High",
                    "title":    "[SECURITY-INFERRED] Cross-checkout: voucher của checkout A apply vào checkout B",
                    "condition": "X-Checkout-Token của checkout B; voucher_code thuộc checkout A",
                    "steps":    "1. Lấy voucher_code từ checkout A\n2. Apply vào checkout B (khác X-Checkout-Token)",
                    "expected": "HTTP 200\nsuccess = false\nVoucher không được áp cross-checkout",
                    "security": True,
                },
                {
                    "priority": "Medium",
                    "title":    "[SECURITY-INFERRED] Gửi array vouchers quá lớn (100+ items) — không DoS",
                    "condition": "vouchers = array 100 items",
                    "steps":    "Tạo array 100 voucher items\nPOST /public/v1/voucher/apply",
                    "expected": "HTTP 200\nServer trả lỗi \"giới hạn số lượng\" hoặc xử lý được — không timeout, không crash",
                    "security": True,
                },
            ],
        },
    ],
}


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    totals = {}
    for api in [API_01, API_02, API_03]:
        ws = wb.create_sheet(title=api["sheet"])
        n = build_sheet(wb, api)
        totals[api["code"]] = n
        print(f"  [OK]  {api['code']}  {api['sheet']:25s}  {n} TCs")

    wb.save(OUT_FILE)
    grand = sum(totals.values())
    sec   = sum(
        1 for api in [API_01, API_02, API_03]
        for grp in api["groups"]
        for tc  in grp["tcs"]
        if "[SECURITY-INFERRED]" in tc.get("title","") or tc.get("security",False)
    )
    print(f"\n  Saved  : {OUT_FILE}")
    print(f"  Total  : {grand} TCs  ({sec} Security-Inferred)")


if __name__ == "__main__":
    main()
