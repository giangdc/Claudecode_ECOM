# -*- coding: utf-8 -*-
"""
Sinh TC Web cho module CHECKOUT (chucnang_checkout) — 2 sheet:
  - Checkout_Common  (TC_CKCOMMON)  : màn checkout chung (dịch vụ có Địa chỉ lắp đặt)
  - Checkout_Internet (TC_INTERNET) : đăng ký Internet 3 bước
Nguồn: 02_analyze-requirements/chucnang_checkout/{test_scenario_map.md, MEMORY.md}
Theo template: .claude/template/template-testcase-web_mobile.md
Chung cư (SC-CKCOMMON-037..040) DEFERRED -> không gen. Popup content pending -> BLOCKED.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

OUT = "ecom-pdh/03_test-cases/functional/chucnang_checkout/AI_ISC_ecom-pdh_v1.1_TC_checkout_v1.0.xlsx"

# ---- styles ----
HDR_FILL = PatternFill("solid", fgColor="4472C4")
GRP_FILL = PatternFill("solid", fgColor="A9D08E")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
GRP_FONT = Font(bold=True, size=11, name="Calibri")
AI_FONT  = Font(bold=True, color="7030A0", size=11, name="Calibri")
NORMAL   = Font(size=11, name="Calibri")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP_TL = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_TC = Alignment(wrap_text=True, vertical="top", horizontal="center")
CTR = Alignment(wrap_text=True, vertical="center", horizontal="center")
GRP_AL = Alignment(wrap_text=True, vertical="center", horizontal="left")

COLW = dict(A=6, B=14, C=10, D=45, E=45, F=50, G=58, H=8, I=12, J=15, K=12, L=20)
HEADERS = ["QC/AI","Testcase ID","Mức Độ Ưu Tiên","Nội Dung Test",
           "Điều Kiện / Dữ Liệu Test","Các Bước Thực Hiện","Kết Quả Mong Đợi",
           "Có thể\nTự Động Hóa"]
ROUND_SUB = ["Kết Quả Thực Hiện","Người Thực Hiện","ID Bugs","Ghi Chú"]

def est_height(row):
    # row = (pri,title,pre,steps,exp,auto)
    mx = 1
    for txt in (row[1], row[2], row[3], row[4]):
        for line in str(txt).split("\n"):
            mx += max(1, (len(line)//48) + (1 if line else 0))
    return min(max(30, mx*15), 430)

def build_sheet(wb, sheet_name, func_id, func_name, groups, first=False):
    ws = wb.create_sheet(title=sheet_name)
    for c,w in COLW.items():
        ws.column_dimensions[c].width = w
    # function header
    ws["C3"] = "Mã chức năng:"; ws["C3"].font = GRP_FONT
    ws["D3"] = func_id;         ws["D3"].font = GRP_FONT
    ws["C4"] = "Tên chức năng:"; ws["C4"].font = GRP_FONT
    ws["D4"] = func_name;        ws["D4"].font = GRP_FONT
    # table header rows 7-8
    for i,h in enumerate(HEADERS, start=1):
        cell = ws.cell(7, i, h); cell.fill=HDR_FILL; cell.font=HDR_FONT; cell.alignment=CTR; cell.border=BORDER
        ws.merge_cells(start_row=7,start_column=i,end_row=8,end_column=i)
        ws.cell(8,i).fill=HDR_FILL; ws.cell(8,i).border=BORDER
    ws.cell(7,9,"Round 1").fill=HDR_FILL; ws.cell(7,9).font=HDR_FONT; ws.cell(7,9).alignment=CTR
    ws.merge_cells("I7:L7")
    for j,sub in enumerate(ROUND_SUB, start=9):
        c=ws.cell(8,j,sub); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CTR; c.border=BORDER
    for col in range(9,13):
        ws.cell(7,col).fill=HDR_FILL; ws.cell(7,col).border=BORDER
    ws.row_dimensions[7].height=30; ws.row_dimensions[8].height=20
    ws.freeze_panes="A9"

    r = 9
    for gname, tcs in groups:
        # group header
        ws.cell(r,2,gname); ws.cell(r,2).fill=GRP_FILL; ws.cell(r,2).font=GRP_FONT; ws.cell(r,2).alignment=GRP_AL
        ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=12)
        for col in range(2,13):
            ws.cell(r,col).fill=GRP_FILL; ws.cell(r,col).border=BORDER
        ws.row_dimensions[r].height=18
        r += 1
        for pri,title,pre,steps,exp,auto in tcs:
            ws.cell(r,1,"AI").font=AI_FONT; ws.cell(r,1).alignment=WRAP_TC
            # TC ID formula (template 2.1)
            ws.cell(r,2,'=IF(D%d="","",$D$3&"."&COUNTA($D$10:D%d)&"")'%(r,r)).font=NORMAL
            ws.cell(r,2).alignment=WRAP_TC
            ws.cell(r,3,pri).font=NORMAL;  ws.cell(r,3).alignment=WRAP_TC
            ws.cell(r,4,title).font=NORMAL; ws.cell(r,4).alignment=WRAP_TL
            ws.cell(r,5,pre).font=NORMAL;   ws.cell(r,5).alignment=WRAP_TL
            ws.cell(r,6,steps).font=NORMAL; ws.cell(r,6).alignment=WRAP_TL
            ws.cell(r,7,exp).font=NORMAL;   ws.cell(r,7).alignment=WRAP_TL
            ws.cell(r,8,auto).font=NORMAL;  ws.cell(r,8).alignment=WRAP_TC
            for col in range(1,13):
                ws.cell(r,col).border=BORDER
            ws.row_dimensions[r].height=est_height((pri,title,pre,steps,exp,auto))
            r += 1
    return ws

# ====================== DATA: Checkout_Common ======================
PRE_CK = "- Đã chọn 1 gói dịch vụ (có Địa chỉ lắp đặt) ở màn Chi tiết\n- Đang ở màn hình Checkout"
COMMON = [
 ("Nhóm 1: Hiển thị tổng thể & Điều hướng", [
  ("High","Kiểm tra hiển thị tổng thể màn hình Checkout (dịch vụ có Địa chỉ lắp đặt)",
   PRE_CK,
   "1. Quan sát toàn bộ màn hình Checkout",
   "Hiển thị đầy đủ các thành phần:\n- Tiến trình các bước\n- Block Sản phẩm dịch vụ đã chọn\n- Block Thông tin cá nhân\n- Block Địa chỉ lắp đặt\n- Block Phương thức thanh toán\n- Block Thông tin khách hàng\n- Block Thông tin thanh toán\n- Button Thanh toán, Button Quay lại\n- Text điều khoản FPT Telecom","N"),
  ("Medium","Kiểm tra click Logo FPT điều hướng về Home",
   "- Đang ở màn hình Checkout",
   "1. Click Logo FPT Telecom ở header",
   "- Vào trực tiếp link: điều hướng Home fpt.vn\n- Vào từ tongdaiwifi: điều hướng về https://staging.tongdaiwifi.vn/","Y"),
  ("Medium","Kiểm tra click icon back khi mở trực tiếp link đăng ký",
   "- Mở trực tiếp link đăng ký dịch vụ\n- Đang ở màn hình Checkout",
   "1. Click icon back trên màn hình",
   "Điều hướng vào Home fpt.vn","Y"),
  ("Medium","Kiểm tra click icon back khi vào từ tongdaiwifi.vn",
   "- Vào từ tongdaiwifi.vn rồi click Đăng ký ngay\n- Đang ở màn hình Checkout",
   "1. Click icon back trên màn hình",
   "Điều hướng về https://staging.tongdaiwifi.vn/","Y"),
  ("Low","Kiểm tra màu sắc mặc định tiến trình các bước và click vào bước khi đang ở Bước 1",
   "- Đang ở Bước 1 màn hình nhập thông tin",
   "1. Quan sát màu sắc các bước trên tiến trình\n2. Click lần lượt icon Bước 1, 2, 3",
   "- Bước 1 màu xanh dương; Bước 2, 3 màu xám\n- Click các bước không có action, vẫn ở màn hình hiện tại","N"),
  ("Medium","Kiểm tra click Bước 1 khi đang ở Bước 2 (luồng nhiều bước)",
   "- Luồng checkout nhiều bước, đang ở Bước 2",
   "1. Click icon Bước 1 trên tiến trình",
   "Cho phép quay về Bước 1","Y"),
 ]),
 ("Nhóm 2: Block Sản phẩm dịch vụ đã chọn", [
  ("High","Kiểm tra Block Sản phẩm dịch vụ đã chọn load đúng thông tin",
   "- Đã chọn gói dịch vụ cụ thể (tên, chu kỳ, số lượng, giá) ở màn Chi tiết",
   "1. Quan sát Block Sản phẩm dịch vụ đã chọn",
   "Hiển thị đúng: số lượng, tên gói, chu kỳ (VD 3/6/12 tháng), giá (sau ưu đãi nếu có), icon theo config Product Hub","Y"),
 ]),
 ("Nhóm 3: Block Thông tin cá nhân — Họ tên", [
  ("High","Kiểm tra nhập Họ tên hợp lệ (có khoảng trắng đầu/cuối)",
   PRE_CK,
   "1. Nhập Họ tên hợp lệ chỉ gồm chữ + khoảng trắng, có khoảng trắng đầu/cuối\n2. Nhập các trường bắt buộc khác\n3. Submit",
   "Chấp nhận, không báo lỗi; hệ thống tự trim khoảng trắng đầu/cuối khi submit","Y"),
  ("High","Kiểm tra Họ tên để trống hoặc chỉ khoảng trắng",
   PRE_CK,
   "1. Để trống (hoặc nhập khoảng trắng) trường Họ tên\n2. Click ra ngoài trường",
   'Border đỏ trường Họ tên và hiển thị "Vui lòng nhập họ tên."',"Y"),
  ("Medium","Kiểm tra Họ tên chứa số hoặc ký tự đặc biệt",
   PRE_CK,
   "1. Nhập/paste Họ tên có chứa số hoặc ký tự đặc biệt\n2. Click ra ngoài trường",
   'Hiển thị "Họ tên không hợp lệ." (chỉ cho phép chữ + khoảng trắng)',"Y"),
  ("Medium","Kiểm tra Họ tên nhập/paste vượt quá 100 ký tự",
   PRE_CK,
   "1. Nhập/paste Họ tên dài hơn 100 ký tự",
   "Chỉ cho nhập và hiển thị tối đa 100 ký tự","Y"),
  ("Low","Kiểm tra icon X hiển thị và xóa dữ liệu trường Họ tên",
   PRE_CK,
   "1. Nhập ký tự bất kỳ vào Họ tên\n2. Click icon X cuối textbox",
   "Icon X hiển thị khi có dữ liệu; click X xóa toàn bộ và ẩn icon X","Y"),
 ]),
 ("Nhóm 4: Block Thông tin cá nhân — Số điện thoại", [
  ("High","Kiểm tra nhập Số điện thoại hợp lệ",
   PRE_CK,
   "1. Nhập SĐT 10 số bắt đầu bằng 0 (VD 0901234567)",
   "Chấp nhận, không báo lỗi; icon X hiển thị cuối textbox","Y"),
  ("High","Kiểm tra Số điện thoại để trống",
   PRE_CK,
   "1. Để trống trường Số điện thoại\n2. Click ra ngoài (hoặc submit)",
   'Border đỏ và hiển thị "Vui lòng nhập số điện thoại."',"Y"),
  ("Medium","Kiểm tra Số điện thoại chứa ký tự không phải số",
   PRE_CK,
   "1. Nhập/paste SĐT có ký tự không phải số",
   'Hiển thị "Số điện thoại không đúng."',"Y"),
  ("Medium","Kiểm tra Số điện thoại 10 số nhưng không bắt đầu bằng 0",
   PRE_CK,
   "1. Nhập SĐT 10 số bắt đầu khác 0 (VD 1901234567)",
   'Hiển thị "Số điện thoại không đúng."',"Y"),
  ("Medium","Kiểm tra Số điện thoại nhập/paste nhiều hơn 10 số",
   PRE_CK,
   "1. Nhập/paste SĐT nhiều hơn 10 chữ số",
   "Chỉ nhận 10 số đầu, tự cắt bỏ từ số thứ 11","Y"),
  ("Low","Kiểm tra icon X xóa dữ liệu trường Số điện thoại",
   PRE_CK,
   "1. Nhập ký tự vào SĐT\n2. Click icon X",
   "Xóa toàn bộ dữ liệu SĐT và ẩn icon X","Y"),
 ]),
 ("Nhóm 5: Block Thông tin cá nhân — Email (chỉ Hyperfast/UltraFast)", [
  ("Medium","Kiểm tra trường Email placeholder và cho phép để trống",
   "- Đang Checkout dịch vụ Hyperfast/UltraFast",
   "1. Quan sát placeholder trường Email\n2. Để trống Email rồi click ra ngoài",
   'Placeholder "Nhập email"; cho phép để trống, không báo lỗi (Email không bắt buộc)',"Y"),
  ("Medium","Kiểm tra nhập Email sai định dạng",
   "- Đang Checkout dịch vụ Hyperfast/UltraFast",
   "1. Nhập Email sai định dạng (VD: email_sai_format)\n2. Click ra ngoài",
   'Hiển thị "Email không hợp lệ"',"Y"),
  ("Medium","Kiểm tra nhập Email không có domain",
   "- Đang Checkout dịch vụ Hyperfast/UltraFast",
   "1. Nhập Email không có domain (VD: test@)\n2. Click ra ngoài",
   'Hiển thị "Email không hợp lệ"',"Y"),
  ("Medium","Kiểm tra Email hiển thị realtime sang block Thông tin khách hàng và icon X",
   "- Đang Checkout dịch vụ Hyperfast/UltraFast",
   "1. Nhập Email hợp lệ (VD: Acb@gmail.com)\n2. Quan sát block Thông tin khách hàng\n3. Click icon X",
   "Email hiển thị realtime sang block Thông tin khách hàng; icon X xóa được Email","Y"),
 ]),
 ("Nhóm 6: Block Địa chỉ lắp đặt — Tỉnh/Thành phố & pre-fill", [
  ("Low","Kiểm tra Tỉnh/Thành phố mặc định rỗng và placeholder",
   PRE_CK,
   "1. Quan sát trường Tỉnh/Thành phố khi chưa chọn",
   'Mặc định rỗng; placeholder "Chọn tỉnh/thành phố"',"Y"),
  ("Medium","Kiểm tra không chọn Tỉnh/Thành phố",
   PRE_CK,
   "1. Click trường Tỉnh/Thành phố\n2. Không chọn và click ra ngoài",
   'Border đỏ và hiển thị "Vui lòng chọn tỉnh/thành phố."',"Y"),
  ("Medium","Kiểm tra load danh sách Tỉnh/Thành phố theo ĐCHC mới",
   PRE_CK,
   "1. Click dropdown Tỉnh/Thành phố",
   "Hiển thị đủ tỉnh/thành toàn quốc theo ĐCHC mới; HCM, HN, Đà Nẵng load đầu danh sách","Y"),
  ("Medium","Kiểm tra tìm kiếm Tỉnh/Thành phố",
   PRE_CK,
   "1. Click dropdown Tỉnh/Thành phố\n2. Nhập từ khóa cần tìm",
   "Hiển thị danh sách khớp từ khóa (quy tắc contains)","Y"),
  ("Medium","Kiểm tra chọn Tỉnh/Thành phố load thêm trường địa chỉ",
   PRE_CK,
   "1. Chọn 1 Tỉnh/Thành phố bất kỳ",
   "Chỉ chọn được 1; load thêm Phường/Xã, Tên đường, Số nhà, Ghi chú","Y"),
  ("Medium","Kiểm tra đổi Tỉnh/Thành phố reload các trường phụ thuộc",
   "- Đã chọn 1 Tỉnh/Thành phố",
   "1. Chọn 1 Tỉnh/Thành phố khác",
   "Reload Phường/Xã tương ứng, reset Tên đường, trigger kiểm tra chính sách giá","Y"),
  ("Medium",'Kiểm tra pre-fill "Địa chỉ trước sáp nhập" điền toàn bộ địa chỉ',
   '- Account đã có địa chỉ lưu sẵn từ đơn trước\n- Dịch vụ có link "Địa chỉ trước sáp nhập"',
   '1. Click link "Địa chỉ trước sáp nhập"',
   "Hệ thống tự điền TOÀN BỘ địa chỉ đã lưu (Tỉnh/TP, Phường/Xã, Tên đường, Số nhà...)","Y"),
 ]),
 ("Nhóm 7: Block Địa chỉ lắp đặt — Phường/Xã & kiểm tra chính sách giá", [
  ("Medium","Kiểm tra Phường/Xã placeholder và bắt buộc chọn",
   "- Đã chọn Tỉnh/Thành phố",
   "1. Quan sát placeholder Phường/Xã\n2. Không chọn và click ra ngoài",
   'Placeholder "Chọn phường/xã"; border đỏ + "Vui lòng chọn phường/xã."',"Y"),
  ("Medium","Kiểm tra tìm kiếm và chọn Phường/Xã",
   "- Đã chọn Tỉnh/Thành phố",
   "1. Tìm Phường/Xã theo tên\n2. Chọn 1 Phường/Xã",
   "Hiển thị danh sách khớp; chỉ chọn được 1 và hiển thị đúng giá trị","Y"),
  ("High","Kiểm tra đổi Phường/Xã gọi API chính sách kiểm tra lại giá",
   "- Đã chọn Tỉnh/Thành phố và Phường/Xã",
   "1. Đổi sang Phường/Xã khác\n2. Quan sát Tạm tính/Giá",
   "Reset Tên đường; gọi API chính sách: giá không đổi → giữ Tạm tính; giá đổi → cập nhật lại Tạm tính","Y"),
  ("High",'Kiểm tra chọn địa chỉ không có chính sách hiển thị popup "Chưa hỗ trợ chính sách!"',
   "- Block Địa chỉ lắp đặt đang hiển thị",
   "1. Chọn Tỉnh/Phường-Xã không có chính sách\n2. Quan sát popup\n3. Click btn Đóng",
   '[BLOCKED – cần confirm nội dung đầy đủ popup (CLA-CKCOMMON-007 BA bổ sung sau)]. Theo rule (Rule common R5): hiển thị popup "Chưa hỗ trợ chính sách!" → đồng thời đẩy KHTN → tắt thông báo và quay về homepage (KHÔNG cho qua bước thanh toán)',"Y"),
  ("Medium","Kiểm tra đẩy KHTN khi chọn địa chỉ không có chính sách",
   "- Đã chọn địa chỉ không có chính sách\n- Có thể đẩy KHTN thành công",
   "1. Đóng popup Chưa hỗ trợ chính sách\n2. Truy cập web-admin (saleplatform-stag) kiểm tra KHTN theo SĐT đã nhập",
   "Đẩy KHTN thành công tương ứng SĐT đã nhập","N"),
 ]),
 ("Nhóm 8: Block Địa chỉ lắp đặt — Tên đường", [
  ("Medium","Kiểm tra Tên đường placeholder và bắt buộc chọn",
   "- Đã chọn Phường/Xã",
   "1. Quan sát placeholder Tên đường\n2. Không chọn và click ra ngoài",
   'Placeholder "Chọn tên đường"; border đỏ + "Vui lòng chọn tên đường."',"Y"),
  ("Medium","Kiểm tra tìm kiếm và chọn Tên đường",
   "- Đã chọn Phường/Xã",
   "1. Nhập từ khóa tìm Tên đường\n2. Chọn 1 Tên đường",
   "Tìm theo quy tắc contains, không trim khoảng trắng trước khi tìm; chỉ chọn được 1","Y"),
 ]),
 ("Nhóm 9: Block Địa chỉ lắp đặt — Số nhà (radio Nhà riêng/Chung cư: đặc thù dịch vụ thiết bị; Chung cư DEFERRED)", [
  ("Medium","Kiểm tra radio Nhà riêng/Chung cư (đặc thù dịch vụ thiết bị)",
   "- Block Địa chỉ lắp đặt đang hiển thị\n- Dịch vụ thiết bị (VD Camera) — Rule common không định nghĩa radio này",
   '1. Quan sát Block Địa chỉ lắp đặt',
   'Rule common KHÔNG có radio Nhà riêng/Chung cư. Internet: trường Số nhà* hiển thị trực tiếp sau khi chọn đủ địa chỉ. Dịch vụ thiết bị (Camera) có radio: chọn Nhà riêng → hiển thị Số nhà* (placeholder "Nhập Số nhà"). [Chung cư DEFERRED — CLA-CKCOMMON-006]',"N"),
  ("Medium","Kiểm tra Số nhà để trống → báo lỗi bắt buộc",
   "- Đã chọn đủ Tỉnh/Phường-Xã/Tên đường (trường Số nhà* đang hiển thị)",
   "1. Để trống Số nhà và click ra ngoài / submit",
   'Border đỏ và hiển thị "Vui lòng nhập số nhà."',"Y"),
  ("Medium","Kiểm tra Số nhà nhập/paste vượt quá 50 ký tự",
   "- Trường Số nhà* đang hiển thị",
   "1. Nhập/paste Số nhà dài hơn 50 ký tự",
   "Chỉ cho nhập và hiển thị tối đa 50 ký tự","Y"),
 ]),
 ("Nhóm 10: Block Địa chỉ lắp đặt — Ghi chú", [
  ("Low","Kiểm tra Ghi chú placeholder, không bắt buộc và giới hạn 100 ký tự",
   "- Block Địa chỉ lắp đặt đang hiển thị",
   "1. Quan sát placeholder Ghi chú\n2. Để trống rồi click ra ngoài\n3. Nhập/paste > 100 ký tự",
   'Placeholder "Gọi cho tôi trước 30 phút nhé!"; để trống không báo lỗi; nhập > 100 → chỉ nhận 100 ký tự',"Y"),
 ]),
 ("Nhóm 11: Popup Địa chỉ hành chính cũ (3 cấp → 2 cấp)", [
  ("Medium","Kiểm tra hiển thị UI popup Địa chỉ hành chính cũ",
   "- Block Địa chỉ lắp đặt đang hiển thị",
   '1. Click link "Địa chỉ trước sáp nhập"',
   'Popup hiển thị: title "Địa chỉ hành chính cũ", label hướng dẫn, 4 dropdown (Tỉnh/TP, Quận/Huyện, Phường/Xã, Tên đường), icon X, button Xác nhận',"Y"),
  ("Medium","Kiểm tra load phân cấp và tìm kiếm trong popup Địa chỉ hành chính cũ",
   "- Popup Địa chỉ hành chính cũ đang mở",
   "1. Chọn Tỉnh/TP → Quận/Huyện → Phường/Xã → Tên đường\n2. Tìm kiếm ở từng cấp",
   "63 tỉnh theo ĐCHC 3 cấp; mỗi cấp load theo cấp trên; tìm kiếm contains, không trim","Y"),
  ("Medium","Kiểm tra button Xác nhận disable khi chưa chọn đủ 4 cấp",
   "- Popup Địa chỉ hành chính cũ đang mở",
   "1. Để trống bất kỳ 1 trong 4 cấp\n2. Quan sát button Xác nhận",
   "Button Xác nhận ở trạng thái disable","Y"),
  ("High","Kiểm tra chọn đủ 4 cấp hiển thị Địa chỉ hành chính mới và enable Xác nhận",
   "- Popup Địa chỉ hành chính cũ đang mở",
   "1. Chọn đủ Tỉnh/TP + Quận/Huyện + Phường/Xã + Tên đường",
   'Hiển thị field "Địa chỉ hành chính mới" (convert 3 cấp → 2 cấp); enable button Xác nhận (disable nếu convert thất bại)',"Y"),
  ("Medium","Kiểm tra click icon X đóng popup không cập nhật địa chỉ",
   "- Popup có dữ liệu đã chọn",
   "1. Click icon X trên popup",
   "Không cập nhật địa chỉ mới; đóng popup, về form đăng ký","Y"),
  ("High","Kiểm tra click Xác nhận đẩy địa chỉ 2 cấp vào form",
   "- Popup đã chọn đủ 4 cấp, button Xác nhận enable",
   "1. Click button Xác nhận",
   "Đẩy Tỉnh/TP, Phường/Xã, Tên đường (2 cấp) vào dropdown form đăng ký; đóng popup","Y"),
 ]),
 ("Nhóm 12: Block Thông tin khách hàng", [
  ("Low","Kiểm tra thu gọn/mở rộng block Thông tin khách hàng",
   "- Đang ở màn hình Checkout",
   "1. Click icon Collapse\n2. Click icon Expand",
   "Thu gọn/mở rộng các trường; icon đổi collapse ↔ expand","Y"),
  ("Medium","Kiểm tra format hiển thị địa chỉ Nhà riêng trong block Thông tin khách hàng",
   "- Đã nhập đủ địa chỉ loại Nhà riêng",
   "1. Quan sát địa chỉ trong block Thông tin khách hàng",
   "Hiển thị theo format: Số nhà, Tên đường, Phường/Xã, Tỉnh/Thành","Y"),
  ("Medium","Kiểm tra không cho chỉnh sửa thông tin trong block Thông tin khách hàng",
   "- Đang ở màn hình Checkout",
   "1. Cố chỉnh sửa Họ tên/SĐT/Email/Địa chỉ trong block Thông tin khách hàng",
   "Không thể chỉnh sửa (read-only)","Y"),
 ]),
 ("Nhóm 13: Block Phương thức thanh toán", [
  ("High","Kiểm tra hiển thị UI block Phương thức thanh toán",
   "- Dịch vụ có ít nhất 1 PTTT trong chính sách",
   "1. Quan sát block Phương thức thanh toán",
   "Hiển thị: radio chọn, icon PTTT, tên PTTT, miêu tả, số lượng ưu đãi, button Xem thêm/Rút gọn (tùy số PTTT)","Y"),
  ("High","Kiểm tra hiển thị khi chính sách có ≤ 4 PTTT",
   "- Chính sách có dưới 4 PTTT",
   "1. Quan sát block Phương thức thanh toán",
   "Hiển thị tất cả PTTT, không có button Xem thêm","Y"),
  ("High","Kiểm tra hiển thị và Xem thêm/Thu gọn khi > 4 PTTT",
   "- Chính sách có hơn 4 PTTT",
   "1. Quan sát số PTTT hiển thị mặc định\n2. Click Xem thêm\n3. Click Thu gọn",
   "Mặc định hiển thị 4 PTTT đầu + Xem thêm; Xem thêm → hiển thị hết; Thu gọn → còn 4 PTTT đầu","Y"),
  ("Medium","Kiểm tra thứ tự ưu tiên PTTT có CTKM",
   "- Có PTTT kèm chương trình khuyến mãi",
   "1. Quan sát thứ tự các PTTT",
   "PTTT có CTKM được ưu tiên hiển thị lên trên","Y"),
  ("High","Kiểm tra default option, chỉ chọn 1 và không bỏ chọn được",
   "- Block PTTT đang hiển thị",
   "1. Quan sát option mặc định\n2. Cố chọn nhiều option / bỏ chọn option đầu",
   "Option đầu được chọn mặc định, không bỏ chọn được; chỉ chọn được 1 PTTT tại một thời điểm","Y"),
 ]),
 ("Nhóm 14: Block Thông tin thanh toán & Mã ưu đãi", [
  ("Low","Kiểm tra thu gọn/mở rộng và hyperlink điều khoản block Thông tin thanh toán",
   "- Đang ở màn hình Checkout",
   "1. Click Collapse/Expand block Thông tin thanh toán\n2. Click hyperlink điều khoản",
   "Thu gọn/mở rộng hoạt động; click điều khoản → màn hình điều khoản FPT Telecom","Y"),
  ("Medium","Kiểm tra hiển thị tên gói và giá theo chính sách",
   "- Có chính sách giá (chung cư/nhà phố)",
   "1. Quan sát tên gói và giá tại block Thông tin thanh toán",
   "Tên gói đúng; giá đúng theo chính sách; mặc định hiển thị giá gói 1 tháng","Y"),
  ("Medium","Kiểm tra ô mã khuyến mãi, link Chọn ưu đãi và trạng thái button Áp dụng",
   "- Block Thông tin thanh toán đang hiển thị",
   '1. Quan sát ô mã khi rỗng (button Áp dụng)\n2. Nhập 1 ký tự vào ô mã',
   'Ô input placeholder "Nhập mã khuyến mãi" + link "Chọn ưu đãi" (badge số) + button Áp dụng: disable khi rỗng, enable khi có ký tự',"Y"),
  ("Medium","Kiểm tra click link Chọn ưu đãi mở danh sách ưu đãi",
   "- Có ưu đãi đang hoạt động",
   '1. Click link "Chọn ưu đãi"',
   "Hiển thị danh sách ưu đãi khả dụng; badge đúng số lượng ưu đãi","Y"),
  ("High","Kiểm tra áp dụng mã khuyến mãi hợp lệ",
   "- Có mã khuyến mãi hợp lệ đang hoạt động",
   "1. Nhập mã khuyến mãi hợp lệ\n2. Click Áp dụng\n3. Quan sát Cần thanh toán",
   "Hiển thị giá trị giảm; Cần thanh toán = giá gốc − giá trị giảm (cập nhật đúng)","Y"),
  ("High","Kiểm tra áp dụng mã khuyến mãi không hợp lệ/hết hạn",
   "- Mã khuyến mãi không tồn tại hoặc đã hết hạn",
   "1. Nhập mã không hợp lệ/hết hạn\n2. Click Áp dụng",
   "Hiển thị thông báo lỗi (mã không hợp lệ/hết hạn); Cần thanh toán không thay đổi","Y"),
 ]),
 ("Nhóm 15: Luồng thanh toán", [
  ("High","Kiểm tra click Thanh toán khi chưa nhập đủ trường bắt buộc",
   PRE_CK,
   "1. Bỏ trống ít nhất 1 trường bắt buộc\n2. Click button Thanh toán",
   "Hiển thị lỗi ở các trường bắt buộc; không thực hiện thanh toán","Y"),
  ("High","Kiểm tra thanh toán với PTTT = COD",
   PRE_CK,
   "1. Nhập đầy đủ trường bắt buộc\n2. Chọn PTTT = COD\n3. Click Thanh toán",
   'Điều hướng màn hình Hoàn tất đơn hàng với trạng thái "Chưa thanh toán"',"Y"),
  ("High","Kiểm tra thanh toán với PTTT Online điều hướng sang cổng 3rd party",
   PRE_CK,
   "1. Nhập đầy đủ trường bắt buộc\n2. Chọn PTTT Online (VD: thẻ ATM/Visa)\n3. Click Thanh toán",
   "Điều hướng sang màn hình thanh toán của 3rd party tương ứng, hiển thị đúng số tiền cần thanh toán","Y"),
  ("High","Kiểm tra double-click button Thanh toán không tạo trùng đơn",
   PRE_CK,
   "1. Nhập đầy đủ trường bắt buộc, chọn PTTT Online\n2. Double-click button Thanh toán\n3. Kiểm tra log/đơn hàng tạo ra",
   "Hệ thống chỉ tạo 1 đơn hàng đúng thông tin, không bị double","Y"),
  ("High","Kiểm tra click Thanh toán khi quá session checkout (20 phút)",
   PRE_CK,
   "1. Nhập đầy đủ trường bắt buộc, chọn PTTT Online\n2. Chờ hơn 20 phút\n3. Click Thanh toán",
   "Hiển thị thông báo lỗi hợp lý; đẩy thông tin KH tiềm năng","Y"),
  ("Medium","Kiểm tra hết countdown thanh toán ở cổng 3rd party (~15 phút)",
   "- Đã điều hướng sang cổng thanh toán 3rd party",
   "1. Chờ hết thời gian thanh toán của kênh (~15 phút)",
   "Quay về tongdaiwifi và hiển thị màn hình không thành công","N"),
  ("Medium","Kiểm tra back từ cổng 3rd party về màn hình thanh toán",
   "- Đã điều hướng sang cổng thanh toán 3rd party",
   "1. Click icon back trên trình duyệt",
   "Về màn hình dịch vụ; disable các thông tin, chỉ cho cập nhật PTTT và mã ưu đãi","N"),
  ("High","Kiểm tra hủy/thanh toán thất bại tại cổng 3rd party",
   "- Đã điều hướng sang cổng thanh toán 3rd party",
   "1. Hủy thanh toán hoặc nhập sai thông tin thẻ nhiều lần",
   "Quay về tongdaiwifi, hiển thị màn hình không thành công; đẩy KHTN","N"),
  ("High","Kiểm tra click Thanh toán khi chính sách gói đã hết hiệu lực",
   "- Gói có ít nhất 1 chính sách đã hết hiệu lực trên QLCS",
   "1. Nhập đầy đủ trường bắt buộc, chọn PTTT\n2. Click Thanh toán",
   "Hệ thống hiển thị thông báo lỗi, không lên đơn hàng","Y"),
 ]),
 ("Nhóm 16: Màn hình Hoàn tất đơn hàng", [
  ("High","Kiểm tra hiển thị Mã đơn hàng và hyperlink Theo dõi ĐH",
   "- Đã hoàn tất đơn hàng thành công",
   "1. Quan sát khu vực Mã đơn hàng",
   'Hiển thị "Mã đơn hàng: <mã>" kèm hyperlink "Theo dõi ĐH"',"Y"),
  ("Medium","Kiểm tra click hyperlink Theo dõi đơn hàng",
   "- Đang ở màn hình Hoàn tất đơn hàng",
   "1. Click hyperlink Theo dõi đơn hàng",
   "Điều hướng đến màn hình theo dõi đơn hàng","Y"),
  ("Medium","Kiểm tra hiển thị và thu gọn/mở rộng block Thông tin khách hàng & Thông tin thanh toán",
   "- Đang ở màn hình Hoàn tất đơn hàng",
   "1. Quan sát thông tin 2 block\n2. Thu gọn/mở rộng từng block",
   "Hiển thị đúng thông tin đã nhập ở các bước trước; thu gọn/mở rộng hoạt động, icon đổi v ↔ ^","Y"),
  ("High","Kiểm tra trạng thái hoàn tất với COD và Online",
   "- Hoàn tất qua COD hoặc Online thành công",
   "1. Quan sát label trạng thái trên màn Hoàn tất",
   'COD → "Chưa thanh toán"; Online thành công → "Thanh toán thành công"; các bước header màu xanh lá',"Y"),
  ("High","Kiểm tra đối soát đơn hàng và hợp đồng sau thanh toán thành công",
   "- Đã thanh toán thành công",
   "1. Truy cập web-admin kiểm tra đơn hàng\n2. Truy cập inside kiểm tra hợp đồng",
   "Đơn hàng ghi nhận đúng thông tin đã nhập; Hợp đồng ghi nhận đúng thông tin đơn hàng","N"),
 ]),
 ("Nhóm 17: Empty/Error state, Quyền truy cập & Mobile (Mandatory)", [
  ("Medium","Kiểm tra truy cập màn hình Checkout khi chưa đăng nhập (guest)",
   "- Chưa đăng nhập tài khoản; có link đăng ký dịch vụ",
   "1. Mở link đăng ký dịch vụ ở trạng thái chưa đăng nhập\n2. Thực hiện checkout",
   "Cho phép checkout ở trạng thái guest (luồng đăng ký dịch vụ không bắt buộc đăng nhập)","Y"),
  ("Medium","Kiểm tra Block Phương thức thanh toán khi chính sách không khai báo PTTT nào",
   "- Gói có chính sách nhưng không khai báo PTTT nào",
   "1. Truy cập Checkout của gói này\n2. Quan sát Block Phương thức thanh toán",
   "[BLOCKED – cần confirm: khi chính sách không khai báo PTTT nào thì Block Phương thức thanh toán hiển thị thế nào (ẩn/placeholder/thông báo) và có cho thanh toán không?]","Y"),
  ("Medium","Kiểm tra xử lý khi API chính sách lỗi/timeout lúc kiểm tra giá",
   "- Block Địa chỉ lắp đặt đang hiển thị",
   "1. Chọn đủ địa chỉ để trigger API chính sách\n2. Giả lập API chính sách lỗi/timeout",
   "[BLOCKED – cần confirm: hành vi khi API chính sách lỗi/timeout — hiển thị thông báo gì, có chặn bước thanh toán không?]","Y"),
  ("Medium","Kiểm tra hiển thị màn hình Checkout trên mobile (≤ 768px)",
   "- Mở màn hình Checkout trên viewport ≤ 768px",
   "1. Quan sát layout tổng thể các block trên mobile",
   "Các block hiển thị đúng, không vỡ layout, không tràn ngang; các trường/button thao tác được","N"),
 ]),
]

# ====================== DATA: Checkout_Internet ======================
INTERNET = [
 ("Nhóm 1: Hiển thị tổng thể & Điều hướng 3 bước", [
  ("High","Kiểm tra hiển thị tổng thể luồng Checkout Internet (3 bước)",
   "- Đã chọn 1 gói Internet ở màn Chi tiết",
   "1. Quan sát tiến trình và bố cục các bước checkout Internet",
   "Luồng 3 bước: B1 Thông tin đăng ký → B2 Thanh toán → B3 Hoàn tất đơn hàng; B1 hiển thị các block Thông tin cá nhân, Địa chỉ lắp đặt, Thông tin lắp đặt, Thông tin thanh toán","N"),
  ("High",'Kiểm tra click "Đăng ký ngay" điều hướng vào luồng checkout 3 bước',
   "- Đang ở màn Chi tiết gói Internet",
   '1. Click button "Đăng ký ngay"',
   "Điều hướng vào luồng checkout gồm 3 bước: Thông tin đăng ký, Thanh toán, Hoàn tất đơn hàng","Y"),
  ("Low",'Kiểm tra button "Đăng ký ngay" hiển thị (có thể cần scroll)',
   "- Đang ở màn Chi tiết gói Internet",
   '1. Quan sát/scroll trang để tìm button "Đăng ký ngay"',
   'Tùy màn hình, button "Đăng ký ngay" có thể cần scroll xuống mới hiển thị',"N"),
 ]),
 ("Nhóm 2: Bước 1 — Thông tin đăng ký", [
  ("High","Kiểm tra hiển thị đầy đủ các block tại Bước 1",
   "- Đang ở Bước 1 luồng đăng ký Internet",
   "1. Quan sát toàn bộ màn Bước 1",
   "Hiển thị: Logo FPT, Block Thông tin cá nhân (Họ tên + SĐT), Block Địa chỉ lắp đặt, Block Thông tin lắp đặt, Block Thông tin thanh toán (giá 1 tháng), text điều khoản, button Tiếp tục","Y"),
  ("High","Kiểm tra button Tiếp tục khi thiếu trường bắt buộc",
   "- Đang ở Bước 1, còn trường bắt buộc trống",
   "1. Bỏ trống ít nhất 1 trường bắt buộc\n2. Click button Tiếp tục",
   "Không điều hướng sang Bước 2; hiển thị lỗi ở các trường bắt buộc","Y"),
  ("High","Kiểm tra button Tiếp tục khi nhập đủ hợp lệ",
   "- Đang ở Bước 1, đã nhập đủ hợp lệ",
   "1. Nhập đầy đủ trường bắt buộc hợp lệ\n2. Click button Tiếp tục",
   "Điều hướng sang Bước 2 — Thanh toán","Y"),
  ("Medium","Kiểm tra Block Thông tin lắp đặt load theo dữ liệu đã nhập",
   "- Đã nhập Thông tin cá nhân và Địa chỉ ở Bước 1",
   "1. Quan sát Block Thông tin lắp đặt",
   "Hiển thị Họ tên, SĐT, địa chỉ load đúng theo Block Thông tin cá nhân và Địa chỉ lắp đặt","Y"),
  ("Medium","Kiểm tra giá tiền Bước 1 thay đổi theo địa chỉ đã chọn",
   "- Đang ở Bước 1, đã chọn 1 địa chỉ",
   "1. Đổi sang địa chỉ lắp đặt khác",
   "Block Thông tin thanh toán cập nhật giá 1 tháng theo dữ liệu QLCS của địa chỉ mới","Y"),
 ]),
 ("Nhóm 3: Bước 2 — Thanh toán (trả trước/trả sau, giá động)", [
  ("High","Kiểm tra gói trả sau chỉ tính Phí lắp đặt khi thanh toán",
   '- Gói Internet được QLCS trả về loại "trả sau" (trả sau thường chu kỳ 1 tháng)\n- Đang ở Bước 2',
   "1. Quan sát Block Thông tin sản phẩm (trả sau)",
   "Chỉ tính tiền Phí lắp đặt lúc thanh toán; load gói + chu kỳ + giá theo QLCS","Y"),
  ("High","Kiểm tra gói trả trước tính Phí dịch vụ + Phí lắp đặt",
   '- Gói Internet được QLCS trả về loại "trả trước"\n- Đang ở Bước 2',
   "1. Quan sát Block Thông tin sản phẩm (trả trước)",
   "Tính tiền Phí dịch vụ + Phí lắp đặt; load gói + chu kỳ + giá theo QLCS","Y"),
  ("High","Kiểm tra Block Thông tin thanh toán hiển thị giá động theo địa chỉ",
   "- Đang ở Bước 2",
   "1. Quan sát Block Thông tin thanh toán\n2. Đối chiếu số tiền theo địa chỉ đã chọn",
   "Load đúng sản phẩm + số tiền tương ứng; giá động theo địa chỉ từ QLCS (mỗi địa chỉ có thể có giá khác)","Y"),
  ("Medium","Kiểm tra button Quay lại giữ nguyên dữ liệu Bước 1",
   "- Đang ở Bước 2",
   "1. Click button Quay lại",
   "Về Bước 1; dữ liệu đã nhập ở Bước 1 vẫn được giữ nguyên","Y"),
  ("High","Kiểm tra Block Phương thức thanh toán có đầy đủ Online và COD",
   "- Đang ở Bước 2",
   "1. Quan sát Block Phương thức thanh toán",
   "Hiển thị đầy đủ PTTT Online và COD theo chính sách QLCS (khác UltraFast — UltraFast không có COD)","Y"),
  ("High","Kiểm tra button Thanh toán validate và kiểm tra chính sách active",
   "- Đang ở Bước 2, đã nhập đủ, chính sách còn active",
   "1. Click button Thanh toán",
   "Validate trường bắt buộc + chính sách còn active → thực hiện luồng thanh toán theo PTTT đã chọn","Y"),
 ]),
 ("Nhóm 4: Bước 3 — Hoàn tất đơn hàng", [
  ("High","Kiểm tra hoàn tất đơn hàng với PTTT = COD",
   "- Thanh toán bằng COD",
   "1. Hoàn tất đặt hàng với PTTT = COD\n2. Quan sát màn hình Hoàn tất",
   'Màn Hoàn tất: "Chưa thanh toán" + nội dung "Đơn hàng đã đăng ký thành công. Kỹ thuật viên FPT sẽ liên hệ triển khai dịch vụ trong 8h-12h..."',"Y"),
  ("High","Kiểm tra hoàn tất đơn hàng với PTTT Online thanh toán thành công",
   "- Thanh toán Online thành công",
   "1. Hoàn tất thanh toán Online thành công\n2. Quan sát màn hình Hoàn tất",
   'Bước Hoàn tất header màu xanh lá; trạng thái "Đã thanh toán" + nội dung thông báo thành công',"N"),
  ("High","Kiểm tra thanh toán Online thất bại quay về màn thanh toán",
   "- Thanh toán Online thất bại",
   "1. Thực hiện thanh toán Online thất bại",
   "Quay về màn Thanh toán giữ nguyên thông tin đã chọn; chỉ cho thay đổi PTTT và nhập mã ưu đãi, các trường còn lại disable","N"),
 ]),
 ("Nhóm 5: Mobile (Mandatory)", [
  ("Medium","Kiểm tra hiển thị luồng Checkout Internet trên mobile (≤ 768px)",
   "- Mở luồng đăng ký Internet trên viewport ≤ 768px",
   "1. Quan sát layout 3 bước trên mobile",
   "Các bước và block hiển thị đúng, không vỡ layout, thao tác được trên mobile","N"),
 ]),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)
build_sheet(wb, "Checkout_Common", "TC_CKCOMMON",
            "Checkout chung — dịch vụ có Địa chỉ lắp đặt (Internet/Camera/Smart Home)", COMMON, first=True)
build_sheet(wb, "Checkout_Internet", "TC_INTERNET",
            "Đăng ký dịch vụ Internet — Checkout 3 bước", INTERNET)
import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)

# summary
def count(groups):
    n=hi=me=lo=auy=aun=bl=0
    for _,tcs in groups:
        for pri,title,pre,steps,exp,auto in tcs:
            n+=1
            hi+=pri=="High"; me+=pri=="Medium"; lo+=pri=="Low"
            auy+=auto=="Y"; aun+=auto=="N"
            bl+= "[BLOCKED" in exp
    return n,hi,me,lo,auy,aun,bl
for name,g in [("Checkout_Common",COMMON),("Checkout_Internet",INTERNET)]:
    n,hi,me,lo,auy,aun,bl=count(g)
    print("%s: %d TC | High:%d Med:%d Low:%d | Auto Y:%d N:%d | BLOCKED:%d"%(name,n,hi,me,lo,auy,aun,bl))
print("Saved:", OUT)
