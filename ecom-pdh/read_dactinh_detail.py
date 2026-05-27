import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r'E:\AI\Ecom\ecom-pdh\00_input\chucnang_QLdactinh\ISC_chucnang_dactinh_v1.0.xlsx')
ws = wb['Đặc tính']

target_rows = [76, 97, 101, 102, 109, 121, 122, 123, 125, 126, 127, 133, 136, 137, 138]

for r in target_rows:
    a = ws.cell(r, 1).value or ''
    b = ws.cell(r, 2).value or ''
    c = ws.cell(r, 3).value or ''
    d = ws.cell(r, 4).value or ''
    e = ws.cell(r, 5).value or ''
    f = ws.cell(r, 6).value or ''
    print(f"\nR{r:03d}")
    print(f"  B (Priority): {b}")
    print(f"  C (Title):    {str(c).replace(chr(10), ' | ')}")
    print(f"  D (Pre-cond): {str(d).replace(chr(10), ' | ')}")
    print(f"  E (Steps):    {str(e).replace(chr(10), ' | ')}")
    print(f"  F (Expected): {str(f).replace(chr(10), ' | ')}")
