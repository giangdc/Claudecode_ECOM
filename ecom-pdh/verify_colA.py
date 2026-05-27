from openpyxl import load_workbook

wb = load_workbook(r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx')
ws = wb.active

ai_rows = []
blank_rows = []

for r in range(9, ws.max_row + 1):
    col_f = ws.cell(r, 6).value  # Test Steps — nếu có thì là TC row
    col_b = ws.cell(r, 2).value  # TC ID / header
    col_a = ws.cell(r, 1).value

    # bỏ qua header rows (col_b chứa text section header)
    if col_b and ('DANH SACH' in str(col_b).upper() or 'TAO MOI' in str(col_b).upper() or 'CHINH SUA' in str(col_b).upper()):
        continue
    if not col_f:
        continue

    if col_a == 'AI':
        ai_rows.append(r)
    else:
        blank_rows.append(r)

print(f"Tong TC rows co du lieu: {len(ai_rows) + len(blank_rows)}")
print(f"  Co AI (updated/new):  {len(ai_rows)} rows -> {ai_rows}")
print(f"  Khong co AI (giu nguyen): {len(blank_rows)} rows")
print()

# Kiem tra cac dong AI co dung khong
EXPECTED_AI = [16,17,25,33,39,48,86,104,113,119,137]  # Step1 updated (pre-insert)
# After inserting 12 rows at 116: step1 rows >= 116 shift +12
# R119 (TC_02.108 CHINHSUA) -> +12 = 131; R137 -> +12 = 149
# After inserting 2 rows at 152: rows >= 152 shift +2
# None of step1 rows >= 152 so no further shift
step1_actual = []
for r in EXPECTED_AI:
    if r >= 116:
        r += 12
    step1_actual.append(r)

new_taomoi = list(range(116, 128))   # 12 new
new_chinhsua = list(range(152, 154)) # 2 new

expected_all = sorted(step1_actual + new_taomoi + new_chinhsua)
print(f"Expected AI rows: {expected_all}")
print(f"Actual   AI rows: {sorted(ai_rows)}")
print()
if sorted(ai_rows) == expected_all:
    print("OK: Col A chinh xac")
else:
    missing = set(expected_all) - set(ai_rows)
    extra   = set(ai_rows) - set(expected_all)
    if missing: print(f"MISSING AI: {sorted(missing)}")
    if extra:   print(f"EXTRA   AI: {sorted(extra)}")
