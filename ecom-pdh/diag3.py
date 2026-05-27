import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

src = load_workbook(r'E:\AI\Ecom\ecom-pdh\00_input\ISC_ECP_QLnoidunggoiban_V1.0_R1.xlsx')
ws = src.active

print("=== SOURCE: Alignment ALL cols for R10 (existing TC) ===")
for c in range(1, 8):
    cell = ws.cell(10, c)
    a = cell.alignment
    print(f"  col{c}: horiz={a.horizontal!r}  vert={a.vertical!r}  wrap={a.wrap_text}")

print("\n=== SOURCE: Row heights ===")
for r in [9, 10, 11, 31, 32, 116, 117]:
    h = ws.row_dimensions[r].height
    print(f"  R{r}: height={h}")

dst = load_workbook(r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx')
ws2 = dst.active

print("\n=== DST: Alignment ALL cols R10 vs R116 ===")
for r in [10, 116]:
    print(f"  Row {r}:")
    for c in range(1, 8):
        cell = ws2.cell(r, c)
        a = cell.alignment
        print(f"    col{c}: horiz={a.horizontal!r}  vert={a.vertical!r}  wrap={a.wrap_text}")

print("\n=== DST: Row heights existing vs new ===")
for r in [9, 10, 31, 32, 116, 117, 128, 129]:
    h = ws2.row_dimensions[r].height
    print(f"  R{r}: height={h}")

print("\n=== DST: Border check R10 vs R116 (col 4) ===")
for r in [10, 116]:
    cell = ws2.cell(r, 4)
    b = cell.border
    print(f"  R{r}: left={b.left.border_style}  right={b.right.border_style}  top={b.top.border_style}  bottom={b.bottom.border_style}")
