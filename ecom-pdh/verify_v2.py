import re
from openpyxl import load_workbook

wb = load_workbook(r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx')
ws = wb.active

print("=== Ghost merge check (rows 114-130) ===")
for rng in ws.merged_cells.ranges:
    nums = [int(x) for x in re.findall(r"\d+", str(rng))]
    if any(114 <= n <= 130 for n in nums):
        print(f"  merge: {rng}")

print()
print("=== R116 cells ===")
for c in range(1, 12):
    cell = ws.cell(116, c)
    t = type(cell).__name__
    v = cell.value
    short = str(v)[:60] if v else None
    print(f"  col{c}: type={t} val={short!r}")

print()
print("=== R116-R127 col1/col4 preview ===")
for r in range(116, 128):
    col1 = ws.cell(r, 1).value
    col4 = ws.cell(r, 4).value
    title = str(col4)[:70] if col4 else "(empty)"
    print(f"  R{r}: A={col1!r} | {title}")
