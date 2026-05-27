"""
So sanh FORMAT (fill, alignment, row height, font) giua file hien tai va reference.
"""
import sys, io, os, subprocess
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

CURRENT = r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx'
TEMP    = r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\_TEMP_reference.xlsx'
SCRIPT  = r'E:\AI\Ecom\ecom-pdh\build_tc_v2.py'

src = open(SCRIPT, encoding='utf-8').read()
src_patched = src.replace(
    r"dst = r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx'",
    f"dst = r'{TEMP}'"
)
tmp_script = r'E:\AI\Ecom\ecom-pdh\_build_ref_tmp.py'
with open(tmp_script, 'w', encoding='utf-8') as f:
    f.write(src_patched)
subprocess.run(['python', tmp_script], capture_output=True)
os.remove(tmp_script)

wb_cur = load_workbook(CURRENT)
wb_ref = load_workbook(TEMP)
ws_cur = wb_cur.active
ws_ref = wb_ref.active

def fill_str(cell):
    f = cell.fill
    if f.patternType == 'solid':
        return f"solid:{f.fgColor.rgb if f.fgColor.type=='rgb' else f'theme{f.fgColor.theme}'}"
    return 'none'

def align_str(cell):
    a = cell.alignment
    return f"h={a.horizontal} v={a.vertical} w={a.wrap_text}"

def font_str(cell):
    f = cell.font
    return f"bold={f.bold} size={f.size} name={f.name}"

diffs = []
max_row = max(ws_cur.max_row, ws_ref.max_row)

for r in range(9, max_row + 1):
    # Row height
    h_cur = ws_cur.row_dimensions[r].height
    h_ref = ws_ref.row_dimensions[r].height
    if h_cur != h_ref:
        diffs.append((r, 0, f'row_height', str(h_ref), str(h_cur)))

    for c in range(1, 12):
        cell_cur = ws_cur.cell(r, c)
        cell_ref = ws_ref.cell(r, c)

        # Fill
        fc = fill_str(cell_cur)
        fr = fill_str(cell_ref)
        if fc != fr:
            diffs.append((r, c, 'fill', fr, fc))

        # Alignment (only on data cols)
        if c in range(1, 8):
            ac = align_str(cell_cur)
            ar = align_str(cell_ref)
            if ac != ar:
                diffs.append((r, c, 'align', ar, ac))

        # Font bold (skip formula/value cols for noise)
        if c in [1, 2, 3, 4]:
            fc2 = font_str(cell_cur)
            fr2 = font_str(cell_ref)
            if fc2 != fr2:
                diffs.append((r, c, 'font', fr2, fc2))

os.remove(TEMP)

if not diffs:
    print("Khong co khac biet format nao!")
else:
    print(f"Tong so khac biet format: {len(diffs)}\n")
    cur_row = None
    for r, c, prop, v_ref, v_cur in diffs:
        if r != cur_row:
            col_a = ws_cur.cell(r, 1).value
            print(f"\n{'='*60}")
            print(f"ROW {r}  (ColA={col_a!r})")
            print(f"{'='*60}")
            cur_row = r
        col_lbl = f"col{c}" if c > 0 else "ROW"
        print(f"  [{prop} @ {col_lbl}]  SCRIPT: {v_ref!r}  ->  CURRENT: {v_cur!r}")
