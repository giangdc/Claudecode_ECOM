import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx')
ws = wb.active

AI_ROWS = [16,17,25,33,39,48,86,104,113,116,117,118,119,120,121,122,123,124,125,126,127,131,149,152,153]

print("=== SAMPLE CONTENT (AI rows — tiếng Việt có dấu) ===\n")
for r in AI_ROWS:
    col_a = ws.cell(r, 1).value
    title = ws.cell(r, 4).value or ''
    note  = ws.cell(r, 11).value or ''
    print(f"R{r:3d} | {str(title)[:70]}")
    print(f"     | Note: {str(note)[:60]}")
    print()
