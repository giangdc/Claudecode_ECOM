"""
So sanh file hien tai (manual-edited) voi output goc cua build script.
"""
import sys, io, os, shutil, subprocess
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

CURRENT = r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx'
TEMP    = r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\_TEMP_reference.xlsx'
SCRIPT  = r'E:\AI\Ecom\ecom-pdh\build_tc_v2.py'

# Build reference to TEMP by patching dst path
src = open(SCRIPT, encoding='utf-8').read()
src_patched = src.replace(
    r"dst = r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx'",
    f"dst = r'{TEMP}'"
)
tmp_script = r'E:\AI\Ecom\ecom-pdh\_build_ref_tmp.py'
with open(tmp_script, 'w', encoding='utf-8') as f:
    f.write(src_patched)

result = subprocess.run(['python', tmp_script], capture_output=True, text=True, encoding='utf-8')
os.remove(tmp_script)
if result.returncode != 0:
    print("Build reference FAILED:")
    print(result.stderr)
    sys.exit(1)

# Compare
wb_cur = load_workbook(CURRENT)
wb_ref = load_workbook(TEMP)
ws_cur = wb_cur.active
ws_ref = wb_ref.active

COL_NAMES = {
    1: 'A(QC/AI)',
    2: 'B(TC_ID_formula)',
    3: 'C(Priority)',
    4: 'D(NoiDungTest)',
    5: 'E(PreCondition)',
    6: 'F(Steps)',
    7: 'G(ExpResult)',
    11: 'K(ChangeNote)',
}

diffs = []
max_row = max(ws_cur.max_row, ws_ref.max_row)
for r in range(1, max_row + 1):
    for c in COL_NAMES:
        v_cur = ws_cur.cell(r, c).value
        v_ref = ws_ref.cell(r, c).value
        vc = str(v_cur).strip() if v_cur is not None else ''
        vr = str(v_ref).strip() if v_ref is not None else ''
        if vc != vr:
            diffs.append((r, c, vr, vc))

os.remove(TEMP)

if not diffs:
    print("Khong co khac biet nao!")
else:
    print(f"Tong so o khac nhau: {len(diffs)}\n")
    cur_row = None
    for r, c, v_ref, v_cur in diffs:
        if r != cur_row:
            col_a = ws_cur.cell(r, 1).value
            print(f"\n{'='*70}")
            print(f"ROW {r}  (ColA={col_a!r})")
            print(f"{'='*70}")
            cur_row = r
        col_name = COL_NAMES.get(c, f'col{c}')
        print(f"\n  [{col_name}]")
        ref_short = v_ref[:200] if v_ref else '(empty)'
        cur_short = v_cur[:200] if v_cur else '(empty)'
        print(f"  SCRIPT : {ref_short!r}")
        print(f"  MANUAL : {cur_short!r}")
