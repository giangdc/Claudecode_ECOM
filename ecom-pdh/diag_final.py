import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx')
ws = wb.active

print("=== 1. SECTION HEADER MERGES ===")
for rng in sorted(ws.merged_cells.ranges, key=lambda x: x.min_row):
    if rng.min_row >= 9:
        v = ws.cell(rng.min_row, 2).value
        cell = ws.cell(rng.min_row, 2)
        fill = cell.fill.fgColor.rgb if cell.fill.fgColor.type == 'rgb' else 'no-rgb'
        print(f"  {rng}  fill={fill!r}  val={str(v)[:35] if v else None!r}")

print("\n=== 2. ALIGNMENT CHECK (existing R10 vs new R116) ===")
for r, label in [(10, 'existing'), (116, 'new')]:
    print(f"  Row {r} ({label}):")
    for c in range(1, 8):
        cell = ws.cell(r, c)
        a = cell.alignment
        b = ws.cell(r, c).border.left.border_style
        print(f"    col{c}: horiz={a.horizontal!r} vert={a.vertical!r} wrap={a.wrap_text}  border={b!r}")

print("\n=== 3. HIGHLIGHT CHECK ===")
ai_highlighted = []
old_highlight = []
for r in range(10, ws.max_row + 1):
    cell4 = ws.cell(r, 4)
    ft = cell4.fill.patternType
    fg = cell4.fill.fgColor.rgb if cell4.fill.fgColor.type == 'rgb' else None
    col_a = ws.cell(r, 1).value
    if ft == 'solid' and fg == 'FFFFF2CC':
        if col_a == 'AI':
            ai_highlighted.append(r)
        else:
            old_highlight.append(r)
print(f"  AI rows highlighted (yellow): {len(ai_highlighted)} -> {ai_highlighted}")
print(f"  Non-AI rows still highlighted (should be 0): {old_highlight}")
