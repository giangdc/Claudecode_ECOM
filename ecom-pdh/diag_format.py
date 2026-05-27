from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook(r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx')
ws = wb.active

# 1. All merged ranges
print("=== ALL MERGED RANGES ===")
for rng in sorted(ws.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
    print(f"  {rng}")

# 2. Row 128 detail
print("\n=== ROW 128 (CHINH SUA header?) ===")
for c in range(1, 10):
    cell = ws.cell(128, c)
    t = type(cell).__name__
    v = cell.value
    print(f"  col{c}: type={t} val={str(v)[:50] if v else None!r}")

# 3. Check alignment of existing TCs vs new TCs
print("\n=== ALIGNMENT SAMPLE ===")
print("  Existing TCs (R10, R11, R12):")
for r in [10, 11, 12]:
    for c in [4, 5, 6, 7]:
        cell = ws.cell(r, c)
        a = cell.alignment
        print(f"    R{r}C{c}: horiz={a.horizontal} vert={a.vertical} wrap={a.wrap_text}")
    break  # just row 10

print("  New TCs (R116, R117):")
for r in [116, 117]:
    for c in [4, 5, 6, 7]:
        cell = ws.cell(r, c)
        a = cell.alignment
        print(f"    R{r}C{c}: horiz={a.horizontal} vert={a.vertical} wrap={a.wrap_text}")
    break

# 4. Highlight/fill check
print("\n=== FILL/HIGHLIGHT SAMPLE ===")
rows_to_check = [10, 11, 16, 17, 25, 32, 33, 116, 117]
for r in rows_to_check:
    cell = ws.cell(r, 4)
    f = cell.fill
    fg = f.fgColor.rgb if f.fgColor and f.fgColor.type == 'rgb' else f.fgColor.theme if f.fgColor else None
    fill_type = f.fill_type if hasattr(f, 'fill_type') else f.patternType
    print(f"  R{r}: fill_type={fill_type!r} fgColor={f.fgColor.rgb if f.fgColor.type=='rgb' else f.fgColor.type!r}")
