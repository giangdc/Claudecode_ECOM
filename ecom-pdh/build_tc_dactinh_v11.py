import sys, io, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

INPUT  = r'E:\AI\Ecom\ecom-pdh\00_input\chucnang_QLdactinh\ISC_chucnang_dactinh_v1.0.xlsx'
OUTPUT = r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_dactinh_v1.1.xlsx'

shutil.copy2(INPUT, OUTPUT)
print(f"Copied: {OUTPUT}")

wb = load_workbook(OUTPUT)
ws = wb['Đặc tính']

# ---- Styles ----
YELLOW  = PatternFill(fill_type='solid', fgColor='FFFFF2CC')
NO_FILL = PatternFill(fill_type=None)
_thin   = Side(style='thin')
TC_BDR  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HDR_FILL = PatternFill(fill_type='solid', fgColor='FFA8D08D')
CN = 11  # col K = Change Note

def note(r, txt):
    ws.cell(r, CN).value = txt

def tc_formula(r):
    return f'=IF(C{r}="","",$C$3&"."&COUNTA($C$9:C{r})&"")'

def write_new_tc(r, pri, title, pre, steps, exp, cn_text):
    ws.cell(r, 1).value = tc_formula(r)
    ws.cell(r, 2).value = pri
    ws.cell(r, 3).value = title
    ws.cell(r, 4).value = pre
    ws.cell(r, 5).value = steps
    ws.cell(r, 6).value = exp
    ws.cell(r, CN).value = cn_text
    # Alignment
    ws.cell(r, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.cell(r, 2).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for c in [3, 4, 5, 6]:
        ws.cell(r, c).alignment = Alignment(vertical='top', wrap_text=True)
    # Borders
    for c in range(1, 8):
        ws.cell(r, c).border = TC_BDR
    ws.row_dimensions[r].height = 60.0

# ============================================================
# STEP 1 — UPDATE existing rows (BEFORE any inserts)
# ============================================================

# R97: GUI Tạo mới — thêm Icon field
ws.cell(97, 6).value = (
    'Popup "Tạo mới đặc tính" hiển thị:\n'
    '- droplist Nhóm đặc tính\n'
    '- Icon đặc tính (upload ảnh JPG/PNG) [V1.2]\n'
    '- Tên đặc tính\n'
    '- droplist: Kiểu dữ liệu\n'
    '- Textbox: Giá trị đặc tính\n'
    '- btn Thêm giá trị\n'
    '- btn Thêm đặc tính\n'
    '- btn Tạo\n'
    '- Hủy\n'
    '- icon Close'
)
note(97, '[UPDATE V1.2] Bổ sung field "Icon đặc tính" vào GUI popup Tạo mới')

# R109: Tạo thành công — Icon là optional
ws.cell(109, 4).value = 'Popup Tạo mới đặc tính đang mở'
ws.cell(109, 5).value = (
    '1. Nhập Tên đặc tính hợp lệ\n'
    '2. Chọn Kiểu dữ liệu\n'
    '3. Nhập ít nhất 1 Giá trị đặc tính\n'
    '4. (Tuỳ chọn) Upload Icon đặc tính\n'
    '5. Click "Tạo"'
)
ws.cell(109, 6).value = (
    'Hệ thống hiển thị thông báo "Tạo thành công";\n'
    'popup đóng; đặc tính mới xuất hiện đầu danh sách\n'
    '- Tạo thành công dù có hoặc không upload Icon đặc tính'
)
note(109, '[UPDATE V1.2] Icon đặc tính là optional trong flow Tạo thành công')

# R121: GUI Chi tiết — thêm Icon field
ws.cell(121, 6).value = (
    'popup Hiển thị đúng các field:\n'
    '- Nhóm đặc tính\n'
    '- Icon đặc tính (nếu đặc tính có Icon) [V1.2]\n'
    '- Tên đặc tính\n'
    '- Kiểu dữ liệu\n'
    '- Giá trị đặc tính\n'
    '- btn Đóng'
)
note(121, '[UPDATE V1.2] Bổ sung field "Icon đặc tính" vào danh sách field popup Chi tiết')

# R123: Read-only check — thêm Icon
ws.cell(123, 5).value = '1. Click vào các field (Tên, Kiểu dữ liệu, Giá trị, Icon đặc tính)'
ws.cell(123, 6).value = (
    'Không cho chỉnh sửa, tất cả field ở chế độ read-only;\n'
    'Icon đặc tính chỉ hiển thị ảnh, không thể upload/thay đổi'
)
note(123, '[UPDATE V1.2] Bổ sung Icon đặc tính vào kiểm tra chế độ read-only')

# R133: Chỉnh sửa load data — thêm Icon
ws.cell(133, 6).value = (
    'Tất cả field hiển thị đúng dữ liệu hiện có;\n'
    'bao gồm Icon đặc tính (nếu đặc tính có Icon đã upload trước đó) [V1.2]'
)
note(133, '[UPDATE V1.2] Bổ sung kiểm tra Icon đặc tính hiển thị đúng trong popup Chỉnh sửa')

print("STEP 1 done: 5 rows updated")

# ============================================================
# STEP 2 — INSERT mới (BOTTOM TO TOP)
# ============================================================

# --- 2a: Insert 4 TCs mới vào section Chỉnh sửa sau R133 ---
ws.insert_rows(134, 4)
print("Inserted 4 rows at R134 (Chỉnh sửa)")

write_new_tc(134, 'Medium',
    'Kiểm tra field Icon đặc tính trong popup Chỉnh sửa — upload icon mới [V1.2]',
    'Popup Chỉnh sửa đặc tính đang mở (đặc tính có hoặc không có Icon)',
    '1. Click vào field Icon đặc tính\n2. Upload file JPG hoặc PNG hợp lệ (≤1MB)\n3. Click Lưu',
    'Icon mới được lưu thành công;\nnếu đặc tính đã có Icon cũ, Icon cũ bị thay thế',
    '[NEW V1.2] Icon đặc tính editable trong Chỉnh sửa — upload đè')

write_new_tc(135, 'Medium',
    'Kiểm tra Chỉnh sửa — không có nút xóa icon riêng [V1.2]',
    'Popup Chỉnh sửa đặc tính đang mở; đặc tính đang có Icon',
    '1. Quan sát field Icon đặc tính trong popup Chỉnh sửa\n2. Tìm nút xóa (remove) icon',
    'Không có nút xóa icon riêng;\nIcon chỉ có thể được thay thế bằng cách upload file mới đè lên',
    '[NEW V1.2] BA confirm: không có nút xóa icon riêng, chỉ upload đè')

write_new_tc(136, 'Medium',
    'Kiểm tra thay đổi Kiểu dữ liệu trong Chỉnh sửa [V1.2]',
    'Popup Chỉnh sửa đặc tính đang mở;\nđặc tính hiện có Kiểu dữ liệu = Dropdown;\nđặc tính đã được gán cho ít nhất 1 SKU',
    '1. Thay đổi Kiểu dữ liệu sang Multi-select Dropdown\n2. Click Lưu',
    'Cho phép thay đổi Kiểu dữ liệu;\nLưu thành công;\nSKU đã gán không bị ảnh hưởng',
    '[NEW V1.2] BA confirm: được phép đổi Kiểu dữ liệu; SKU không bị ảnh hưởng')

write_new_tc(137, 'Medium',
    'Kiểm tra KDL = Text trong Chỉnh sửa — không có nút Thêm giá trị [V1.2]',
    'Popup Chỉnh sửa đặc tính đang mở;\nKiểu dữ liệu của đặc tính = Text',
    '1. Quan sát Kiểu dữ liệu = "Text"\n2. Quan sát field Giá trị đặc tính',
    'Field Giá trị đặc tính hiển thị;\nKHÔNG có nút "Thêm giá trị";\nnhất quán với hành vi tại màn hình Tạo mới',
    '[NEW V1.2] BA confirm: KDL Text = Giá trị hiển thị, không có nút Thêm')

print("Wrote 4 new Chỉnh sửa TCs at R134-R137")

# --- 2b: Insert 2 TCs mới vào section Chi tiết sau R127 ---
ws.insert_rows(128, 2)
# Fix ghost merge for A131:F131 (shifted to A133:F133 but ghost left at A131:F131)
for rng in list(ws.merged_cells.ranges):
    if str(rng) == 'A131:F131':
        ws.merged_cells.ranges.discard(rng)
        print("Discarded ghost merge A131:F131")
        break
print("Inserted 2 rows at R128 (Chi tiết)")

write_new_tc(128, 'Medium',
    'Kiểm tra Icon đặc tính hiển thị read-only trong popup Chi tiết [V1.2]',
    'Đặc tính có Icon đặc tính đã được upload;\npopup Chi tiết đặc tính đang mở',
    '1. Mở popup Chi tiết của đặc tính có Icon\n2. Quan sát field Icon đặc tính',
    'Icon đặc tính hiển thị đúng ảnh đã upload;\nfield ở chế độ read-only (không thể upload/thay đổi)',
    '[NEW V1.2] Icon đặc tính hiển thị read-only trong Chi tiết')

write_new_tc(129, 'Medium',
    'Kiểm tra popup Chi tiết khi đặc tính không có Icon [V1.2]',
    'Đặc tính không có Icon (không upload trước đó);\npopup Chi tiết đang mở',
    '1. Mở popup Chi tiết của đặc tính không có Icon\n2. Quan sát field Icon đặc tính',
    'Field Icon không hiển thị ảnh (trống hoặc placeholder mặc định);\nkhông có lỗi',
    '[NEW V1.2] Kiểm tra Chi tiết khi đặc tính không có Icon')

print("Wrote 2 new Chi tiết TCs at R128-R129")

# --- 2c: Insert 6 TCs mới vào section Tạo mới sau R109 ---
ws.insert_rows(110, 6)
# Fix ghost merges for A118:F118 and A133:F133 (shifted up, ghosts left behind)
for target in ('A118:F118', 'A133:F133'):
    for rng in list(ws.merged_cells.ranges):
        if str(rng) == target:
            ws.merged_cells.ranges.discard(rng)
            print(f"Discarded ghost merge {target}")
            break
print("Inserted 6 rows at R110 (Tạo mới)")

write_new_tc(110, 'Medium',
    'Kiểm tra field Icon đặc tính hiển thị trong popup Tạo mới [V1.2]',
    'Popup Tạo mới đặc tính đang mở',
    '1. Mở popup Tạo mới đặc tính\n2. Quan sát các field trong popup',
    'Field "Icon đặc tính" hiển thị trong popup;\nfield là optional (không có dấu * bắt buộc)',
    '[NEW V1.2] Hiển thị field Icon đặc tính trong form Tạo mới')

write_new_tc(111, 'Medium',
    'Kiểm tra upload icon hợp lệ (JPG/PNG, ≤1MB) — thành công [V1.2]',
    'Popup Tạo mới đặc tính đang mở;\ncó file JPG hoặc PNG kích thước ≤1MB',
    '1. Click vào field Icon đặc tính\n2. Chọn file JPG hoặc PNG kích thước ≤1MB\n3. Quan sát kết quả',
    'Icon được upload thành công;\nhiển thị preview ảnh trong form;\nkhông có thông báo lỗi',
    '[NEW V1.2] Upload icon hợp lệ — JPG/PNG ≤1MB')

write_new_tc(112, 'Medium',
    'Kiểm tra upload icon — format không hợp lệ (không phải JPG/PNG) [V1.2]',
    'Popup Tạo mới đặc tính đang mở',
    '1. Click vào field Icon đặc tính\n2. Chọn file không phải JPG/PNG (ví dụ: .gif, .svg, .pdf)\n3. Quan sát kết quả',
    'Hệ thống từ chối file;\nhiển thị thông báo lỗi yêu cầu định dạng JPG hoặc PNG;\nfile không được upload',
    '[NEW V1.2] Upload icon format không hợp lệ → lỗi. BA confirm: chỉ nhận JPG/PNG')

write_new_tc(113, 'Medium',
    'Kiểm tra upload icon — kích thước vượt 1MB [V1.2]',
    'Popup Tạo mới đặc tính đang mở;\ncó file JPG/PNG kích thước > 1MB',
    '1. Click vào field Icon đặc tính\n2. Chọn file JPG/PNG có kích thước > 1MB\n3. Quan sát kết quả',
    'Hệ thống từ chối file;\nhiển thị thông báo lỗi kích thước tối đa 1MB;\nfile không được upload',
    '[NEW V1.2] Upload icon > 1MB → lỗi. BA confirm: max 1MB')

write_new_tc(114, 'Medium',
    'Kiểm tra upload icon lần 2 thay thế icon cũ (max 1 icon) [V1.2]',
    'Popup Tạo mới đặc tính đang mở;\nđã upload 1 icon JPG/PNG hợp lệ',
    '1. Upload thêm 1 file ảnh JPG/PNG hợp lệ thứ 2\n2. Quan sát kết quả',
    'Icon thứ 2 thay thế icon cũ;\nchỉ hiển thị 1 icon trong form;\nkhông có thông báo lỗi',
    '[NEW V1.2] Max 1 icon — upload đè thay thế icon cũ. BA confirm')

write_new_tc(115, 'Medium',
    'Kiểm tra Kiểu dữ liệu = Text — không có nút Thêm giá trị [V1.2]',
    'Popup Tạo mới đặc tính đang mở',
    '1. Chọn Kiểu dữ liệu = "Text"\n2. Quan sát field Giá trị đặc tính',
    'Field Giá trị đặc tính vẫn hiển thị;\nKHÔNG có nút "Thêm giá trị";\nkhông thể thêm predefined value khi Kiểu dữ liệu = Text',
    '[NEW V1.2] KDL Text: Giá trị hiển thị nhưng không có nút Thêm. BA confirm')

print("Wrote 6 new Tạo mới TCs at R110-R115")

# ============================================================
# STEP 3 — HIGHLIGHT AI rows + clear non-AI row fills
# ============================================================
# Final row positions after all 3 inserts:
# Updates (original → final):
#   R97  → R97  (no shift)
#   R109 → R109 (no shift)
#   R121 → R121+6 = R127  (shifted by insert_rows(110,6))
#   R123 → R123+6 = R129  (shifted by insert_rows(110,6))
#   R133 → R133+2+6 = R141 (shifted by insert_rows(128,2) then insert_rows(110,6))
#
# New TCs (position at time of writing → final):
#   R110-R115 (written after insert at 110 — no further inserts above) → R110-R115
#   R128-R129 (written after insert at 128 — then shifted by +6) → R134-R135
#   R134-R137 (written after insert at 134 — then shifted by +2+6) → R142-R145

AI_ROWS = {
    97, 109, 127, 129, 141,          # Updated rows
    110, 111, 112, 113, 114, 115,    # New Tạo mới
    134, 135,                         # New Chi tiết
    142, 143, 144, 145,               # New Chỉnh sửa
}

# Section headers to SKIP (keep their green fill)
HEADER_ROWS = {9, 28, 43, 54, 70, 96, 124, 139}
# Note: 124 = orig R118+6, 139 = orig R131+2+6

for r in range(10, ws.max_row + 1):
    if r in HEADER_ROWS:
        continue
    # Check if row has TC content (col B has priority text)
    b_val = ws.cell(r, 2).value
    if not b_val:
        continue
    target_fill = YELLOW if r in AI_ROWS else NO_FILL
    for c in range(1, 8):
        cell = ws.cell(r, c)
        if type(cell).__name__ == 'Cell':
            cell.fill = target_fill

print(f"STEP 3 done: highlighted {len(AI_ROWS)} AI rows, cleared others")

# ============================================================
# STEP 4 — Add Change Note column header
# ============================================================
ws.cell(7, CN).value = 'Change Note (Sprint V1.2)'
ws.cell(7, CN).font  = Font(bold=True)
ws.cell(7, CN).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.column_dimensions['K'].width = 45

# ============================================================
# STEP 5 — Verify section header merges are intact
# ============================================================
current_merges = {str(m) for m in ws.merged_cells.ranges}
print(f"\nCurrent merged ranges: {sorted(current_merges)}")

# Expected final section headers after inserts:
# A9:F9, A28:F28, A43:F43, A54:F54, A70:F70, A96:F96 (unchanged)
# A124:F124 (was A118:F118, shifted +6)
# A139:F139 (was A131:F131, shifted +2+6)
for expected in ('A9:F9', 'A28:F28', 'A43:F43', 'A54:F54', 'A70:F70', 'A96:F96', 'A124:F124', 'A139:F139'):
    status = '✓' if expected in current_merges else '✗ MISSING'
    print(f"  {status} {expected}")

# ============================================================
# STEP 6 — Save
# ============================================================
wb.save(OUTPUT)
print(f"\n✅ Saved: {OUTPUT}")
print(f"Total rows: {ws.max_row}")
print(f"\nSummary:")
print(f"  Reuse:    ~127 TCs")
print(f"  Update:   5 TCs (R97, R109, R127, R129, R141)")
print(f"  New:      12 TCs (R110-R115 + R134-R135 + R142-R145)")
print(f"  Obsolete: 0")
print(f"  AI rows highlighted: {sorted(AI_ROWS)}")
