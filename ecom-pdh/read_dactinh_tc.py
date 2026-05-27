import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r'E:\AI\Ecom\ecom-pdh\00_input\chucnang_QLdactinh\ISC_chucnang_dactinh_v1.0.xlsx')
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}  (max_row={ws.max_row}, max_col={ws.max_column})")
    print(f"{'='*80}")

    # Print merged cells
    merges = [str(m) for m in ws.merged_cells.ranges]
    print(f"Merged ranges ({len(merges)}): {merges[:20]}")

    # Print first 10 rows
    print("\n--- First rows ---")
    for r in range(1, min(12, ws.max_row+1)):
        row_vals = []
        for c in range(1, min(9, ws.max_column+1)):
            v = ws.cell(r, c).value
            row_vals.append(str(v)[:40] if v else '')
        print(f"R{r:03d}: {' | '.join(row_vals)}")

    # Print all rows with content
    print("\n--- All data rows ---")
    for r in range(1, ws.max_row+1):
        cols = [ws.cell(r, c).value for c in range(1, 9)]
        # Print header rows (col B merged) and TC rows (col C = priority)
        col_b = str(cols[1] or '').strip()[:80]
        col_c = str(cols[2] or '').strip()[:20]
        col_d = str(cols[3] or '').strip()[:60]
        if col_b or col_c:
            fill = ws.cell(r, 2).fill
            fill_color = fill.fgColor.rgb if fill and fill.fill_type == 'solid' else 'none'
            print(f"R{r:03d} [fill:{fill_color}] B={col_b} | C={col_c} | D={col_d}")
