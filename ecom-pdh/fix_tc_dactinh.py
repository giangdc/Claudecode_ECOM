import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

FILE = r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_dactinh_v1.1.xlsx'

wb = load_workbook(FILE)
ws = wb['Đặc tính']

# ============================================================
# FIX 1 — Add missing section header merges
# ============================================================
missing_merges = ['A124:F124', 'A139:F139']
for m in missing_merges:
    # make sure not already there (avoid duplicate)
    current = {str(r) for r in ws.merged_cells.ranges}
    if m not in current:
        ws.merge_cells(m)
        print(f"Added merge: {m}")
    else:
        print(f"Already exists: {m}")

# Style the header cells
HDR_FONT = Font(bold=True)
for r in (124, 139):
    cell = ws.cell(r, 1)
    cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')

# ============================================================
# FIX 2 — Correct TC ID formulas where row ref != actual row
# ============================================================
FORMULA_PATTERN = re.compile(r'=IF\(C(\d+)="",""', re.IGNORECASE)

def tc_formula(r):
    return f'=IF(C{r}="","",$C$3&"."&COUNTA($C$9:C{r})&"")'

fixed = 0
for r in range(9, ws.max_row + 1):
    cell = ws.cell(r, 1)
    v = cell.value
    if v and isinstance(v, str) and v.startswith('=IF(C'):
        m = FORMULA_PATTERN.match(v)
        if m:
            ref_row = int(m.group(1))
            if ref_row != r:
                cell.value = tc_formula(r)
                fixed += 1
                print(f"  R{r:03d}: fixed formula C{ref_row} → C{r}")

print(f"\nFIX 2 done: corrected {fixed} TC ID formulas")

# ============================================================
# Verify final merges
# ============================================================
current = sorted(str(r) for r in ws.merged_cells.ranges)
print(f"\nFinal merged ranges: {current}")
for expected in ('A9:F9', 'A28:F28', 'A43:F43', 'A54:F54', 'A70:F70', 'A96:F96', 'A124:F124', 'A139:F139'):
    status = '✓' if expected in current else '✗ MISSING'
    print(f"  {status} {expected}")

wb.save(FILE)
print(f"\n✅ Saved: {FILE}")
