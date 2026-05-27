"""
Build AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx
- Col A: 'AI' for all updated/new TCs
- Col K (11): Change Notes
- New TAOMOI TCs inserted BEFORE CHINH SUA header (R116)
- New CHINHSUA TCs inserted AFTER last existing CHINH SUA TC (R152)
"""
import sys, io, os, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# ── style constants ───────────────────────────────────────────────────
_thin = Side(style='thin')
TC_BORDER   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
YELLOW_FILL = PatternFill(fill_type='solid', fgColor='FFFFF2CC')
NO_FILL     = PatternFill(fill_type=None)
HDR_FILL    = PatternFill(fill_type='solid', fgColor='FF93C47D')

CN = 11  # Change Note column K

src = r'E:\AI\Ecom\ecom-pdh\00_input\ISC_ECP_QLnoidunggoiban_V1.0_R1.xlsx'
os.makedirs(r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional', exist_ok=True)
dst = r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx'
shutil.copy(src, dst)
wb = load_workbook(dst)
ws = wb.active

# Clear col A for all data rows — will re-mark AI only where changes were made
for r in range(9, ws.max_row + 1):
    ws.cell(r, 1).value = None

# ── helpers ──────────────────────────────────────────────────────────
def ai(row):
    ws.cell(row, 1).value = 'AI'

def note(row, text):
    ws.cell(row, CN).value = text

def tc_formula(r):
    return f'=IF(F{r}="","",$D$3&"."&COUNTA($F$9:F{r})&"")'

def write_tc(r, pri, title, pre, steps, exp, note_text):
    ws.cell(r, 1).value = 'AI'
    ws.cell(r, 2).value = tc_formula(r)
    ws.cell(r, 3).value = pri
    ws.cell(r, 4).value = title
    ws.cell(r, 5).value = pre
    ws.cell(r, 6).value = steps
    ws.cell(r, 7).value = exp
    ws.cell(r, CN).value = note_text
    # Alignment — match existing TC format
    for c in [1, 2, 3]:
        ws.cell(r, c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for c in [4, 5, 6, 7]:
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
    # Borders — thin on all 4 sides (match existing TCs)
    for c in range(1, 8):
        ws.cell(r, c).border = TC_BORDER


# ════════════════════════════════════════════════════════════════════
# STEP 1 — UPDATE existing TC rows
# ════════════════════════════════════════════════════════════════════

# DANH SACH ──────────────────────────────────────────────────────────
# TC_02.7 R16
ai(16)
ws.cell(16,4).value = 'Kiểm tra phân trang - hiển thị 20 bản ghi mặc định mỗi trang'
ws.cell(16,5).value = '1. Đã đăng nhập, có quyền xem.\n2. Hệ thống có hơn 21 bản ghi gói bán.'
ws.cell(16,7).value = 'Trang đầu tiên hiển thị đúng 20 bản ghi. Thanh điều hướng trang được hiển thị.'
note(16, '[UPDATE V1.2] Phân trang 20 bản ghi/trang (v1.0: 10)')

# TC_02.8 R17
ai(17)
ws.cell(17,7).value = 'Trang 2 hiển thị đúng các bản ghi tiếp theo. STT tiếp tục từ bản ghi thứ 21.'
note(17, '[UPDATE V1.2] STT trang 2 bắt đầu từ 21 (v1.0: 11)')

# TC_02.16 R25
ai(25)
ws.cell(25,7).value = 'Hệ thống hiển thị toàn bộ danh sách gói bán (20 bản ghi mặc định/trang).'
note(25, '[UPDATE V1.2] 20 bản ghi/trang (v1.0: 10)')

# TAO MOI ────────────────────────────────────────────────────────────
# TC_02.23 R33
ai(33)
ws.cell(33,4).value = 'Kiểm tra giao diện màn hình Tạo mới - hiển thị đủ các trường theo v1.2'
ws.cell(33,7).value = (
    'Form Tạo mới hiển thị đầy đủ theo v1.2:\n'
    ' - Dropdown Gói bán (bắt buộc)\n'
    ' - Dropdown Kênh bán (bắt buộc, multi-select)\n'
    ' - Bảng line items: Sản phẩm (SKU), Tên hiển thị trên kênh, Phương thức hiển thị, '
    'Sản phẩm nhận trong gói (checkbox), Hình ảnh SKU nhận trong gói\n'
    ' - Tên hiển thị trên kênh (cấp Gói bán)\n'
    ' - Icon gói bán (upload JPG/PNG/SVG <=1MB)\n'
    ' - Mô tả card gói (text, không bắt buộc)\n'
    ' - Mô tả ngắn gói bán (text editor, max 2000 ký tự)\n'
    ' - Mô tả dài gói bán (text editor, max 12000 ký tự)\n'
    ' - Ảnh gói bán (JPG/PNG <=1MB, max 10 ảnh)\n'
    ' - Đặc tính gói bán (section read-only, load từ cấu hình)\n'
    ' - Block Đặc quyền (Collapse/Expand)\n'
    ' - Link video gói bán (URL http/https, không bắt buộc)\n'
    ' - Hình ảnh banner đầu trang (max 10 JPG/PNG)\n'
    ' - Hình ảnh banner giữa trang (bảng: STT/Ảnh/Hoạt động/Xóa)\n'
    ' - Button Lưu, Button Hủy'
)
note(33, '[UPDATE V1.2] Thêm: Kênh bán multi-select, Icon gói bán, Mô tả card gói, Đặc tính read-only, Link video')

# TC_02.29 R39
ai(39)
ws.cell(39,4).value = 'Kiểm tra Kênh bán Multi-select - chọn nhiều kênh cùng lúc [V1.2]'
ws.cell(39,5).value = '1. Đã đăng nhập, có quyền tạo mới.\n 2. Gói bán có nhiều kênh hợp lệ (FPT.vn, tongdaiwifi.vn).'
ws.cell(39,6).value = ('1. Mở màn hình Tạo mới.\n'
    ' 2. Chọn Gói bán.\n'
    ' 3. Chọn Kênh bán A (FPT.vn).\n'
    ' 4. Chọn thêm Kênh bán B (tongdaiwifi.vn).\n'
    ' 5. Quan sát dropdown Kênh bán.')
ws.cell(39,7).value = 'Cả Kênh bán A và B đều được chọn và hiển thị. Hệ thống cho phép multi-select. Không deselect kênh khi chọn thêm.'
note(39, '[UPDATE V1.2] Kênh bán đổi từ single-select sang multi-select')

# TC_02.38 R48
ai(48)
ws.cell(48,6).value = ('1. Mở màn hình Tạo mới.\n'
    " 2. Chọn Gói bán, Kênh bán.\n"
    " 3. Uncheck checkbox 'Sản phẩm nhận trong gói' tại 1 line item.\n"
    " 4. Nhấn 'Lưu'.\n"
    ' 5. Kiểm tra thông tin sản phẩm nhận trong gói trên kênh đã thiết lập.')
ws.cell(48,7).value = 'Sản phẩm uncheck sẽ không hiển thị trên kênh bán (ẩn toàn bộ thông tin sản phẩm trên kênh).'
note(48, '[UPDATE V1.2] Mở rộng expected result: áp dụng mọi kênh (v1.0: chỉ ghi tongdaiwifi)')

# TC_02.76 R86
ai(86)
ws.cell(86,7).value = "Hệ thống hiển thị lỗi validation: 'Tiêu đề đặc quyền là bắt buộc'. Dữ liệu không được lưu."
note(86, '[UPDATE V1.2] Tiêu đề đặc quyền là bắt buộc khi có dòng (v1.0: không bắt buộc)')

# TC_02.94 R104
ai(104)
ws.cell(104,4).value = "Kiểm tra toggle 'Hoạt động' banner giữa trang mặc định = OFF (BR-B08) [V1.2]"
ws.cell(104,7).value = 'Toggle mặc định = Không hoạt động (OFF). Ảnh sẽ không hiển thị trên kênh nếu không bật toggle.'
note(104, '[UPDATE V1.2] Toggle mặc định OFF (v1.0: ON)')

# TC_02.103 R113
ai(113)
ws.cell(113,4).value = "Kiểm tra nút 'Hủy' (Tạo mới) hiển thị popup xác nhận hủy đúng nội dung [V1.2]"
ws.cell(113,6).value = ("1. Mở màn hình Tạo mới.\n"
    " 2. Nhập một số thông tin.\n"
    " 3. Nhấn nút 'Hủy'.\n"
    " 4. Quan sát popup hiển thị.")
ws.cell(113,7).value = ("Popup hiển thị với nội dung: 'Bạn có chắc muốn hủy? Dữ liệu chưa lưu sẽ bị mất.'"
    " Có 2 button: 'Quay lại' và 'Xác nhận'.")
note(113, '[UPDATE V1.2] Hủy có popup xác nhận (v1.0: không có popup, về Danh sách liền)')

# TC_02.108 R119
ai(119)
ws.cell(119,4).value = 'Kiểm tra load đúng dữ liệu bản ghi vào màn hình Chỉnh sửa theo v1.2'
ws.cell(119,7).value = ('Màn hình Chỉnh sửa load đúng toàn bộ dữ liệu theo v1.2:\n'
    ' - Gói bán, Kênh bán (read-only)\n'
    ' - Bảng line items (Sản phẩm, Tên hiển thị trên kênh, Phương thức hiển thị, Sản phẩm nhận trong gói, Hình ảnh SKU)\n'
    ' - Tên hiển thị trên kênh (cấp Gói bán)\n'
    ' - Icon gói bán (hiển thị ảnh đã upload nếu có)\n'
    ' - Mô tả card gói (text)\n'
    ' - Mô tả ngắn, Mô tả dài (text editor)\n'
    ' - Ảnh gói bán\n'
    ' - Đặc tính gói bán (read-only section)\n'
    ' - Block Đặc quyền (Tiêu đề chung, Hình ảnh đặc quyền, danh sách đặc quyền)\n'
    ' - Hình ảnh banner đầu trang\n'
    ' - Hình ảnh banner giữa trang (bảng STT/Ảnh/Toggle/Xóa)\n'
    ' - Link video gói bán')
note(119, '[UPDATE V1.2] Thêm kiểm tra: Icon gói bán, Mô tả card gói, Đặc tính (read-only), Link video')

# TC_02.126 R137
ai(137)
ws.cell(137,4).value = "Kiểm tra nút 'Hủy' (Chỉnh sửa) hiển thị popup xác nhận hủy đúng nội dung [V1.2]"
ws.cell(137,6).value = ("1. Mở màn hình Chỉnh sửa.\n"
    " 2. Thay đổi 1 trường bất kỳ.\n"
    " 3. Nhấn nút 'Hủy'.\n"
    " 4. Quan sát popup hiển thị.")
ws.cell(137,7).value = ("Popup hiển thị với nội dung: 'Bạn có chắc muốn hủy? Dữ liệu chưa lưu sẽ bị mất.'"
    " Có 2 button: 'Quay lại' và 'Xác nhận'.")
note(137, '[UPDATE V1.2] Hủy có popup xác nhận (v1.0: không có popup, về Danh sách liền)')

print('Step 1 done: 11 TC rows updated')


# ════════════════════════════════════════════════════════════════════
# STEP 2 — Insert 12 NEW TAOMOI TCs before CHINH SUA header (R116)
#   After insertion: rows 116-127 = new TCs; R128 = CHINH SUA header
# ════════════════════════════════════════════════════════════════════
TAOMOI_NEW = [
    ('High',
     'Kiểm tra kênh bán đã sử dụng không hiển thị trong dropdown Tạo mới [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.\n 2. Gói bán "Gói A" đã có nội dung tạo với kênh "FPT.vn".',
     '1. Mở màn hình Tạo mới.\n 2. Chọn Gói bán "Gói A".\n 3. Mở dropdown Kênh bán.\n 4. Quan sát danh sách.',
     'Kênh "FPT.vn" không xuất hiện trong dropdown (đã được sử dụng). Chỉ hiển thị kênh chưa tạo nội dung cho Gói A.',
     '[NEW V1.2] SC-TAOMOI-002'),
    ('Medium',
     'Kiểm tra upload Icon gói bán hợp lệ (JPG/PNG/SVG <=1MB) - Tạo mới [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.\n 2. Có file icon.png (PNG, 500KB).',
     '1. Mở màn hình Tạo mới.\n 2. Tại field "Icon gói bán" nhấn upload.\n 3. Chọn file icon.png (500KB).\n 4. Quan sát kết quả.',
     'File icon upload thành công. Hiển thị preview trong field. Không hiển thị thông báo lỗi.',
     '[NEW V1.2] SC-TAOMOI-006'),
    ('Medium',
     'Kiểm tra upload Icon gói bán - từ chối sai định dạng (GIF) - Tạo mới [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.\n 2. Có file icon.gif (GIF, 200KB).',
     '1. Mở màn hình Tạo mới.\n 2. Tại field "Icon gói bán" nhấn upload.\n 3. Chọn file icon.gif.',
     'Hệ thống từ chối file. Hiển thị thông báo lỗi: chỉ chấp nhận JPG, PNG hoặc SVG.',
     '[NEW V1.2] SC-TAOMOI-007'),
    ('Medium',
     'Kiểm tra upload Icon gói bán - từ chối vượt 1MB - Tạo mới [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.\n 2. Có file icon_large.jpg (1.5MB).',
     '1. Mở màn hình Tạo mới.\n 2. Tại field "Icon gói bán" nhấn upload.\n 3. Chọn file 1.5MB.\n 4. Quan sát kết quả.',
     'Hệ thống từ chối file. Hiển thị thông báo lỗi: dung lượng tối đa 1MB/ảnh.',
     '[NEW V1.2] SC-TAOMOI-006 negative'),
    ('Medium',
     'Kiểm tra trường "Mô tả card gói" - không bắt buộc, lưu thành công khi để trống - Tạo mới [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.\n 2. Đã điền đủ trường bắt buộc (Gói bán, Kênh bán, Tên hiển thị trên kênh).',
     '1. Mở màn hình Tạo mới.\n 2. Điền đủ trường bắt buộc.\n 3. Để trống field "Mô tả card gói".\n 4. Nhấn "Lưu".',
     'Hệ thống lưu thành công. Không có lỗi validation cho trường Mô tả card gói.',
     '[NEW V1.2] REQ-TAOMOI-012'),
    ('Medium',
     'Kiểm tra section "Đặc tính gói bán" hiển thị read-only - Tạo mới [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.\n 2. Gói bán "Gói A" có cấu hình đặc tính gói bán.',
     '1. Mở màn hình Tạo mới.\n 2. Chọn Gói bán "Gói A".\n 3. Quan sát section "Đặc tính gói bán".\n 4. Thử chỉnh sửa bất kỳ trường nào trong section.',
     'Section "Đặc tính gói bán" hiển thị dữ liệu từ cấu hình Gói bán. Tất cả trường ở trạng thái read-only, không chỉnh sửa được.',
     '[NEW V1.2] SC-TAOMOI-017 | CLA-008'),
    ('Medium',
     'Kiểm tra section "Đặc tính gói bán" load đúng theo Gói bán đã chọn - Tạo mới [V1.2]',
     '1. Đã đăng nhập.\n 2. Có "Gói A" và "Gói B" với cấu hình đặc tính khác nhau.',
     '1. Mở màn hình Tạo mới.\n 2. Chọn Gói bán "Gói A", quan sát đặc tính.\n 3. Đổi sang Gói bán "Gói B".\n 4. Quan sát lại đặc tính.',
     'Khi chọn Gói A: hiển thị đặc tính Gói A. Đổi sang Gói B: section cập nhật hiển thị đặc tính Gói B. Không hiển thị đặc tính gói khác.',
     '[NEW V1.2] SC-TAOMOI-018 | CLA-008'),
    ('Medium',
     'Kiểm tra field "Link video gói bán" - URL hợp lệ http/https - Tạo mới [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.',
     '1. Mở màn hình Tạo mới.\n 2. Nhập "https://youtube.com/watch?v=example123" vào field "Link video gói bán".\n 3. Nhấn "Lưu" (đã điền đủ trường bắt buộc).',
     'Hệ thống chấp nhận URL hợp lệ. Không hiển thị thông báo lỗi. Dữ liệu được lưu thành công.',
     '[NEW V1.2] SC-TAOMOI-028'),
    ('Medium',
     'Kiểm tra field "Link video gói bán" - URL sai định dạng thiếu http/https - Tạo mới [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.',
     '1. Mở màn hình Tạo mới.\n 2. Nhập "youtube.com/watch?v=example" (thiếu http/https) vào "Link video".\n 3. Nhấn "Lưu".',
     'Hệ thống hiển thị lỗi: "Link video không hợp lệ. Vui lòng nhập đúng định dạng URL (http:// hoặc https://)". Không lưu.',
     '[NEW V1.2] SC-TAOMOI-029'),
    ('High',
     'Kiểm tra lưu với nhiều kênh bán - tạo N records theo N kênh đã chọn [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.\n 2. Gói bán "Gói A" có giá trên 2 kênh: FPT.vn và tongdaiwifi.vn. Cả 2 kênh chưa có nội dung.',
     '1. Mở màn hình Tạo mới.\n 2. Chọn Gói bán "Gói A".\n 3. Chọn Kênh bán: FPT.vn VÀ tongdaiwifi.vn (multi-select).\n 4. Điền đủ thông tin bắt buộc.\n 5. Nhấn "Lưu".',
     'Hệ thống lưu thành công. Tạo 2 records riêng biệt trong DB. Hiển thị thông báo thành công. Chuyển về Danh sách với 2 dòng mới: "Gói A - FPT.vn" và "Gói A - tongdaiwifi.vn".',
     '[NEW V1.2] SC-TAOMOI-033 | CLA-005'),
    ('High',
     'Kiểm tra popup Hủy (Tạo mới) - nhấn "Xác nhận" -> về Danh sách, không lưu [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.\n 2. Đang ở màn hình Tạo mới, đã nhập thông tin.\n 3. Popup "Bạn có chắc muốn hủy? Dữ liệu chưa lưu sẽ bị mất." đang hiển thị.',
     '1. Mở màn hình Tạo mới.\n 2. Nhập một số thông tin.\n 3. Nhấn "Hủy" để popup hiện ra.\n 4. Nhấn button "Xác nhận".',
     'Popup đóng. Dữ liệu không được lưu. Hệ thống quay về màn hình Danh sách. Không có bản ghi mới.',
     '[NEW V1.2] SC-TAOMOI-036 | CLA-001'),
    ('High',
     'Kiểm tra popup Hủy (Tạo mới) - nhấn "Quay lại" -> đóng popup, tiếp tục nhập [V1.2]',
     '1. Đã đăng nhập, có quyền tạo mới.\n 2. Đang ở màn hình Tạo mới, đã nhập thông tin.\n 3. Popup "Bạn có chắc muốn hủy? Dữ liệu chưa lưu sẽ bị mất." đang hiển thị.',
     '1. Mở màn hình Tạo mới.\n 2. Nhập một số thông tin.\n 3. Nhấn "Hủy" để popup hiện ra.\n 4. Nhấn button "Quay lại".',
     'Popup đóng. Hệ thống quay về form Tạo mới. Tất cả dữ liệu đã nhập trước đó vẫn còn nguyên trong form.',
     '[NEW V1.2] SC-TAOMOI-037 | CLA-001'),
]

ws.insert_rows(116, amount=12)
# openpyxl bug: insert_rows leaves a ghost merge at the original position;
# standard unmerge_cells fails on ghost ranges, so remove directly from the set
for rng in list(ws.merged_cells.ranges):
    if str(rng) == 'B116:G116':
        ws.merged_cells.ranges.discard(rng)
        break
for i, (pri, title, pre, steps, exp, note_text) in enumerate(TAOMOI_NEW):
    write_tc(116 + i, pri, title, pre, steps, exp, note_text)
    ws.row_dimensions[116 + i].height = 60.0

print('Step 2 done: 12 new TAOMOI rows inserted at R116-R127')
# After this: CHINH SUA header at R128, old CHINH SUA TCs at R129-R151, empty at R152


# ════════════════════════════════════════════════════════════════════
# STEP 3 — Insert 2 NEW CHINHSUA TCs at R152 (after last CHINH SUA TC)
#   Old CHINH SUA TCs: R129-R151; empty row was R152, now moves to R154
# ════════════════════════════════════════════════════════════════════
CHINHSUA_NEW = [
    ('High',
     'Kiểm tra popup Hủy (Chỉnh sửa) - nhấn "Xác nhận" -> về Danh sách, không lưu [V1.2]',
     '1. Đã đăng nhập, có quyền chỉnh sửa.\n 2. Đang ở màn hình Chỉnh sửa, đã thay đổi một số field.\n 3. Popup "Bạn có chắc muốn hủy? Dữ liệu chưa lưu sẽ bị mất." đang hiển thị.',
     '1. Mở màn hình Chỉnh sửa.\n 2. Thay đổi 1 trường bất kỳ.\n 3. Nhấn "Hủy" để popup hiện ra.\n 4. Nhấn button "Xác nhận".',
     'Popup đóng. Thay đổi không được lưu. Hệ thống quay về màn hình Danh sách. Dữ liệu bản ghi không bị thay đổi.',
     '[NEW V1.2] SC-CHINHSUA-011 | CLA-001'),
    ('High',
     'Kiểm tra popup Hủy (Chỉnh sửa) - nhấn "Quay lại" -> đóng popup, tiếp tục sửa [V1.2]',
     '1. Đã đăng nhập, có quyền chỉnh sửa.\n 2. Đang ở màn hình Chỉnh sửa, đã thay đổi một số field.\n 3. Popup "Bạn có chắc muốn hủy? Dữ liệu chưa lưu sẽ bị mất." đang hiển thị.',
     '1. Mở màn hình Chỉnh sửa.\n 2. Thay đổi 1 trường bất kỳ.\n 3. Nhấn "Hủy" để popup hiện ra.\n 4. Nhấn button "Quay lại".',
     'Popup đóng. Hệ thống quay về form Chỉnh sửa. Tất cả thay đổi vẫn còn nguyên trong form. Người dùng có thể tiếp tục chỉnh sửa.',
     '[NEW V1.2] SC-CHINHSUA-012 | CLA-001'),
]

ws.insert_rows(152, amount=2)
for i, (pri, title, pre, steps, exp, note_text) in enumerate(CHINHSUA_NEW):
    write_tc(152 + i, pri, title, pre, steps, exp, note_text)
    ws.row_dimensions[152 + i].height = 60.0

print('Step 3 done: 2 new CHINHSUA rows inserted at R152-R153')


# ════════════════════════════════════════════════════════════════════
# STEP 4 — Rewrite ALL col-B formulas with correct row numbers
# ════════════════════════════════════════════════════════════════════
rewritten = 0
for r in range(10, ws.max_row + 1):
    cell_f = ws.cell(r, 6)
    if cell_f.value is not None and str(cell_f.value).strip():
        ws.cell(r, 2).value = tc_formula(r)
        rewritten += 1

print(f'Step 4 done: {rewritten} col-B formulas rewritten')


# ════════════════════════════════════════════════════════════════════
# STEP 5 — Re-merge & re-style CHINH SUA header at R128
#   insert_rows(116,12) shifts B116:G116 → should be B128:G128,
#   but openpyxl drops it when we remove the ghost; add it back manually.
# ════════════════════════════════════════════════════════════════════
ws.merge_cells('B128:G128')
hdr128 = ws.cell(128, 2)
hdr128.fill = HDR_FILL
ref_font = ws.cell(9, 2).font
hdr128.font = Font(name=ref_font.name, bold=True, size=ref_font.size,
                   color=ref_font.color.rgb if ref_font.color.type == 'rgb' else 'FF000000')
hdr128.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[128].height = 15.75  # match DANH SACH / TAO MOI header height
print('Step 5 done: B128:G128 merged and styled (CHỈNH SỬA header)')


# ════════════════════════════════════════════════════════════════════
# STEP 6 — Reset all highlights; apply yellow only to AI rows
# ════════════════════════════════════════════════════════════════════
AI_ROWS = {16,17,25,33,39,48,86,104,113,
           116,117,118,119,120,121,122,123,124,125,126,127,
           131,149,152,153}

SECTION_KEYWORDS = ('DANH S', 'TẠO MỚI', 'CHỈNH SỮA',
                    'DANH SACH', 'TAO MOI', 'CHINH SUA')

highlighted = 0
cleared = 0
for r in range(10, ws.max_row + 1):
    v2 = str(ws.cell(r, 2).value or '')
    # Skip section header rows
    if any(k in v2.upper() for k in SECTION_KEYWORDS):
        continue
    # Only process rows that have test steps (real TC rows)
    if not ws.cell(r, 6).value:
        continue
    fill = YELLOW_FILL if r in AI_ROWS else NO_FILL
    for c in range(1, 8):
        cell = ws.cell(r, c)
        if type(cell).__name__ == 'Cell':
            cell.fill = fill
    if r in AI_ROWS:
        highlighted += 1
    else:
        cleared += 1

print(f'Step 6 done: {highlighted} rows highlighted (AI), {cleared} rows cleared')


# ════════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════════
wb.save(dst)
print(f'\nSaved: {dst}')
print(f'Total rows: {ws.max_row}')

# Quick summary
taomoi_header_row = None
chinhsua_header_row = None
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 2).value
    if v and 'TAO MOI' in str(v).upper():
        taomoi_header_row = r
    if v and 'CHINH SUA' in str(v).upper():
        chinhsua_header_row = r

# count TCs per section
danh_sach_count = sum(1 for r in range(10, 31) if ws.cell(r,6).value)
taomoi_count = sum(1 for r in range(32, 128) if ws.cell(r,6).value)
chinhsua_count = sum(1 for r in range(129, ws.max_row+1) if ws.cell(r,6).value)

print(f'\nTC count by section:')
print(f'  DANH SACH:  {danh_sach_count} TCs (TC_02.1 - TC_02.{danh_sach_count})')
print(f'  TAO MOI:    {taomoi_count} TCs (TC_02.{danh_sach_count+1} - TC_02.{danh_sach_count+taomoi_count})')
print(f'  CHINH SUA:  {chinhsua_count} TCs (TC_02.{danh_sach_count+taomoi_count+1} - TC_02.{danh_sach_count+taomoi_count+chinhsua_count})')
print(f'  TOTAL:      {danh_sach_count+taomoi_count+chinhsua_count} TCs')
