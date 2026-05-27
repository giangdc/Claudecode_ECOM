import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

# Check SOURCE file for original styling reference
src = load_workbook(r'E:\AI\Ecom\ecom-pdh\00_input\ISC_ECP_QLnoidunggoiban_V1.0_R1.xlsx')
ws_src = src.active

print("=== SOURCE: Merges (section headers) ===")
for rng in sorted(ws_src.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
    r = rng.min_row
    if r in (9, 31, 116):
        v = ws_src.cell(r, 2).value
        print(f"  {rng}  val={v!r}")

print("\n=== SOURCE: Section header R9 format (col B) ===")
c9 = ws_src.cell(9, 2)
print(f"  fill_type={c9.fill.patternType!r}  fgRGB={c9.fill.fgColor.rgb!r}")
print(f"  font bold={c9.font.bold}  font color={c9.font.color.rgb!r}")
print(f"  align horiz={c9.alignment.horizontal}  vert={c9.alignment.vertical}")

print("\n=== SOURCE: Section header R31 format (col B) ===")
c31 = ws_src.cell(31, 2)
print(f"  fill_type={c31.fill.patternType!r}  fgRGB={c31.fill.fgColor.rgb!r}")
print(f"  font bold={c31.font.bold}  font color={c31.font.color.rgb!r}")

print("\n=== SOURCE: Normal TC R10 alignment (cols 4-7) ===")
for c in range(4, 8):
    cell = ws_src.cell(10, c)
    a = cell.alignment
    print(f"  col{c}: horiz={a.horizontal!r} vert={a.vertical!r} wrap={a.wrap_text}")

print("\n=== SOURCE: Fill of various TC rows ===")
for r in [10, 11, 12, 16, 17, 25, 32, 33]:
    cell = ws_src.cell(r, 4)
    ft = cell.fill.patternType
    fg = cell.fill.fgColor.rgb if cell.fill.fgColor.type == 'rgb' else f"theme:{cell.fill.fgColor.theme}"
    print(f"  R{r}: fill_type={ft!r}  fgColor={fg!r}")

# Check DST file
print("\n\n=== DST: Merges ===")
dst = load_workbook(r'E:\AI\Ecom\ecom-pdh\03_test-cases\functional\AI_ISC_ecom-pdh_v1.1_TC_v2.0.xlsx')
ws_dst = dst.active
for rng in sorted(ws_dst.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
    if rng.min_row >= 9:
        v = ws_dst.cell(rng.min_row, 2).value
        print(f"  {rng}  val={str(v)[:40] if v else None!r}")

print("\n=== DST: Row 128 (CHINH SUA header expected) ===")
for c in range(1, 8):
    cell = ws_dst.cell(128, c)
    print(f"  col{c}: type={type(cell).__name__}  val={str(cell.value)[:50] if cell.value else None!r}")

print("\n=== DST: New TC R116 alignment (cols 4-7) ===")
for c in range(4, 8):
    cell = ws_dst.cell(116, c)
    a = cell.alignment
    print(f"  col{c}: horiz={a.horizontal!r} vert={a.vertical!r} wrap={a.wrap_text}")

print("\n=== DST: Fill of AI rows vs normal rows ===")
for r in [10, 11, 16, 17, 116, 117, 131, 149]:
    cell = ws_dst.cell(r, 4)
    ft = cell.fill.patternType
    fg = cell.fill.fgColor.rgb if cell.fill.fgColor.type == 'rgb' else f"theme:{cell.fill.fgColor.theme}"
    col_a = ws_dst.cell(r, 1).value
    print(f"  R{r} (A={col_a!r}): fill_type={ft!r}  fgColor={fg!r}")
