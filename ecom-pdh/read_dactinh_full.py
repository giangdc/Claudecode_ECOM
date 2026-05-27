import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook(r'E:\AI\Ecom\ecom-pdh\00_input\chucnang_QLdactinh\ISC_chucnang_dactinh_v1.0.xlsx')
ws = wb['Đặc tính']

print(f"Sheet 'Đặc tính': {ws.max_row} rows x {ws.max_column} cols")
print(f"Merged: {[str(m) for m in ws.merged_cells.ranges]}\n")

# Print full content A-F for all rows
for r in range(1, ws.max_row+1):
    cells = []
    for c in range(1, 8):
        v = ws.cell(r, c).value
        cells.append(str(v or '').replace('\n',' ').strip()[:70])

    col_a = cells[0]
    col_b = cells[1]
    col_c = cells[2]
    col_d = cells[3]
    col_e = cells[4]
    col_f = cells[5]

    # Detect fill
    fill = ws.cell(r, 1).fill
    if fill and fill.fill_type == 'solid':
        rgb = fill.fgColor.rgb
    else:
        # try col B
        fill = ws.cell(r, 2).fill
        rgb = fill.fgColor.rgb if fill and fill.fill_type == 'solid' else 'none'

    if col_a or col_b or col_c:
        is_header = (rgb not in ('none', 'FF000000', '00000000')) and col_c == ''
        marker = '[HDR]' if is_header else '     '
        print(f"R{r:03d}{marker} A={col_a[:30]} | B={col_b[:30]} | C={col_c[:50]} | D={col_d[:40]}")
