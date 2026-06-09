"""
Generate API TC Excel — chucnang_Voucher (ECP_API_voucher_v1.xlsx)
Version: v1.2  |  Date: 2026-05-28
CLARY-001 Resolved: CHECKOUT_TOKEN_INVALID→401; missing/notfound/noPTTT→400
CLARY-004 Resolved: data=null content case removed
CLARY-005 Resolved: type mismatch → VOUCHER_INVALID 400
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "ecom-pdh/03_test-cases/api/AI_ISC_ecom-pdh_v1.1_TC_API_v1.2.xlsx"
BASE_URL = "http://ecp-api-stag.fpt.net/ordering"

# Colors
C_DARK_BLUE  = "1F3864"
C_MED_BLUE   = "2E75B6"
C_GROUP_BLUE = "D9E1F2"
C_COL_HEAD   = "4472C4"
C_SECURITY   = "FFF2CC"   # yellow tint for security-inferred rows
C_BLOCKED    = "FFE0E0"   # pink for BLOCKED

FILL_INFO   = PatternFill("solid", fgColor=C_DARK_BLUE)
FILL_COL    = PatternFill("solid", fgColor=C_COL_HEAD)
FILL_GROUP  = PatternFill("solid", fgColor=C_GROUP_BLUE)
FILL_WHITE  = PatternFill("solid", fgColor="FFFFFF")
FILL_SEC    = PatternFill("solid", fgColor=C_SECURITY)
FILL_BLK    = PatternFill("solid", fgColor=C_BLOCKED)
FONT_WHITE  = Font(color="FFFFFF", bold=True, size=10)
FONT_BOLD   = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)
FONT_BLUE_B = Font(color=C_MED_BLUE, bold=True, size=10)
ALIGN_WRAP  = Alignment(wrap_text=True, vertical="top")
ALIGN_CENTER= Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def build_sheet(wb, sheet_name, api_code, api_name, method, path, tc_data):
    ws = wb.create_sheet(sheet_name)

    # Column widths
    widths = [7, 14, 9, 52, 48, 52, 62, 11, 16, 11, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Rows 1-4: header block ────────────────────────────────────────
    info_rows = [
        ("Mã API",      api_code),
        ("Tên API",     api_name),
        ("Phương thức", method),
        ("URL",         f"{BASE_URL}{path}"),
    ]
    for r, (lbl, val) in enumerate(info_rows, 1):
        c1 = ws.cell(row=r, column=1, value=lbl)
        c2 = ws.cell(row=r, column=2, value=val)
        c1.fill, c1.font, c1.alignment = FILL_INFO, FONT_WHITE, ALIGN_CENTER
        c2.fill = FILL_INFO
        c2.font = Font(color="FFFFFF", size=10)
        c2.alignment = ALIGN_WRAP
        ws.row_dimensions[r].height = 20
        for col in range(3, 12):
            ws.cell(row=r, column=col).fill = FILL_INFO

    # ── Row 5: column headers ────────────────────────────────────────
    col_headers = [
        "QC/AI", "Testcase ID", "Priority",
        "Nội Dung Test",
        "Điều Kiện / Dữ Liệu Test",
        "Các Bước Thực Hiện",
        "Kết Quả Mong Đợi (Expected Response)",
        "Round 1\nResult", "Round 1\nExecuted By",
        "Round 1\nBug ID", "Round 1\nRemark",
    ]
    for c, h in enumerate(col_headers, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.fill      = FILL_COL
        cell.font      = FONT_WHITE
        cell.alignment = ALIGN_CENTER
        cell.border    = thin_border()
    ws.row_dimensions[5].height = 30

    # ── Rows 6+: groups & TCs ────────────────────────────────────────
    row = 6
    counter = 0
    for item in tc_data:
        if item[0] == "G":
            # Group header
            cell = ws.cell(row=row, column=1, value=item[1])
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            cell.fill      = FILL_GROUP
            cell.font      = FONT_BOLD
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.border    = thin_border()
            ws.row_dimensions[row].height = 18
        else:
            # TC row: ("T", priority, title, condition, steps, expected, is_security, is_blocked)
            _, priority, title, condition, steps, expected = item[:6]
            is_sec = len(item) > 6 and item[6]
            is_blk = len(item) > 7 and item[7]

            counter += 1
            tc_id = f"{api_code}.{counter}"
            row_vals = ["AI", tc_id, priority, title, condition, steps, expected,
                        None, None, None, None]
            fill = FILL_BLK if is_blk else (FILL_SEC if is_sec else FILL_WHITE)
            for c, v in enumerate(row_vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.fill      = fill
                cell.font      = FONT_NORMAL
                cell.alignment = ALIGN_WRAP
                cell.border    = thin_border()
            ws.cell(row=row, column=1).alignment = ALIGN_CENTER
            ws.cell(row=row, column=2).alignment = ALIGN_CENTER
            ws.cell(row=row, column=3).alignment = ALIGN_CENTER
            ws.row_dimensions[row].height = 80
        row += 1

    # Freeze panes
    ws.freeze_panes = "A6"
    return counter


# ════════════════════════════════════════════════════════════════════
# TC DATA
# ════════════════════════════════════════════════════════════════════

# Helper: G = Group header, T = TC row
G = lambda label: ("G", label)
T = lambda pri, title, cond, steps, exp, sec=False, blk=False: ("T", pri, title, cond, steps, exp, sec, blk)

# ────────────────────────────────────────────────────────────────────
# API_16 — POST /public/v1/voucher/list
# ────────────────────────────────────────────────────────────────────
LIST_TCS = [
    G("Authentication / Authorization  —  X-Checkout-Token"),
    T("High",
      "Thiếu header X-Checkout-Token  [/voucher/list]",
      "Không truyền X-Checkout-Token trong header;\nAccept-Language: vi",
      "POST /public/v1/voucher/list\nHeaders: Accept-Language: vi\n(Không có X-Checkout-Token)",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\"\nerror.message = \"Phiên đặt hàng không hợp lệ. Vui lòng thử lại.\"\ndata = null"),
    T("High",
      "X-Checkout-Token = \"\" (chuỗi rỗng)  [/voucher/list]",
      "X-Checkout-Token: \"\" (empty string);\nAccept-Language: vi",
      "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: \"\"",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\"\ndata = null"),
    T("High",
      "X-Checkout-Token sai giá trị (invalid string)  [/voucher/list]",
      "X-Checkout-Token = \"invalid_token_xyz_000\"",
      "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: invalid_token_xyz_000",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nerror.message = \"Phiên đặt hàng không hợp lệ. Vui lòng thử lại.\"\ndata = null"),
    T("High",
      "X-Checkout-Token đã hết hạn (expired)  [/voucher/list]",
      "X-Checkout-Token = token hợp lệ về format nhưng đã expired",
      "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: <expired_token>",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
    T("High",
      "[SECURITY-INFERRED] X-Checkout-Token bị tampered (payload chỉnh sửa)  [/voucher/list]",
      "Decode token hợp lệ → sửa payload (vd: đổi checkout_id) → encode lại",
      "1. Lấy X-Checkout-Token hợp lệ\n2. Tamper payload (sửa checkout_id hoặc user_id)\n3. POST /public/v1/voucher/list với token đã sửa",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nChữ ký không khớp — không trả data voucher",
      True),
    T("High",
      "[SECURITY-INFERRED] Replay X-Checkout-Token từ checkout đã hoàn tất  [/voucher/list]",
      "X-Checkout-Token từ checkout session đã completed hoặc đã expired",
      "1. Lấy X-Checkout-Token từ checkout cũ (đã hoàn tất)\n2. POST /public/v1/voucher/list với token đó",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nToken replay bị từ chối — không trả data",
      True),
    G("Authentication / Authorization  —  Client-Id (Header Tùy Chọn)"),
    T("Medium",
      "Client-Id không truyền — API vẫn hoạt động bình thường  [/voucher/list]",
      "X-Checkout-Token hợp lệ; Client-Id: không truyền",
      "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: <valid>  (không có Client-Id)",
      "HTTP 200\nAPI xử lý bình thường\nClient-Id là header không bắt buộc — không ảnh hưởng kết quả"),
    T("Low",
      "Client-Id có giá trị — không ảnh hưởng kết quả business  [/voucher/list]",
      "X-Checkout-Token hợp lệ; Client-Id = \"test_client_001\"",
      "POST /public/v1/voucher/list 2 lần:\nLần 1: có Client-Id = \"test_client_001\"\nLần 2: không có Client-Id\nSo sánh kết quả",
      "HTTP 200 cả 2 lần\nKết quả data[] giống nhau\nClient-Id không ảnh hưởng danh sách voucher trả về"),
    G("Business Flow — Happy Path"),
    T("High",
      "Lấy danh sách voucher thành công — có ≥1 voucher khả dụng  [/voucher/list]",
      "X-Checkout-Token hợp lệ (checkout đã chọn PTTT, có voucher phù hợp)",
      "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: <valid_with_vouchers>",
      "HTTP 200\nsuccess = true\ndata[] có ≥1 item\nMỗi item có: voucher_code (non-null string), description, to_date (format dd/MM/yyyy), voucher_type (\"General\"/\"Individual\"/\"Unknown\")"),
    T("High",
      "Không có voucher khả dụng — data trả về mảng rỗng  [/voucher/list]",
      "X-Checkout-Token hợp lệ nhưng không có voucher nào khớp với context checkout",
      "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: <valid_no_voucher>",
      "HTTP 200\nsuccess = true\ndata = [] (mảng rỗng)\nKhông có error"),
    T("Medium",
      "Cấu trúc từng item trong data[] đầy đủ fields bắt buộc  [/voucher/list]",
      "X-Checkout-Token hợp lệ với context có ≥1 voucher",
      "POST /public/v1/voucher/list\nDuyệt từng item trong data[]",
      "HTTP 200\nMỗi item có đủ keys: voucher_code, description, note, to_date, register_type_id, voucher_type, policy_group_id, apply_type_id, promotion_type_id\nvoucher_code ≠ null"),
    G("Business Flow — Error Cases"),
    T("High",
      "Checkout không tồn tại — CHECKOUT_NOT_FOUND  [/voucher/list]",
      "X-Checkout-Token hợp lệ về format nhưng checkout session không tồn tại/đã xóa",
      "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: <token_nonexistent_checkout>",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\"\nerror.message = \"Không tìm thấy phiên đặt hàng.\"\ndata = null"),
    T("High",
      "Chưa chọn hình thức thanh toán — CHECKOUT_PAYMENT_REQUIRED  [/voucher/list]",
      "Checkout session tồn tại nhưng chưa chọn phương thức thanh toán",
      "POST /public/v1/voucher/list\nHeaders: X-Checkout-Token: <checkout_no_payment>",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_PAYMENT_REQUIRED\"\nerror.message = \"Vui lòng chọn phương thức thanh toán.\"\ndata = null"),
    G("Data Integrity & Response Structure"),
    T("Medium",
      "Response 200 có cấu trúc đầy đủ: success, error, meta, data  [/voucher/list]",
      "X-Checkout-Token hợp lệ",
      "POST /public/v1/voucher/list\nInspect toàn bộ response body",
      "HTTP 200\nsuccess = true\nerror = null\nmeta.request_id: string (có thể rỗng)\nmeta.trace_id: string (có thể rỗng)\nmeta.timestamp: string (datetime format)\nmeta.pagination = null\ndata: array"),
]

# ────────────────────────────────────────────────────────────────────
# API_17 — POST /public/v1/voucher/content
# ────────────────────────────────────────────────────────────────────
CONTENT_TCS = [
    G("Authentication / Authorization  —  X-Checkout-Token"),
    T("High",
      "Thiếu header X-Checkout-Token  [/voucher/content]",
      "Không truyền X-Checkout-Token;\nbody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
      "POST /public/v1/voucher/content\nHeaders: (không có X-Checkout-Token)\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\"\ndata = null"),
    T("High",
      "X-Checkout-Token không hợp lệ (invalid)  [/voucher/content]",
      "X-Checkout-Token = \"invalid_xyz\";\nbody hợp lệ",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: invalid_xyz\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
    T("High",
      "X-Checkout-Token đã hết hạn  [/voucher/content]",
      "X-Checkout-Token = expired token;\nbody hợp lệ",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <expired>\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
    T("High",
      "[SECURITY-INFERRED] X-Checkout-Token bị tampered  [/voucher/content]",
      "Token payload bị chỉnh sửa (checkout_id thay đổi)",
      "1. Tamper payload của X-Checkout-Token\n2. POST /public/v1/voucher/content với token đã sửa\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nChữ ký không khớp — không trả content",
      True),
    T("High",
      "[SECURITY-INFERRED] Replay X-Checkout-Token cũ  [/voucher/content]",
      "X-Checkout-Token của checkout đã completed",
      "1. Lấy X-Checkout-Token từ checkout cũ\n2. POST /public/v1/voucher/content với token đó",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nToken replay bị từ chối",
      True),
    G("Authentication / Authorization  —  Client-Id (Header Tùy Chọn)"),
    T("Medium",
      "Client-Id không truyền — API hoạt động bình thường  [/voucher/content]",
      "X-Checkout-Token hợp lệ; không có Client-Id;\nbody có voucher_code hợp lệ",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <valid>  (không có Client-Id)\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
      "HTTP 200\nAPI trả content bình thường\nClient-Id không bắt buộc"),
    G("Validation — Required Fields (Request Body)"),
    T("High",
      "Thiếu field voucher_code trong body  [/voucher/content]",
      "X-Checkout-Token hợp lệ;\nbody: {}",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <valid>\nBody: {}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\nerror.message = \"Mã voucher là bắt buộc.\"\ndata = null"),
    T("High",
      "voucher_code = null  [/voucher/content]",
      "X-Checkout-Token hợp lệ;\nbody: {\"voucher_code\": null}",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": null}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\ndata = null"),
    T("High",
      "voucher_code = \"\" (chuỗi rỗng — boundary min=1)  [/voucher/content]",
      "X-Checkout-Token hợp lệ;\nbody: {\"voucher_code\": \"\"}",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"\"}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\ndata = null"),
    G("Business Flow — Happy Path"),
    T("High",
      "Lấy nội dung voucher thành công — content1..content6 trả về  [/voucher/content]",
      "X-Checkout-Token hợp lệ;\nvoucher_code = \"CA21060100KTHIETBIKHOFG039\" (có nội dung)",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
      "HTTP 200\nsuccess = true\ndata.voucher_code = \"CA21060100KTHIETBIKHOFG039\"\ndata.content1 ≠ null (tiêu đề voucher)\ndata.content2..content6 trả về (có thể null cho field chưa có nội dung)"),
    G("Business Flow — Error Cases"),
    T("High",
      "Voucher không hợp lệ hoặc đã hết hạn — VOUCHER_INVALID  [/voucher/content]",
      "X-Checkout-Token hợp lệ;\nvoucher_code đã hết hạn hoặc không tồn tại",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"EXPIRED_CODE_999\"}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\"\nerror.message = \"Mã voucher không hợp lệ hoặc đã hết hạn.\"\ndata = null"),
    T("High",
      "Checkout không tồn tại — CHECKOUT_NOT_FOUND  [/voucher/content]",
      "X-Checkout-Token trỏ đến checkout không tồn tại;\nvoucher_code hợp lệ",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <token_nonexistent>\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\"\ndata = null"),
    G("Security [SECURITY-INFERRED]"),
    T("High",
      "[SECURITY-INFERRED] XSS payload trong voucher_code  [/voucher/content]",
      "X-Checkout-Token hợp lệ;\nvoucher_code = \"<script>alert(document.cookie)</script>\"",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"<script>alert(document.cookie)</script>\"}",
      "HTTP 400\nsuccess = false\nXSS payload không được reflect/execute trong response body\nKhông có HTML rendering trong response",
      True),
    T("High",
      "[SECURITY-INFERRED] SQL Injection trong voucher_code  [/voucher/content]",
      "X-Checkout-Token hợp lệ;\nvoucher_code = \"'; DROP TABLE vouchers; --\"",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"'; DROP TABLE vouchers; --\"}",
      "HTTP 400\nsuccess = false\nKhông có SQL error lộ ra\nKhông crash server (không HTTP 500)",
      True),
    T("Medium",
      "[SECURITY-INFERRED] voucher_code quá dài (>200 ký tự) — không gây crash  [/voucher/content]",
      "X-Checkout-Token hợp lệ;\nvoucher_code = chuỗi 201 ký tự 'A'",
      "POST /public/v1/voucher/content\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"AAAA...\" (201 ký tự)}",
      "HTTP 400\nsuccess = false\nKhông HTTP 500\nKhông crash server",
      True),
]

# ────────────────────────────────────────────────────────────────────
# API_18 — POST /public/v1/voucher/apply
# ────────────────────────────────────────────────────────────────────
APPLY_TCS = [
    G("Authentication / Authorization  —  X-Checkout-Token"),
    T("High",
      "Thiếu header X-Checkout-Token  [/voucher/apply]",
      "Không truyền X-Checkout-Token;\nbody: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}]}",
      "POST /public/v1/voucher/apply\nHeaders: (không có X-Checkout-Token)\nBody: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}]}",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\"\ndata = null"),
    T("High",
      "X-Checkout-Token không hợp lệ (invalid)  [/voucher/apply]",
      "X-Checkout-Token = \"invalid_xyz\";\nbody hợp lệ",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: invalid_xyz\nBody: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}]}",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
    T("High",
      "X-Checkout-Token đã hết hạn  [/voucher/apply]",
      "X-Checkout-Token = expired token;\nbody hợp lệ",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <expired>\nBody: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}]}",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
    T("High",
      "[SECURITY-INFERRED] X-Checkout-Token bị tampered  [/voucher/apply]",
      "Token payload bị chỉnh sửa (checkout_id thay đổi)",
      "1. Tamper payload của X-Checkout-Token\n2. POST /public/v1/voucher/apply với token đã sửa",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nKhông apply voucher",
      True),
    T("High",
      "[SECURITY-INFERRED] Replay X-Checkout-Token cũ  [/voucher/apply]",
      "X-Checkout-Token từ checkout đã completed",
      "1. Lấy X-Checkout-Token từ checkout cũ\n2. POST /public/v1/voucher/apply với token đó",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nToken replay bị từ chối",
      True),
    T("High",
      "[SECURITY-INFERRED] X-Checkout-Token của checkout người dùng khác (IDOR)  [/voucher/apply]",
      "X-Checkout-Token là token của user B;\nvouchers là voucher của user A",
      "1. Lấy X-Checkout-Token hợp lệ của user B\n2. POST /public/v1/voucher/apply với voucher của user A",
      "HTTP 401 hoặc 400\nsuccess = false\nTừ chối — không cho phép apply voucher cho checkout của user khác",
      True),
    G("Authentication / Authorization  —  Client-Id (Header Tùy Chọn)"),
    T("Medium",
      "Client-Id không truyền — API hoạt động bình thường  [/voucher/apply]",
      "X-Checkout-Token hợp lệ; không có Client-Id;\nbody có voucher hợp lệ",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <valid>\nBody: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}]}",
      "HTTP 200\nsuccess = true\nClient-Id không bắt buộc"),
    G("Validation — Required Fields (vouchers[].voucher_code)"),
    T("High",
      "vouchers[].voucher_code bị thiếu trong item  [/voucher/apply]",
      "X-Checkout-Token hợp lệ;\nbody: {\"vouchers\": [{\"voucher_type\": \"General\"}]} (thiếu voucher_code)",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <valid>\nBody: {\"vouchers\": [{\"voucher_type\": \"General\"}]}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\ndata = null"),
    T("High",
      "vouchers[].voucher_code = null  [/voucher/apply]",
      "X-Checkout-Token hợp lệ;\nbody: {\"vouchers\": [{\"voucher_code\": null, \"voucher_type\": \"General\"}]}",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <valid>\nBody: {\"vouchers\": [{\"voucher_code\": null, \"voucher_type\": \"General\"}]}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\ndata = null"),
    T("High",
      "vouchers[].voucher_code = \"\" (rỗng — boundary min=1)  [/voucher/apply]",
      "X-Checkout-Token hợp lệ;\nbody: {\"vouchers\": [{\"voucher_code\": \"\", \"voucher_type\": \"General\"}]}",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <valid>\nBody: {\"vouchers\": [{\"voucher_code\": \"\", \"voucher_type\": \"General\"}]}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\ndata = null"),
    G("Business Flow — Happy Path"),
    T("High",
      "Áp dụng 1 voucher General thành công  [/voucher/apply]",
      "X-Checkout-Token hợp lệ (checkout đã chọn PTTT);\nvoucher_code hợp lệ loại General",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <valid>\nBody: {\n  \"vouchers\": [{\n    \"voucher_code\": \"CA21060100KTHIETBIKHOFG039\",\n    \"voucher_type\": \"General\"\n  }]\n}",
      "HTTP 200\nsuccess = true\ndata[] có 1 item\ndata[0].voucher_code = \"CA21060100KTHIETBIKHOFG039\"\ndata[0].promotion_id ≠ null\ndata[0].discount_value = 0 (chưa calculate)"),
    T("High",
      "Áp dụng 1 voucher Individual thành công  [/voucher/apply]",
      "X-Checkout-Token hợp lệ;\nvoucher_code hợp lệ loại Individual",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <valid>\nBody: {\n  \"vouchers\": [{\n    \"voucher_code\": \"CA23039F50GOLD\",\n    \"voucher_type\": \"Individual\"\n  }]\n}",
      "HTTP 200\nsuccess = true\ndata[0].voucher_code = \"CA23039F50GOLD\"\ndata[0].discount_value = 0"),
    T("High",
      "Áp dụng nhiều voucher cùng lúc (General + Individual)  [/voucher/apply]",
      "X-Checkout-Token hợp lệ;\n2 voucher hợp lệ (1 General, 1 Individual)",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <valid>\nBody: {\n  \"vouchers\": [\n    {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"},\n    {\"voucher_code\": \"CA23039F50GOLD\", \"voucher_type\": \"Individual\"}\n  ]\n}",
      "HTTP 200\nsuccess = true\ndata[] có 2 items\nMỗi item có discount_value = 0"),
    T("High",
      "Xác nhận discount_value = 0 ngay sau apply (business rule)  [/voucher/apply]",
      "X-Checkout-Token hợp lệ;\nvoucher hợp lệ (discount thực tế > 0 về mặt nghiệp vụ)",
      "POST /public/v1/voucher/apply với voucher hợp lệ\nKiểm tra discount_value và discount_ex_vat_value trong response",
      "HTTP 200\ndata[*].discount_value = 0\ndata[*].discount_ex_vat_value = 0\ndata[*].original_discount_value > 0 (giá trị gốc của voucher)\n[Rule: Discount=0 vì sẽ được tính toán sau khi gọi calculate API]"),
    T("Medium",
      "Cấu trúc data[].applies[] có đủ fields bắt buộc  [/voucher/apply]",
      "X-Checkout-Token hợp lệ;\nvoucher hợp lệ có applies",
      "POST /public/v1/voucher/apply với voucher hợp lệ\nKiểm tra structure data[].applies[]",
      "HTTP 200\ndata[0].applies[] có ít nhất 1 item\nMỗi applies[].service_id ≠ null\napplies[].discount ≥ 0\napplies[].dismonth ≥ 0"),
    G("Business Flow — Gỡ Bỏ Voucher"),
    T("High",
      "Gỡ bỏ tất cả voucher — gửi vouchers = [] (empty array)  [/voucher/apply]",
      "Checkout đang có voucher đã áp dụng;\ngửi vouchers=[]",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <valid_with_applied_voucher>\nBody: {\"vouchers\": []}",
      "HTTP 200\nsuccess = true\ndata = [] (mảng rỗng)\nToàn bộ voucher đã áp dụng bị gỡ bỏ"),
    T("Medium",
      "[CLARY-002 Pending BA] vouchers = null — behavior cần xác nhận  [/voucher/apply]",
      "Gửi vouchers=null;\nCLARY-002 chưa resolve",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <valid>\nBody: {\"vouchers\": null}",
      "[BLOCKED – cần confirm BA: vouchers=null có hành vi giống vouchers=[] (gỡ tất cả) hay trả lỗi validation?]",
      False, True),
    G("Business Flow — Error Cases"),
    T("High",
      "Mã voucher trùng lặp trong danh sách — VOUCHER_CODE_DUPLICATE_400  [/voucher/apply]",
      "X-Checkout-Token hợp lệ;\nvouchers chứa 2 items có voucher_code giống nhau",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <valid>\nBody: {\n  \"vouchers\": [\n    {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"},\n    {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}\n  ]\n}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_DUPLICATE_400\"\nerror.message = \"Danh sách mã voucher bị trùng lặp. Vui lòng kiểm tra lại.\"\ndata = null"),
    T("High",
      "Rà soát điều kiện áp dụng voucher thất bại — VOUCHER_RECHECK_FAILED  [/voucher/apply]",
      "X-Checkout-Token hợp lệ;\nvoucher không đáp ứng điều kiện (sai hạng KH, sản phẩm không áp dụng...)",
      "POST /public/v1/voucher/apply với voucher không đủ điều kiện",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_RECHECK_FAILED\"\nerror.message = \"Rà soát voucher thất bại. Vui lòng kiểm tra lại điều kiện áp dụng.\"\ndata = null"),
    T("High",
      "Voucher không hợp lệ hoặc đã hết hạn — VOUCHER_INVALID  [/voucher/apply]",
      "X-Checkout-Token hợp lệ;\nvoucher_code đã hết hạn",
      "POST /public/v1/voucher/apply\nBody: {\"vouchers\": [{\"voucher_code\": \"EXPIRED_CODE\", \"voucher_type\": \"General\"}]}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\"\nerror.message = \"Mã voucher không hợp lệ hoặc đã hết hạn.\"\ndata = null"),
    T("High",
      "Chưa chọn hình thức thanh toán — CHECKOUT_PAYMENT_REQUIRED  [/voucher/apply]",
      "Checkout tồn tại nhưng chưa chọn PTTT;\nvoucher hợp lệ",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <checkout_no_payment>",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_PAYMENT_REQUIRED\"\nerror.message = \"Vui lòng chọn phương thức thanh toán.\"\ndata = null"),
    T("High",
      "Checkout không tồn tại — CHECKOUT_NOT_FOUND  [/voucher/apply]",
      "X-Checkout-Token trỏ đến checkout không tồn tại",
      "POST /public/v1/voucher/apply\nHeaders: X-Checkout-Token: <token_nonexistent>",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\"\ndata = null"),
]

# ────────────────────────────────────────────────────────────────────
# API_19 — POST /public/v1/voucher/check
# ────────────────────────────────────────────────────────────────────
CHECK_TCS = [
    G("Authentication / Authorization  —  X-Checkout-Token"),
    T("High",
      "Thiếu header X-Checkout-Token  [/voucher/check]",
      "Không truyền X-Checkout-Token;\nbody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}",
      "POST /public/v1/voucher/check\nHeaders: (không có X-Checkout-Token)\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\"\ndata = null"),
    T("High",
      "X-Checkout-Token không hợp lệ (invalid)  [/voucher/check]",
      "X-Checkout-Token = \"invalid_xyz\";\nbody hợp lệ",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: invalid_xyz\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
    T("High",
      "X-Checkout-Token đã hết hạn  [/voucher/check]",
      "X-Checkout-Token = expired token;\nbody hợp lệ",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <expired>\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
    T("High",
      "[SECURITY-INFERRED] X-Checkout-Token bị tampered  [/voucher/check]",
      "Token payload bị chỉnh sửa",
      "1. Tamper payload của X-Checkout-Token\n2. POST /public/v1/voucher/check với token đã sửa",
      "HTTP 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"",
      True),
    G("Authentication / Authorization  —  Client-Id (Header Tùy Chọn)"),
    T("Medium",
      "Client-Id không truyền — API hoạt động bình thường  [/voucher/check]",
      "X-Checkout-Token hợp lệ; không có Client-Id;\nbody có voucher_code hợp lệ",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>  (không có Client-Id)\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}",
      "HTTP 200\nAPI hoạt động bình thường\nClient-Id không ảnh hưởng kết quả"),
    G("Validation — Required Fields (voucher_code)"),
    T("High",
      "Thiếu field voucher_code trong body  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nbody: {}",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\nerror.message = \"Mã voucher là bắt buộc.\"\ndata = null"),
    T("High",
      "voucher_code = null  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nbody: {\"voucher_code\": null}",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": null}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\ndata = null"),
    T("High",
      "voucher_code = \"\" (chuỗi rỗng — boundary min=1)  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nbody: {\"voucher_code\": \"\"}",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"\"}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\ndata = null"),
    G("Business Flow — Happy Path"),
    T("High",
      "Check voucher General hợp lệ — is_valid = true  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nvoucher_code hợp lệ loại General",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}",
      "HTTP 200\nsuccess = true\ndata.is_valid = true\ndata.message = \"Voucher is valid.\""),
    T("High",
      "Check voucher Individual hợp lệ — is_valid = true  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nvoucher_code hợp lệ loại Individual",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"CA23039F50GOLD\", \"voucher_type\": \"Individual\"}",
      "HTTP 200\nsuccess = true\ndata.is_valid = true\ndata.message ≠ null"),
    T("Medium",
      "Không truyền voucher_type (Optional) — API vẫn trả kết quả  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nbody chỉ có voucher_code, không có voucher_type",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
      "HTTP 200\nsuccess = true\ndata.is_valid = true hoặc xử lý phù hợp\nvoucher_type là Optional — không gây lỗi khi thiếu"),
    G("Business Flow — voucher_type Logic"),
    T("Medium",
      "voucher_type = General, actual type = General → hợp lệ  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nvoucher_code là General;\nvoucher_type = \"General\" (khớp)",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}",
      "HTTP 200\nsuccess = true\ndata.is_valid = true"),
    T("Medium",
      "voucher_type = Individual, actual type = Individual → hợp lệ  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nvoucher_code là Individual;\nvoucher_type = \"Individual\" (khớp)",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"CA23039F50GOLD\", \"voucher_type\": \"Individual\"}",
      "HTTP 200\nsuccess = true\ndata.is_valid = true"),
    T("High",
      "voucher_type = General nhưng actual type = Individual → VOUCHER_INVALID  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nvoucher_code là Individual nhưng gửi voucher_type=\"General\" (không khớp)\n[CLARY-005 Resolved: → VOUCHER_INVALID 400]",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"CA23039F50GOLD\", \"voucher_type\": \"General\"}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\"\ndata = null\n[CLARY-005 Resolved]"),
    T("High",
      "voucher_type = Individual nhưng actual type = General → VOUCHER_INVALID  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nvoucher_code là General nhưng gửi voucher_type=\"Individual\" (không khớp)\n[CLARY-005 Resolved]",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"Individual\"}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\"\ndata = null\n[CLARY-005 Resolved]"),
    G("Business Flow — Error Cases"),
    T("High",
      "Check voucher không hợp lệ / đã hết hạn — VOUCHER_INVALID  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nvoucher_code đã hết hạn hoặc không tồn tại",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"EXPIRED_OR_INVALID_CODE\", \"voucher_type\": \"General\"}",
      "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\"\nerror.message = \"Mã voucher không hợp lệ hoặc đã hết hạn.\"\ndata = null"),
    T("High",
      "Checkout không tồn tại — CHECKOUT_NOT_FOUND  [/voucher/check]",
      "X-Checkout-Token trỏ đến checkout không tồn tại;\nvoucher_code hợp lệ",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <token_nonexistent>\nBody: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}",
      "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\"\ndata = null"),
    G("Security [SECURITY-INFERRED]"),
    T("High",
      "[SECURITY-INFERRED] XSS payload trong voucher_code  [/voucher/check]",
      "X-Checkout-Token hợp lệ;\nvoucher_code = \"<script>alert(1)</script>\"",
      "POST /public/v1/voucher/check\nHeaders: X-Checkout-Token: <valid>\nBody: {\"voucher_code\": \"<script>alert(1)</script>\", \"voucher_type\": \"General\"}",
      "HTTP 400\nsuccess = false\nXSS payload không reflect trong response",
      True),
]

# ════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

totals = {}
totals["API_16"] = build_sheet(wb, "Danh sách Voucher",
    "API_16", "Lấy danh sách Voucher khả dụng", "POST",
    "/public/v1/voucher/list", LIST_TCS)

totals["API_17"] = build_sheet(wb, "Nội dung Voucher",
    "API_17", "Lấy nội dung hiển thị Voucher", "POST",
    "/public/v1/voucher/content", CONTENT_TCS)

totals["API_18"] = build_sheet(wb, "Áp dụng Voucher",
    "API_18", "Áp dụng / Gỡ bỏ eVoucher vào checkout", "POST",
    "/public/v1/voucher/apply", APPLY_TCS)

totals["API_19"] = build_sheet(wb, "Kiểm tra Voucher",
    "API_19", "Kiểm tra tính hợp lệ Voucher", "POST",
    "/public/v1/voucher/check", CHECK_TCS)

wb.save(OUTPUT)
total_tc = sum(totals.values())
import sys, io
out = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
out.write(f"[OK] Saved: {OUTPUT}\n")
for k, v in totals.items():
    out.write(f"   {k}: {v} TCs\n")
out.write(f"   TOTAL: {total_tc} TCs\n")
out.flush()
