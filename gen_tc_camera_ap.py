# -*- coding: utf-8 -*-
"""Sinh 2 sheet TC Checkout: Camera (TC_CAMERA) + AP (TC_AP).
Thêm vào file 03_test-cases/functional/chucnang_checkout/AI_ISC_ecom-pdh_v1.1_TC_checkout_v1.0.xlsx
Theo chiến lLược gen-testcase-checkout-service: chỉ viết TC đặc thù dịch vụ, KHÔNG clone TC_CKCOMMON.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

PATH = 'ecom-pdh/03_test-cases/functional/chucnang_checkout/AI_ISC_ecom-pdh_v1.1_TC_checkout_v1.0.xlsx'

C_HDR = "4472C4"; C_LVL1 = "A4C2F4"; C_LVL2 = "A9D08E"; C_BLOCK = "FCE4D6"  # cam nhạt — dòng BLOCKED
WIDTHS = dict(A=6, B=14, C=10, D=45, E=45, F=50, G=58, H=8, I=12, J=15, K=12, L=20)
HDR7 = {1:"QC/AI",2:"Testcase ID",3:"Mức Độ Ưu Tiên",4:"Nội Dung Test",
        5:"Điều Kiện / Dữ Liệu Test",6:"Các Bước Thực Hiện",7:"Kết Quả Mong Đợi",
        8:"Có thể\nTự Động Hóa",9:"Round 1"}
HDR8 = {9:"Kết Quả Thực Hiện",10:"Người Thực Hiện",11:"ID Bugs",12:"Ghi Chú"}

def hfill(c): return PatternFill("solid", fgColor=c)

def build_sheet(wb, sheet_name, func_id, func_name, rows):
    """rows: list of dicts {type:'grp'|'note'|'tc', ...}"""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for col,w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    # Header meta
    ws['C3'] = "Mã chức năng:"; ws['D3'] = func_id; ws['D3'].font = Font(bold=True)
    ws['C3'].font = Font(bold=True)
    ws['C4'] = "Tên chức năng:"; ws['D4'] = func_name; ws['D4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)
    # Column header rows 7-8
    whdr = Font(bold=True, color="FFFFFF", sz=11)
    ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col,val in HDR7.items():
        cell = ws.cell(7, col, val); cell.fill = hfill(C_HDR); cell.font = whdr; cell.alignment = ac
    for col,val in HDR8.items():
        cell = ws.cell(8, col, val); cell.fill = hfill(C_HDR); cell.font = whdr; cell.alignment = ac
    for col in range(1,9):
        ws.merge_cells(start_row=7, start_column=col, end_row=8, end_column=col)
    ws.merge_cells('I7:L7')
    ws.row_dimensions[7].height = 30; ws.row_dimensions[8].height = 20
    ws.freeze_panes = 'A9'
    # Body
    r = 9
    al_top_l = Alignment(horizontal="left", vertical="top", wrap_text=True)
    al_top_c = Alignment(horizontal="center", vertical="top", wrap_text=True)
    al_grp = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for item in rows:
        t = item['type']
        if t in ('grp','note'):
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
            cell = ws.cell(r, 2, item['label'])
            cell.fill = hfill(C_LVL2 if t=='grp' else C_LVL1)
            cell.font = Font(bold=True); cell.alignment = al_grp
            ws.row_dimensions[r].height = 18
        else:
            ws.cell(r,1,"AI").font = Font(bold=True, color="7030A0")
            ws.cell(r,1).alignment = al_top_c
            ws.cell(r,2, f'=IF(D{r}="","",$D$3&"."&COUNTA($D$10:D{r})&"")').alignment = al_top_c
            ws.cell(r,3, item['pri']).alignment = al_top_c
            tc = ws.cell(r,4, item['title']); tc.font = Font(bold=True); tc.alignment = al_top_l
            ws.cell(r,5, item['pre']).alignment = al_top_l
            ws.cell(r,6, item['steps']).alignment = al_top_l
            ws.cell(r,7, item['exp']).alignment = al_top_l
            ws.cell(r,8, item['auto']).alignment = al_top_c
            if item.get('note'):
                ws.cell(r,12, item['note']).alignment = al_top_l
                if str(item['note']).startswith('BLOCKED'):
                    for col in range(1,13):
                        ws.cell(r,col).fill = hfill(C_BLOCK)
        r += 1
    return ws

def tc(pri,title,pre,steps,exp,auto,note=None):
    return dict(type='tc',pri=pri,title=title,pre=pre,steps=steps,exp=exp,auto=auto,note=note)
def grp(label): return dict(type='grp',label=label)
def note(label): return dict(type='note',label=label)

REFER = note("Lưu ý: Field validation (Họ tên, SĐT, Email, dropdown địa chỉ Tỉnh/Phường/Đường, "
             "popup Địa chỉ hành chính cũ, danh sách Phương thức thanh toán, luồng OTP) dùng chung — "
             "xem sheet Checkout_Common (TC_CKCOMMON). Sheet này chỉ kiểm tra phần ĐẶC THÙ dịch vụ.")

# ============ CAMERA ============
camera_rows = [
    REFER,
    grp("Nhóm 1: Hiển thị tổng thể & Điều hướng vào Checkout (chu kỳ + số lượng)"),
    tc("High","Kiểm tra hiển thị tổng thể màn Checkout Camera (2 bước)",
       "Đã chọn gói Camera ở màn Chi tiết",
       "1. Quan sát tiến trình và bố cục màn checkout Camera",
       'Header 2 bước: "Thanh toán" (xanh dương) → "Hoàn tất đơn hàng" (xám); hiển thị đủ block: Sản phẩm dịch vụ đã chọn, Thông tin cá nhân, Địa chỉ lắp đặt, Phương thức thanh toán, Thông tin khách hàng, Thông tin thanh toán',"N"),
    tc("High","Kiểm tra chọn chu kỳ + số lượng rồi Mua ngay điều hướng sang Checkout",
       "Ở màn Chi tiết gói Camera",
       "1. Chọn chu kỳ bất kỳ + số lượng\n2. Click button \"Mua ngay\"",
       "Điều hướng sang màn Thanh toán; load đúng tên gói, chu kỳ, số lượng, số tiền như đã chọn ở Chi tiết","Y"),
    tc("Medium","Kiểm tra đổi số lượng thì số tiền cập nhật đúng trên checkout",
       "Ở Chi tiết chọn số lượng > 1",
       "1. Chọn số lượng > 1\n2. Mua ngay sang checkout\n3. Đối chiếu số tiền",
       "Số tiền trên checkout = đơn giá × số lượng theo chu kỳ đã chọn (itemized: dòng Camera + dòng Gói Cloud)","Y"),
    grp("Nhóm 2: Block Sản phẩm & Thông tin thanh toán (đặc thù Camera)"),
    tc("High","Kiểm tra Block Sản phẩm dịch vụ đã chọn load đúng thông tin Camera",
       "Đang ở màn Thanh toán Camera",
       "1. Quan sát Block \"Sản phẩm dịch vụ đã chọn\"",
       "Load đúng tên gói (vd Camera SE S2), dòng Cloud (vd Cloud 3 ngày), chu kỳ (vd 6 tháng), số lượng (x1), số tiền","Y"),
    tc("High","Kiểm tra Block Thông tin thanh toán itemized + Cần thanh toán",
       "Đang ở màn Thanh toán Camera",
       "1. Quan sát Block \"Thông tin thanh toán\"\n2. Đối chiếu dòng Cần thanh toán",
       'Itemized đúng (vd "Camera SE S2 - 1 cái: 600.000đ", "Gói Cloud 3D - 6 tháng: 240.000đ"); "Cần thanh toán" = tổng tiền (vd 840.000đ)',"Y"),
    grp("Nhóm 3: Phương thức thanh toán & Thông tin khách hàng (đặc thù Camera)"),
    tc("Medium","Kiểm tra Camera CÓ phương thức COD + online theo QLCS",
       "Gói Camera cấu hình COD + online trên QLCS",
       "1. Quan sát Block Phương thức thanh toán",
       'Hiển thị COD "Thanh toán tại nhà" (kèm CTKM "Giảm trực tiếp 50% giá trị đơn hàng tối đa 200.000 VND") + các PTTT online theo QLCS; mỗi PTTT có badge ưu đãi (khác UltraFast: Camera CÓ COD). Cơ chế danh sách PTTT xem TC_CKCOMMON',"Y"),
    tc("Medium","Kiểm tra Block Thông tin khách hàng auto-load + Thời gian lắp đặt",
       "Đã nhập Thông tin cá nhân + Địa chỉ lắp đặt",
       "1. Quan sát Block Thông tin khách hàng (cột phải)",
       '[BLOCKED – cần confirm: "Thời gian lắp đặt" (vd Thứ 2, 02/01/2024 10:00-11:10) được set ở đâu/khi nào?]. Phần auto-load Họ tên/SĐT/Địa chỉ đúng theo dữ liệu đã nhập',"N",
       "BLOCKED – CLA-CAMERA-001 (nguồn data Thời gian lắp đặt)"),
    grp("Nhóm 4: Button Thanh toán & Luồng thanh toán"),
    tc("High","Kiểm tra Thanh toán khi chính sách không còn active trên QLCS",
       "Gói/chính sách Camera không còn active trên QLCS",
       "1. Click button \"Thanh toán\"",
       "Báo lỗi chính sách không còn hiệu lực; KHÔNG thực hiện thanh toán","Y"),
    tc("High","Kiểm tra Thanh toán khi data hợp lệ + chính sách active",
       "Nhập đủ trường bắt buộc hợp lệ, chính sách active",
       "1. Click button \"Thanh toán\"",
       "Thực hiện luồng thanh toán theo PTTT đã chọn (COD → sang B3 luôn; Online → điều hướng cổng 3rd party)","Y"),
    grp("Nhóm 5: Bước Hoàn tất đơn hàng"),
    tc("High","Kiểm tra hoàn tất đơn hàng với PTTT = COD",
       "Chọn PTTT = COD, click Thanh toán",
       "1. Hoàn tất đặt hàng với PTTT = COD\n2. Quan sát màn Hoàn tất",
       'Màn Hoàn tất: trạng thái "Chưa thanh toán" + nội dung "Đơn hàng đã đăng ký thành công. Kỹ thuật viên FPT sẽ liên hệ triển khai dịch vụ trong 8h-12h. Mọi thắc mắc vui lòng liên hệ 1900 6600..."',"Y"),
    tc("High","Kiểm tra hoàn tất đơn hàng với PTTT Online thanh toán thành công",
       "Chọn PTTT Online, thanh toán thành công",
       "1. Hoàn tất thanh toán Online thành công\n2. Quan sát màn Hoàn tất",
       'Bước "Hoàn tất đơn hàng" header màu xanh lá; trạng thái "Đã thanh toán" + nội dung thông báo thành công',"N"),
    tc("High","Kiểm tra thanh toán Online thất bại quay về màn Thanh toán",
       "Chọn PTTT Online, thanh toán thất bại",
       "1. Thực hiện thanh toán Online thất bại",
       "Quay về màn Thanh toán giữ nguyên thông tin đã chọn; CHỈ cho thay đổi PTTT + nhập mã ưu đãi, các trường còn lại disable","N"),
    grp("Nhóm 6: Mã ưu đãi & Hạng mục Deferred"),
    tc("Medium","Kiểm tra áp dụng mã ưu đãi / Chọn ưu đãi cho Camera",
       'Đang ở màn Thanh toán, có ô "Nhập mã khuyến mãi" + "Chọn ưu đãi"',
       "1. Nhập mã ưu đãi / chọn ưu đãi\n2. Click \"Áp dụng\"",
       '[BLOCKED – cần confirm: voucher/mã ưu đãi cho Camera chưa implement (như UltraFast SC-DANGKYUF-015)]. Kỳ vọng: trừ tiền vào "Cần thanh toán"',"N",
       "BLOCKED – CLA-CAMAP-001 (voucher chưa implement)"),
    tc("Medium","Kiểm tra chọn loại địa chỉ Chung cư (radio Nhà riêng/Chung cư)",
       "Đang ở Block Địa chỉ lắp đặt",
       "1. Chọn radio \"Chung cư\"",
       "[BLOCKED – Deferred: Chung cư chưa phát triển ở version hiện tại (CLA-CKCOMMON-006), áp dụng chung mọi dịch vụ]","N",
       "BLOCKED/Deferred – CLA-CKCOMMON-006 (Chung cư)"),
    grp("Nhóm 7: Navigation & Mobile"),
    tc("Low","Kiểm tra button Quay lại điều hướng về màn Chi tiết gói",
       "Đang ở màn Thanh toán Camera",
       "1. Click button \"Quay lại\"",
       "Điều hướng về màn Chi tiết gói Camera","Y"),
    tc("Medium","Kiểm tra hiển thị luồng Checkout Camera trên mobile (≤ 768px)",
       "Mở checkout Camera trên viewport mobile",
       "1. Quan sát layout 2 bước + các block trên mobile",
       "Các bước và block hiển thị đúng, không vỡ layout, thao tác được trên mobile","N"),
]

# ============ AP ============
ap_rows = [
    REFER,
    note("Lưu ý: AP giống Camera nhưng B1 CHỈ chọn số lượng (không có chu kỳ). Không có mockup riêng — "
         "tham chiếu layout mockup Camera (camera.png)."),
    grp("Nhóm 1: Hiển thị & Điều hướng vào Checkout (số lượng — không chu kỳ)"),
    tc("High","Kiểm tra hiển thị tổng thể màn Checkout AP (2 bước)",
       "Đã chọn gói AP ở màn Chi tiết",
       "1. Quan sát tiến trình và bố cục màn checkout AP",
       'Header 2 bước: "Thanh toán" → "Hoàn tất đơn hàng"; hiển thị đủ block: Sản phẩm dịch vụ đã chọn, Thông tin cá nhân, Địa chỉ lắp đặt, Phương thức thanh toán, Thông tin khách hàng, Thông tin thanh toán',"N"),
    tc("Medium","Kiểm tra chọn số lượng (không chu kỳ) rồi Mua ngay sang Checkout",
       "Ở màn Chi tiết gói AP",
       "1. Chọn số lượng\n2. Click button \"Mua ngay\"",
       "[BLOCKED – cần confirm: AP chỉ chọn số lượng, KHÔNG có chu kỳ?]. Kỳ vọng: điều hướng sang Thanh toán, load đúng tên gói + số lượng + số tiền","N",
       "BLOCKED – CLA-AP-001 (AP không chu kỳ)"),
    tc("Medium","Kiểm tra Block Sản phẩm dịch vụ đã chọn (không hiển thị chu kỳ)",
       "Đang ở màn Thanh toán AP",
       "1. Quan sát Block \"Sản phẩm dịch vụ đã chọn\"",
       "[BLOCKED – cần confirm: AP không có chu kỳ?]. Kỳ vọng: load đúng tên gói, số lượng, số tiền — KHÔNG hiển thị dòng chu kỳ","N",
       "BLOCKED – CLA-AP-001 (AP không chu kỳ)"),
    grp("Nhóm 2: Block Địa chỉ lắp đặt (đặc thù AP)"),
    tc("Medium","Kiểm tra Block Địa chỉ lắp đặt + note giao hàng / Thời gian lắp đặt",
       "Đang ở màn Thanh toán AP",
       "1. Quan sát Block Địa chỉ lắp đặt",
       "[BLOCKED – cần confirm: AP có note giao hàng 3-7 ngày + Thời gian lắp đặt như Camera không?]. Phần fields địa chỉ (Tỉnh/Phường/Đường/Số nhà/Ghi chú + link địa chỉ cũ) xem TC_CKCOMMON","N",
       "BLOCKED – CLA-AP-002 (note giao hàng/Thời gian lắp đặt AP)"),
    grp("Nhóm 3: Phương thức thanh toán & Thông tin khách hàng"),
    tc("Medium","Kiểm tra AP CÓ phương thức COD + online theo QLCS",
       "Gói AP cấu hình COD + online trên QLCS",
       "1. Quan sát Block Phương thức thanh toán",
       'Hiển thị COD "Thanh toán tại nhà" + các PTTT online theo QLCS; chỉ chọn 1; "Xem thêm" nếu >4. Cơ chế danh sách PTTT xem TC_CKCOMMON',"Y"),
    tc("Medium","Kiểm tra Block Thông tin khách hàng auto-load",
       "Đã nhập Thông tin cá nhân + Địa chỉ lắp đặt",
       "1. Quan sát Block Thông tin khách hàng",
       "[BLOCKED – cần confirm: AP có Thời gian lắp đặt không? (CLA-AP-002)]. Phần auto-load Họ tên/SĐT/Địa chỉ đúng theo dữ liệu đã nhập","N",
       "BLOCKED – CLA-AP-002 (Thời gian lắp đặt AP)"),
    grp("Nhóm 4: Button Thanh toán & Luồng thanh toán"),
    tc("High","Kiểm tra Thanh toán khi chính sách không còn active trên QLCS",
       "Gói/chính sách AP không còn active trên QLCS",
       "1. Click button \"Thanh toán\"",
       "Báo lỗi chính sách không còn hiệu lực; KHÔNG thực hiện thanh toán","Y"),
    tc("High","Kiểm tra Thanh toán khi data hợp lệ + chính sách active",
       "Nhập đủ trường bắt buộc hợp lệ, chính sách active",
       "1. Click button \"Thanh toán\"",
       "Thực hiện luồng thanh toán theo PTTT đã chọn (COD → sang B3 luôn; Online → cổng 3rd party)","Y"),
    grp("Nhóm 5: Bước Hoàn tất đơn hàng"),
    tc("High","Kiểm tra hoàn tất đơn hàng với PTTT = COD",
       "Chọn PTTT = COD, click Thanh toán",
       "1. Hoàn tất đặt hàng với PTTT = COD\n2. Quan sát màn Hoàn tất",
       'Màn Hoàn tất: trạng thái "Chưa thanh toán" + nội dung "Đơn hàng đã đăng ký thành công. Kỹ thuật viên FPT sẽ liên hệ triển khai dịch vụ trong 8h-12h..."',"Y"),
    tc("High","Kiểm tra hoàn tất đơn hàng với PTTT Online thanh toán thành công",
       "Chọn PTTT Online, thanh toán thành công",
       "1. Hoàn tất thanh toán Online thành công\n2. Quan sát màn Hoàn tất",
       'Bước "Hoàn tất đơn hàng" header màu xanh lá; trạng thái "Đã thanh toán" + nội dung thông báo thành công',"N"),
    tc("High","Kiểm tra thanh toán Online thất bại quay về màn Thanh toán",
       "Chọn PTTT Online, thanh toán thất bại",
       "1. Thực hiện thanh toán Online thất bại",
       "Quay về màn Thanh toán giữ nguyên thông tin đã chọn; CHỈ cho thay đổi PTTT + nhập mã ưu đãi, các trường còn lại disable","N"),
    grp("Nhóm 6: Mã ưu đãi & Navigation & Mobile"),
    tc("Medium","Kiểm tra áp dụng mã ưu đãi cho AP",
       "Đang ở màn Thanh toán AP",
       "1. Nhập/áp dụng mã ưu đãi",
       "[BLOCKED – cần confirm: voucher/mã ưu đãi chưa implement (như UltraFast)]. Kỳ vọng: trừ tiền","N",
       "BLOCKED – CLA-CAMAP-001 (voucher chưa implement)"),
    tc("Low","Kiểm tra button Quay lại điều hướng về màn Chi tiết gói",
       "Đang ở màn Thanh toán AP",
       "1. Click button \"Quay lại\"",
       "Điều hướng về màn Chi tiết gói AP","Y"),
    tc("Medium","Kiểm tra hiển thị luồng Checkout AP trên mobile (≤ 768px)",
       "Mở checkout AP trên viewport mobile",
       "1. Quan sát layout 2 bước + các block trên mobile",
       "Các bước và block hiển thị đúng, không vỡ layout, thao tác được trên mobile","N"),
]

wb = openpyxl.load_workbook(PATH)
build_sheet(wb, "Checkout_Camera", "TC_CAMERA", "Đăng ký dịch vụ Camera — Checkout (chu kỳ + COD + địa chỉ lắp đặt)", camera_rows)
build_sheet(wb, "Checkout_AP", "TC_AP", "Đăng ký dịch vụ Access Point (AP) — Checkout (số lượng + COD + địa chỉ lắp đặt)", ap_rows)
wb.save(PATH)

# Thống kê
def count(rows):
    tcs=[x for x in rows if x['type']=='tc']
    blk=[x for x in tcs if x.get('note')]
    pr={'High':0,'Medium':0,'Low':0}
    auto={'Y':0,'N':0}
    for x in tcs:
        pr[x['pri']]+=1; auto[x['auto']]+=1
    return len(tcs), pr, auto, len(blk)
for nm,rows in [("Camera",camera_rows),("AP",ap_rows)]:
    n,pr,au,blk=count(rows)
    print(f"{nm}: {n} TC | High:{pr['High']} Medium:{pr['Medium']} Low:{pr['Low']} | Auto Y:{au['Y']} N:{au['N']} | BLOCKED:{blk}")
print("Sheets:", wb.sheetnames)
print("Saved:", PATH)
