# -*- coding: utf-8 -*-
"""
gen-testcase-checkout-service — Checkout Camera (TC_07)
Thêm sheet "Checkout Camera" vào TC_checkout.xlsx hiện có.

Chỉ viết TCs đặc thù cho dịch vụ Camera:
  - Điều hướng màn hình
  - Block Sản phẩm: Camera + Gói Cloud (đặc thù)
  - Block TT Cá nhân: Hiển thị (chỉ Họ tên + SĐT, không Email)
  - Block Địa chỉ: Hiển thị + pre-fill "Địa chỉ trước"
  - Block TT KH: Có trường Thời gian lắp đặt (đặc thù)
  - Block TT Thanh toán: 2 dòng giá (thiết bị + gói Cloud) + tổng tiền
  - Luồng Thanh toán: còn/hết hiệu lực
  - Màn hình Hoàn tất đơn hàng + verify webadmin/inside

KHÔNG viết lại (đã có ở TC_01):
  - Validation SĐT, Email (format, required, maxlength)
  - Địa chỉ hành chính dropdowns
  - PTTT list + selection
  - OTP flow
  - Logo FPT click
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FILE      = "ecom-pdh/03_test-cases/functional/chucnang_checkout/AI_ISC_ecom-pdh_v1.1_TC_checkout_v1.0.xlsx"
SHEET     = "Checkout Camera"
FUNC_ID   = "TC_07"
FUNC_NAME = "Checkout Camera"

# ─── Palette — theo đúng Smart Home reference ────────────────────────────────
C_HDR  = "4472C4"   # dark blue  — column header background
C_SEP  = "A9D08E"   # green      — section separator (A:G merge)
# "Màn hình Hoàn tất" dùng B:G merge, no fill (giống Smart Home row 26)

def _fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size)

def _align(wrap=True, h="left", v="top"):
    return Alignment(wrap_text=wrap, horizontal=h, vertical=v)

thin = Side(style="thin", color="BFBFBF")
def _border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

# ─── Sheet builders ──────────────────────────────────────────────────────────

def build_header(ws):
    """
    Rows 3-8: info block + column headers.
    Merge pattern matches Checkout Smart Home exactly:
      Row 3 : B3:C3 (label, blue)  |  D3:E3 (value, white)
      Row 4 : B4:C4 (label, blue)  |  D4:E4 (value, white)
      Row 7-8: each of cols A-G merged vertically (A7:A8 … G7:G8)
      Row 7 : H7:K7 merged horizontally
      Row 8 : H8, I8, J8, K8 individual cells
    """
    hdr_fill = _fill(C_HDR)
    hdr_font = _font(bold=True, color="FFFFFF")
    ctr_aln  = Alignment(wrap_text=True, horizontal="center", vertical="center")
    lft_aln  = Alignment(wrap_text=True, horizontal="left",   vertical="center")

    # ── Info block rows 3-4 ──────────────────────────────────────────────────
    # Row 3: label B3:C3 (blue) | value D3:E3 (white, no fill)
    for c in (2, 3):
        cell = ws.cell(row=3, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = lft_aln
    ws.cell(row=3, column=2, value="Mã chức năng\n(Function ID)")
    _merge(ws, 3, 2, 3, 3)

    ws.cell(row=3, column=4, value=FUNC_ID).font = _font(bold=False)
    _merge(ws, 3, 4, 3, 5)

    # Row 4: label B4:C4 (blue) | value D4:E4 (white, bold)
    for c in (2, 3):
        cell = ws.cell(row=4, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = lft_aln
    ws.cell(row=4, column=2, value="Tên chức năng \n(Function Name)")
    _merge(ws, 4, 2, 4, 3)

    ws.cell(row=4, column=4, value=FUNC_NAME).font = _font(bold=True)
    _merge(ws, 4, 4, 4, 5)

    # ── Column headers rows 7-8 ──────────────────────────────────────────────
    # Cols A-G: each column merged vertically across rows 7-8
    main_cols = [
        (1, "QC/AI"),
        (2, "Testcase ID"),
        (3, "Mức Độ Ưu Tiên\n(Priority)"),
        (4, "Nội Dung Test\n(Test Title)"),
        (5, "Điều Kiện/ Dữ Liệu Test\n(Pre-condition/ Test Data)"),
        (6, "Các Bước Thực Hiện\n(Test Steps)"),
        (7, "Kết Quả Mong Đợi\n(Expected Results)"),
    ]
    for col, label in main_cols:
        cell = ws.cell(row=7, column=col, value=label)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ctr_aln
        _merge(ws, 7, col, 8, col)   # merge row 7 + row 8 for this column

    # H7:K7 merged horizontally (execution tracking header)
    ws.cell(row=7, column=8, value="").fill = hdr_fill
    _merge(ws, 7, 8, 7, 11)

    # Row 8: H8-K8 individual execution columns
    exec_cols = [
        (8,  "Kết Quả Thực Hiện\n(Actual Result)"),
        (9,  "Người Thực Hiện\n(Executed By)"),
        (10, "ID Bugs"),
        (11, "Ghi Chú\n(Remark)"),
    ]
    for col, label in exec_cols:
        cell = ws.cell(row=8, column=col, value=label)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = ctr_aln


def write_sep(ws, row_num, label):
    """
    Section separator: A:G merge, green fill.
    Matches Smart Home rows 9 (Điều hướng) and 12 (Màn hình TT thanh toán).
    """
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.fill = _fill(C_SEP)
    cell.font = _font(bold=True, color="375623")
    cell.alignment = _align(wrap=False, h="left", v="center")
    _merge(ws, row_num, 1, row_num, 7)


def write_banner(ws, row_num, label):
    """
    Screen label with B:G merge, no fill.
    Matches Smart Home row 26 (Màn hình Hoàn tất đơn hàng).
    """
    cell = ws.cell(row=row_num, column=2, value=label)
    cell.font = _font(bold=True)
    cell.alignment = _align(wrap=False, h="left", v="center")
    _merge(ws, row_num, 2, row_num, 7)


def write_sub_label(ws, row_num, label):
    """
    Block sub-group label: C:G merge, no fill, label in col C.
    Matches Smart Home rows 13/15/17/20/22.
    Col B gets TC ID formula (evaluates to "" since col F is empty).
    """
    ws.cell(row=row_num, column=2,
            value=f'=IF(F{row_num}="","",$D$3&"."&COUNTA($F$9:F{row_num})&"")')
    cell = ws.cell(row=row_num, column=3, value=label)
    cell.font = _font(bold=True, color="1F3864")
    cell.alignment = _align(wrap=False, h="left", v="center")
    _merge(ws, row_num, 3, row_num, 7)


def write_tc(ws, row_num, first_data_row, priority, title, precond, steps, expected):
    """One TC data row."""
    ws.cell(row=row_num, column=1, value="AI")
    ws.cell(row=row_num, column=2,
            value=f'=IF(F{row_num}="","",$D$3&"."&COUNTA($F${first_data_row}:F{row_num})&"")')
    ws.cell(row=row_num, column=3, value=priority)

    title_c = ws.cell(row=row_num, column=4, value=title)
    title_c.font = _font(bold=True)

    ws.cell(row=row_num, column=5, value=precond)
    ws.cell(row=row_num, column=6, value=steps)
    ws.cell(row=row_num, column=7, value=expected)

    for col in range(1, 12):
        cell = ws.cell(row=row_num, column=col)
        cell.alignment = _align(wrap=True, v="top")
        cell.border = _border()


# ─── TC Content ──────────────────────────────────────────────────────────────

# Row types:
#   "sep"    → A:G merge, green fill  (section header — matches Smart Home rows 9, 12)
#   "banner" → B:G merge, no fill     (matches Smart Home row 26 "Màn hình Hoàn tất")
#   "sub"    → C:G merge, no fill     (block sub-label — matches Smart Home rows 13,15,17,20,22)
#   "tc"     → (priority, title, precond, steps, expected)

TC_DATA = [
    # ── Điều hướng ────────────────────────────────────────────────────────────
    ("sep", "Điều hướng màn hình"),
    ("tc", "High",
     "Kiểm tra điều hướng vào màn hình Checkout Camera + verify URL",
     "",
     "1. Truy cập trang chi tiết sản phẩm Camera (VD: Camera SE S2)\n"
     "2. Click button Mua ngay\n"
     "3. Kiểm tra URL và giao diện trang Checkout Camera",
     "Điều hướng vào trang Checkout Camera đúng URL.\n"
     "Hiển thị đủ các block: Sản phẩm đã chọn (Camera + Gói Cloud), Thông tin cá nhân, "
     "Địa chỉ lắp đặt, Phương thức thanh toán, Panel thông tin KH + thanh toán bên phải."),

    ("tc", "High",
     "Kiểm tra GUI tổng thể màn hình Checkout Camera",
     "Đang ở màn hình Checkout Camera",
     "1. Truy cập màn hình Checkout Camera\n"
     "2. Kiểm tra GUI toàn trang",
     "Hiển thị UI đầy đủ gồm:\n"
     "- Progress bar 2 bước: Thanh toán → Hoàn tất đơn hàng\n"
     "- Block Sản phẩm dịch vụ đã chọn (Camera + thông tin Gói Cloud)\n"
     "- Block Thông tin cá nhân\n"
     "- Block Địa chỉ lắp đặt\n"
     "- Block Phương thức thanh toán\n"
     "- Panel phải: Block Thông tin khách hàng + Block Thông tin thanh toán\n"
     "- Button Thanh toán\n"
     "- Button Quay lại"),

    # ── Màn hình Thông tin thanh toán ─────────────────────────────────────────
    ("sep", "Màn hình Thông tin thanh toán"),

    ("sub", "Block Sản phẩm dịch vụ đã chọn"),
    ("tc", "Medium",
     "Kiểm tra hiển thị thông tin Camera trong block Sản phẩm dịch vụ đã chọn",
     "Chọn sản phẩm Camera từ màn hình Chi tiết sản phẩm",
     "1. Truy cập màn hình Checkout Camera\n"
     "2. Kiểm tra hiển thị block Sản phẩm dịch vụ đã chọn",
     "Block hiển thị đầy đủ thông tin:\n"
     "- Title: Sản phẩm dịch vụ đã chọn (số lượng)\n"
     "- Icon/Hình ảnh Camera\n"
     "- Tên sản phẩm (VD: Camera SE S2)\n"
     "- Thông tin Gói Cloud (VD: Cloud: 3 ngày | Chu kỳ: 6 tháng)\n"
     "- Số lượng: x1\n"
     "- Giá tổng (VD: 840.000đ, màu xanh nổi bật)\n"
     "Không có nút chỉnh sửa số lượng tại màn hình này."),

    ("tc", "Medium",
     "Kiểm tra hiển thị đúng thông tin Gói Cloud kèm Camera",
     "Chọn Camera kèm gói Cloud từ màn hình Chi tiết sản phẩm",
     "1. Truy cập màn hình Checkout Camera\n"
     "2. Quan sát thông tin Gói Cloud trong block Sản phẩm dịch vụ đã chọn",
     "Gói Cloud hiển thị đúng thông tin:\n"
     "- Tên gói (VD: Cloud: 3 ngày)\n"
     "- Chu kỳ (VD: Chu kỳ: 6 tháng)\n"
     "Thông tin khớp với gói đã chọn ở màn hình Chi tiết sản phẩm."),

    ("sub", "Block Thông tin cá nhân"),
    ("tc", "Medium",
     "Kiểm tra hiển thị block Thông tin cá nhân — Camera chỉ có Họ tên + SĐT",
     "Đang ở màn hình Checkout Camera",
     "1. Truy cập màn hình Checkout Camera\n"
     "2. Kiểm tra UI block Thông tin cá nhân",
     "Block hiển thị đúng 2 trường bắt buộc:\n"
     "- Họ và tên (*)\n"
     "- Số điện thoại (*)\n"
     "Không có trường Email (Email không áp dụng cho dịch vụ Camera)."),

    ("sub", "Block Địa chỉ lắp đặt"),
    ("tc", "Medium",
     "Kiểm tra hiển thị block Địa chỉ lắp đặt",
     "Đang ở màn hình Checkout Camera",
     "1. Truy cập màn hình Checkout Camera\n"
     "2. Kiểm tra block Địa chỉ lắp đặt",
     "Block Địa chỉ lắp đặt hiển thị mặc định, gồm các trường bắt buộc:\n"
     "- Tỉnh/Thành phố (*)\n"
     "- Phường/Xã (*)\n"
     "- Tên đường (*)\n"
     "- Radio: Nhà riêng / Chung cư\n"
     "- Số nhà (*)\n"
     "- Ghi chú (không bắt buộc)"),

    ("tc", "Low",
     "Kiểm tra hiển thị label thời gian giao hàng dự kiến 3–7 ngày",
     "Đang ở màn hình Checkout Camera",
     "1. Truy cập màn hình Checkout Camera\n"
     "2. Kiểm tra label thông báo trong block Địa chỉ lắp đặt",
     "Hiển thị đúng text: 'Thời gian giao hàng dự kiến từ 3 đến 7 ngày'."),

    ("tc", "Medium",
     "Kiểm tra chức năng pre-fill 'Địa chỉ trước sắp nhập'",
     "Người dùng đã từng đặt hàng và có địa chỉ lưu sẵn",
     "1. Truy cập màn hình Checkout Camera với account đã có địa chỉ trước\n"
     "2. Quan sát block Địa chỉ lắp đặt — link 'Địa chỉ trước sắp nhập'\n"
     "3. Click link 'Địa chỉ trước sắp nhập'",
     "[BLOCKED – cần confirm BA: Chức năng 'Địa chỉ trước sắp nhập' hoạt động thế nào? "
     "Khi click tự điền toàn bộ hay chỉ một phần? Có áp dụng cho tất cả dịch vụ hay chỉ Camera?]"),

    ("sub", "Block Thông tin khách hàng"),
    ("tc", "Medium",
     "Kiểm tra hiển thị block Thông tin khách hàng — có trường Thời gian lắp đặt",
     "Đang ở màn hình Checkout Camera, đã nhập đầy đủ thông tin",
     "1. Truy cập màn hình Checkout Camera\n"
     "2. Nhập đầy đủ Thông tin cá nhân và Địa chỉ\n"
     "3. Kiểm tra panel Thông tin khách hàng bên phải",
     "Panel Thông tin khách hàng hiển thị đầy đủ:\n"
     "- Họ và tên\n"
     "- Số điện thoại\n"
     "- Địa chỉ (Số nhà + Tên đường + Phường/Xã + Tỉnh/TP)\n"
     "- Thời gian lắp đặt (VD: Thứ 2, 02/01/2024 10:00–11:10)\n"
     "- Icon collapse (mặc định thu gọn)"),

    ("tc", "Medium",
     "Kiểm tra định dạng và nguồn dữ liệu trường Thời gian lắp đặt",
     "Đang ở màn hình Checkout Camera",
     "1. Truy cập màn hình Checkout Camera\n"
     "2. Kiểm tra trường Thời gian lắp đặt trong panel Thông tin khách hàng",
     "Trường Thời gian lắp đặt hiển thị đúng định dạng:\n"
     "Thứ [N], DD/MM/YYYY HH:MM–HH:MM\n"
     "VD: Thứ 2, 02/01/2024 10:00–11:10\n"
     "[BLOCKED – cần confirm BA: Thời gian lắp đặt lấy từ đâu? "
     "Hệ thống tự assign hay khách hàng được chọn? Field này có phải nhập bắt buộc không?]"),

    ("sub", "Block Thông tin thanh toán"),
    ("tc", "Medium",
     "Kiểm tra hiển thị block Thông tin thanh toán — 2 dòng giá (thiết bị + gói Cloud)",
     "Đang ở màn hình Checkout Camera",
     "1. Truy cập màn hình Checkout Camera\n"
     "2. Kiểm tra block Thông tin thanh toán (panel bên phải)",
     "Block Thông tin thanh toán hiển thị:\n"
     "- Dòng 1: [Tên Camera] - 1 cái: <giá thiết bị> (VD: Camera SE S2 - 1 cái: 600.000đ)\n"
     "- Dòng 2: [Tên gói Cloud] - [chu kỳ]: <giá gói> (VD: Gói Cloud 3D - 6 tháng: 240.000đ)\n"
     "- Ô nhập mã khuyến mãi + hyperlink Chọn ưu đãi + Button Áp dụng\n"
     "- Cần thanh toán: <tổng giá> (VD: 840.000đ, chữ to, nổi bật)\n"
     "- Button Thanh toán\n"
     "- Text: 'Bằng việc nhấn vào nút Thanh toán bạn đã đồng ý với các điều khoản của FPT Telecom'"),

    ("tc", "High",
     "Kiểm tra tính đúng tổng tiền cần thanh toán = giá Camera + giá Gói Cloud",
     "Đang ở màn hình Checkout Camera, chưa áp dụng voucher",
     "1. Truy cập màn hình Checkout Camera\n"
     "2. Ghi nhận giá Camera trong block Thông tin thanh toán\n"
     "3. Ghi nhận giá Gói Cloud trong block Thông tin thanh toán\n"
     "4. Tính tổng: Giá Camera + Giá Gói Cloud\n"
     "5. So sánh với số tiền 'Cần thanh toán' hiển thị trên UI",
     "Số tiền Cần thanh toán = Giá Camera + Giá Gói Cloud.\n"
     "VD: 600.000đ + 240.000đ = 840.000đ.\n"
     "Không có sai lệch giữa tổng tính tay và số hiển thị."),

    ("tc", "High",
     "Kiểm tra click button Thanh toán khi chính sách gói Camera còn hiệu lực",
     "Gói Camera có chính sách còn hiệu lực trên QLCS",
     "1. Điều hướng vào trang Checkout Camera thành công\n"
     "2. Nhập đầy đủ thông tin bắt buộc\n"
     "3. Chọn phương thức thanh toán\n"
     "4. Click button Thanh toán",
     "Hệ thống lên đơn hàng thành công, điều hướng qua màn hình thanh toán "
     "của PTTT đã chọn và hiển thị đúng số tiền cần thanh toán."),

    ("tc", "High",
     "Kiểm tra click button Thanh toán khi chính sách gói Camera đã hết hiệu lực",
     "Gói Camera có ít nhất 1 chính sách đã hết hiệu lực trên QLCS",
     "1. Điều hướng vào trang Checkout Camera thành công\n"
     "2. Nhập đầy đủ thông tin bắt buộc\n"
     "3. Chọn phương thức thanh toán\n"
     "4. Click button Thanh toán",
     "Hệ thống hiển thị thông báo lỗi, không lên đơn hàng."),

    # ── Màn hình Hoàn tất đơn hàng ────────────────────────────────────────────
    ("banner", "Màn hình Hoàn tất đơn hàng"),

    ("tc", "High",
     "Kiểm tra giao diện màn hình Hoàn tất đơn hàng — thanh toán COD (Chưa thanh toán)",
     "Thanh toán COD (thanh toán tại nhà)",
     "1. Hoàn tất đặt hàng Camera với PTTT COD\n"
     "2. Kiểm tra giao diện màn hình Hoàn tất đơn hàng",
     "Hiển thị giao diện hoàn tất đơn hàng gồm:\n"
     "- Logo FPT Telecom\n"
     "- Các bước: Thanh toán | Hoàn tất đơn hàng (màu xanh lá)\n"
     "- Label: Hoàn tất đơn hàng\n"
     "- Label trạng thái: Chưa thanh toán\n"
     "- Text thông báo giao hàng Camera\n"
     "  [BLOCKED – cần confirm BA: nội dung chính xác text thông báo giao Camera]\n"
     "- Block Thông tin đơn hàng (ref: sheet Thông tin chung)"),

    ("tc", "High",
     "Kiểm tra giao diện màn hình Hoàn tất đơn hàng — thanh toán khác COD (Đã thanh toán)",
     "Thanh toán qua VietQR / MoMo / Thẻ quốc tế",
     "1. Hoàn tất đặt hàng Camera với PTTT khác COD\n"
     "2. Kiểm tra giao diện màn hình Hoàn tất đơn hàng",
     "Hiển thị giao diện hoàn tất đơn hàng gồm:\n"
     "- Logo FPT Telecom\n"
     "- Các bước: Thanh toán | Hoàn tất đơn hàng (màu xanh lá)\n"
     "- Label: Hoàn tất đơn hàng\n"
     "- Label trạng thái: Thanh toán thành công\n"
     "- Text thông báo giao hàng Camera\n"
     "  [BLOCKED – cần confirm BA: nội dung chính xác text thông báo giao Camera]\n"
     "- Block Thông tin đơn hàng (ref: sheet Thông tin chung)"),

    ("tc", "High",
     "Kiểm tra thông tin đơn hàng Camera trên webadmin sau thanh toán thành công",
     "Đã thanh toán thành công, đang ở bước Hoàn tất",
     "1. Tại bước Hoàn tất đơn hàng Camera\n"
     "2. Truy cập webadmin\n"
     "3. Kiểm tra lại thông tin đơn hàng Camera vừa tạo",
     "Đơn hàng Camera ghi nhận đúng toàn bộ thông tin đã nhập:\n"
     "- Thông tin khách hàng (Họ tên, SĐT)\n"
     "- Địa chỉ lắp đặt\n"
     "- Sản phẩm: Tên Camera + Gói Cloud + số lượng\n"
     "- Số tiền thanh toán (giá thiết bị + giá gói Cloud)\n"
     "- Thời gian lắp đặt"),

    ("tc", "High",
     "Kiểm tra thông tin Hợp đồng Camera trên inside sau thanh toán thành công",
     "Đã thanh toán thành công",
     "1. Tại bước Hoàn tất đơn hàng Camera\n"
     "2. Truy cập inside\n"
     "3. Kiểm tra lại thông tin hợp đồng Camera",
     "Hợp đồng Camera ghi nhận đúng các thông tin đơn hàng:\n"
     "- Thông tin khách hàng và địa chỉ lắp đặt\n"
     "- Sản phẩm Camera + Gói Cloud\n"
     "- Giá trị hợp đồng"),
]


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    wb = load_workbook(FILE)

    # Remove old sheet if exists
    if SHEET in wb.sheetnames:
        del wb[SHEET]

    ws = wb.create_sheet(title=SHEET)

    build_header(ws)

    # Column widths — matches Smart Home exactly
    # A=13, B=default(no set), C=12.4, D=40.6, E=20.1, F=46.1, G=61,
    # H=13.7, I=38.7, J=15.4, K=17
    col_widths = {
        1: 13.0,                  # A
        3: 12.4, 4: 40.6, 5: 20.1, 6: 46.1, 7: 61.0,   # C-G
        8: 13.7, 9: 38.7, 10: 15.4, 11: 17.0,            # H-K
    }
    for col_idx, w in col_widths.items():
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = w

    # Row heights — matches Smart Home
    ws.row_dimensions[7].height = 15.0
    ws.row_dimensions[8].height = 63.0

    # Data starts at row 9 (first separator)
    FIRST_DATA_ROW = 9
    current_row = FIRST_DATA_ROW
    tc_count  = 0
    high_c = medium_c = low_c = blocked_c = 0

    for entry in TC_DATA:
        kind = entry[0]

        if kind == "sep":
            write_sep(ws, current_row, entry[1])
            ws.row_dimensions[current_row].height = 15.75
            current_row += 1

        elif kind == "banner":
            write_banner(ws, current_row, entry[1])
            ws.row_dimensions[current_row].height = 15.75
            current_row += 1

        elif kind == "sub":
            write_sub_label(ws, current_row, entry[1])
            ws.row_dimensions[current_row].height = 15.75
            current_row += 1

        elif kind == "tc":
            _, priority, title, precond, steps, expected = entry
            write_tc(ws, current_row, FIRST_DATA_ROW,
                     priority, title, precond, steps, expected)
            tc_count += 1
            if priority == "High":     high_c += 1
            elif priority == "Medium": medium_c += 1
            elif priority == "Low":    low_c += 1
            if "[BLOCKED" in expected: blocked_c += 1
            current_row += 1

    wb.save(FILE)
    print(f"  [OK]  Sheet '{SHEET}' ({FUNC_ID}) — {tc_count} TCs written")
    print(f"        High:{high_c}  Medium:{medium_c}  Low:{low_c}  BLOCKED:{blocked_c}")
    print(f"  Saved: {FILE}")
    print()
    print("✅ TC Checkout Camera đã tạo xong.")
    print(f"   Sheet: {SHEET} | Function ID: {FUNC_ID}")
    print(f"   Tổng TC: {tc_count} | High:{high_c} Medium:{medium_c} Low:{low_c}")
    print(f"   BLOCKED: {blocked_c} TC cần BA confirm")
    print(f"   Không viết lại: TCs validation fields, PTTT, OTP, Logo đã có ở TC_01")
    print()
    print("Open items:")
    print("   1. [BLOCKED TC_07.8] 'Địa chỉ trước sắp nhập': spec pre-fill cụ thể?")
    print("   2. [BLOCKED TC_07.11] Thời gian lắp đặt: hệ thống assign hay khách chọn? Bắt buộc?")
    print("   3. [BLOCKED TC_07.16-17] Text thông báo hoàn tất giao hàng Camera?")


if __name__ == "__main__":
    main()
