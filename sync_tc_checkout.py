# -*- coding: utf-8 -*-
"""
Sync Playwright JSON report → TC Excel (Checkout) v1.2.
- TC_CKCOMMON: ghi 4 block CỐ ĐỊNH theo dịch vụ
  (Internet=col9, Camera=col13, AP=col17, SmartTV=col21) → nhìn phát biết dịch vụ nào Fail.
  Internet chạy CKCOMMON trên màn Internet, report KHÔNG có prefix [svc] → khớp key (tid, None).
  AP/SmartTV/Camera chạy qua runCkcommonSuite → describe title có prefix [AP]/[SmartTV]/[Camera].
- Các TC khác: ghi Round 1 (col9) như trước.
Chạy từ thư mục gốc: python sync_tc_checkout.py
"""
import json, re, os, shutil, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

ROOT        = os.path.dirname(os.path.abspath(__file__))
# Run 2026-06-11: 3 report tách theo dịch vụ thiết bị (mỗi report = CKCOMMON[svc] + case riêng svc).
# Internet KHÔNG có trong run này → block Internet sẽ là Block.
REPORTS     = [
    os.path.join(ROOT, 'ecom-pdh/06_report/report_ap.json'),
    os.path.join(ROOT, 'ecom-pdh/06_report/report_camera.json'),
    os.path.join(ROOT, 'ecom-pdh/06_report/report_smarttv.json'),
]
EXCEL       = os.path.join(ROOT, 'ecom-pdh/03_test-cases/functional/chucnang_checkout/AI_ISC_ecom-pdh_v1.1_TC_checkout_v1.2.xlsx')
RESULTS_DIR = os.path.join(ROOT, 'ecom-pdh/06_report')
DATE        = '2026-06-11'
EXECUTOR    = 'Auto'

TCID_RE    = re.compile(r'(?:TC_[A-Z0-9_]+\.\d+|API_\d+\.\d+)')
SERVICE_RE = re.compile(r'\[(AP|SmartTV|Camera)\]')

# Sheet Checkout_Common — 4 block cố định theo dịch vụ: (label, col bắt đầu, key dịch vụ trong report)
# svc_key=None nghĩa là khớp test KHÔNG có prefix [svc] (Internet realize CKCOMMON trên màn Internet).
CKCOMMON_SERVICES = [
    ('Internet', 9,  None),       # I–L
    ('Camera',   13, 'Camera'),   # M–P
    ('AP',       17, 'AP'),       # Q–T
    ('SmartTV',  21, 'SmartTV'),  # U–X
]

# ── 1. Parse report → {(TC_ID, service|None): 'Pass'/'Fail'} ───────────────
results:   dict[tuple, str] = {}   # (tid, svc) → status
order_ids: dict[tuple, str] = {}   # (tid, svc) → orderId (nếu có)

def walk(node: dict, service: str | None = None) -> None:
    title = node.get('title', '')
    m = SERVICE_RE.search(title)
    current_service = m.group(1) if m else service

    for child in node.get('suites', []):
        walk(child, current_service)

    for sp in node.get('specs', []):
        ids = TCID_RE.findall(sp.get('title', ''))
        status = 'Pass' if sp.get('ok') else 'Fail'
        # Đọc orderId từ annotation (type='orderId')
        oid = ''
        for t in sp.get('tests', []):
            for ann in t.get('annotations', []):
                if ann.get('type') == 'orderId' and ann.get('description'):
                    oid = ann['description']
                    break
        for tid in ids:
            key = (tid, current_service)
            if results.get(key) != 'Fail':
                results[key] = status
            if status == 'Fail':
                results[key] = 'Fail'
            if oid:
                order_ids[key] = oid

for rep in REPORTS:
    if not os.path.exists(rep):
        print(f'WARN: report khong ton tai, bo qua: {rep}')
        continue
    with open(rep, encoding='utf-8') as f:
        walk(json.load(f))

total_pass = sum(1 for v in results.values() if v == 'Pass')
total_fail = sum(1 for v in results.values() if v == 'Fail')
print(f'Report: {len(results)} entries | Pass={total_pass} Fail={total_fail}')

# ── 2. Styles ───────────────────────────────────────────────────────────────
PASS_F  = PatternFill('solid', fgColor='C6EFCE')
FAIL_F  = PatternFill('solid', fgColor='FFC7CE')
BLOCK_F = PatternFill('solid', fgColor='FFEB9C')
thin    = Side(style='thin', color='BFBFBF')
BORDER  = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP    = Alignment(wrap_text=True, vertical='top')
CTR     = Alignment(wrap_text=True, vertical='top', horizontal='center')

# ── 3. Fill Excel ───────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)

summary   = {'Pass': 0, 'Fail': 0, 'Block': 0}
fails: list[tuple] = []
warnings:  list[str] = []
matched_keys: set[tuple] = set()

def fill_round(ws, row: int, col_start: int, result: str, note: str,
               bug: str = '', order_id: str = '') -> None:
    fill = PASS_F if result == 'Pass' else (FAIL_F if result == 'Fail' else BLOCK_F)
    full_note = f'{note} | Mã ĐH: {order_id}' if order_id else note
    ws.cell(row, col_start,     result).fill      = fill
    ws.cell(row, col_start).alignment             = CTR
    ws.cell(row, col_start).border                = BORDER
    ws.cell(row, col_start + 1, EXECUTOR).alignment = CTR
    ws.cell(row, col_start + 1).border            = BORDER
    ws.cell(row, col_start + 2, bug).alignment    = CTR
    ws.cell(row, col_start + 2).border            = BORDER
    ws.cell(row, col_start + 3, full_note).alignment = WRAP
    ws.cell(row, col_start + 3).border            = BORDER

for ws in wb.worksheets:
    fid         = ws['D3'].value or ''
    is_ckcommon = (fid == 'TC_CKCOMMON')

    seq = 0
    for r in range(10, ws.max_row + 1):
        if not ws.cell(r, 4).value:
            continue
        seq += 1
        tcid = f'{fid}.{seq}'
        exp  = str(ws.cell(r, 7).value or '')
        auto = str(ws.cell(r, 8).value or '').strip().upper()

        if is_ckcommon:
            # Ghi riêng từng dịch vụ vào đúng block (Internet/Camera/AP/SmartTV)
            for label, col_start, svc_key in CKCOMMON_SERVICES:
                key = (tcid, svc_key)
                if key in results:
                    matched_keys.add(key)
                    res  = results[key]
                    note = f'Auto {DATE} — {label}'
                    oid  = order_ids.get(key, '')
                    if res == 'Pass':
                        summary['Pass'] += 1
                    else:
                        summary['Fail'] += 1
                        fails.append((ws.title, r, tcid, label))
                    fill_round(ws, r, col_start, res, note, order_id=oid)
                elif '[BLOCKED' in exp:
                    fill_round(ws, r, col_start, 'Block',
                               f'Block: [BLOCKED] chờ BA confirm — {label}')
                    summary['Block'] += 1
                elif auto == 'N':
                    fill_round(ws, r, col_start, 'Block',
                               f'Block: TC manual (Auto?=N) — {label}')
                    summary['Block'] += 1
                else:
                    fill_round(ws, r, col_start, 'Block',
                               f'Block: chưa chạy cho {label} trong run này')
                    summary['Block'] += 1
        else:
            # Các sheet khác: ghi Round 1 (col 9)
            key = (tcid, None)
            if key in results:
                matched_keys.add(key)
                res  = results[key]
                note = f'Auto {DATE}'
                oid  = order_ids.get(key, '')
                if res == 'Pass':
                    summary['Pass'] += 1
                else:
                    summary['Fail'] += 1
                    fails.append((ws.title, r, tcid, None))
                fill_round(ws, r, 9, res, note, order_id=oid)
            elif '[BLOCKED' in exp:
                fill_round(ws, r, 9, 'Block', 'Block: [BLOCKED] chờ BA confirm')
                summary['Block'] += 1
            elif auto == 'N':
                fill_round(ws, r, 9, 'Block', 'Block: TC manual (Auto?=N) — cần test tay')
                summary['Block'] += 1
            else:
                fill_round(ws, r, 9, 'Block', 'Block: Auto?=Y chưa chạy trong run này')
                summary['Block'] += 1

# TC trong report nhưng không khớp Excel
for key in results:
    if key not in matched_keys:
        warnings.append(str(key))

# ── 4. Save ─────────────────────────────────────────────────────────────────
os.makedirs(RESULTS_DIR, exist_ok=True)
basename = os.path.basename(EXCEL).replace('.xlsx', f'_results_{DATE}.xlsx')
out = os.path.join(RESULTS_DIR, basename)
wb.save(out)

print(f'\n===== SYNC SUMMARY =====')
print(f'Output: {out}')
print(f'Pass={summary["Pass"]} | Fail={summary["Fail"]} | Block={summary["Block"]}')
if fails:
    print('FAILS:')
    for t in fails:
        print(f'  sheet={t[0]} row={t[1]} tc={t[2]} service={t[3]}')
if warnings:
    print('WARN (trong report nhưng không có trong Excel):', warnings)
