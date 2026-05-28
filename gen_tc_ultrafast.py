"""
Generate sheet "DangKy_UltraFast" vào file TC hiện có.
Module: Đăng ký dịch vụ UltraFast — Màn hình Thanh toán
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

TC_FILE = "ecom-pdh/03_test-cases/AI_ISC_ecom-pdh_v1.1_TC_v1.0.xlsx"
SHEET_NAME = "DangKy_UltraFast"
FUNC_ID    = "TC_DANGKYUF"
FUNC_NAME  = "Đăng ký dịch vụ UltraFast — Màn hình Thanh toán"

# ── Colors ─────────────────────────────────────────────────────────────────
BLUE_FILL  = PatternFill("solid", fgColor="4472C4")   # header
GREEN_FILL = PatternFill("solid", fgColor="A9D08E")   # group header
WHITE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BOLD_FONT  = Font(bold=True, name="Calibri", size=11)
NORM_FONT  = Font(name="Calibri", size=11)

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CTR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── TC data ─────────────────────────────────────────────────────────────────
# (priority, title, precondition, steps, expected, auto)
GROUPS = [
    {
        "name": "Nhóm 1: Load & Hiển thị màn hình Checkout",
        "tcs": [
            (
                "Medium",
                "Kiểm tra hiển thị tổng thể màn hình Checkout UltraFast",
                "- Gói UltraFast còn active trên Staging\n- Ít nhất 1 chu kỳ khả dụng\n- URL: http://ecp-stag.fpt.net/",
                "1. Truy cập trang Chi tiết gói UltraFast\n2. Chọn một chu kỳ bất kỳ\n3. Click button \"Mua ngay\"\n4. Quan sát toàn bộ màn hình Checkout sau khi load xong",
                "Màn hình Thanh toán (bước màu xanh dương) hiển thị đầy đủ:\n- Logo FPT ở header\n- Block \"Sản phẩm dịch vụ đã chọn\"\n- Block \"Thông tin cá nhân\" (có trường Số điện thoại; có checkbox Tôi muốn nhận hóa đơn nhưng không test)\n- Block \"Phương thức thanh toán\"\n- Block \"Thông tin khách hàng\"\n- Block \"Thông tin thanh toán\"\n- Dòng \"Cần thanh toán\"\n- Button \"Thanh toán\"\n- Text điều khoản FPT Telecom\n- Button \"Quay lại\"",
                "N"
            ),
            (
                "High",
                "Kiểm tra chọn chu kỳ và click \"Mua ngay\" điều hướng sang màn hình Checkout",
                "- Gói UltraFast có ít nhất 2 chu kỳ bán khả dụng\n- Đang ở màn hình Chi tiết gói UltraFast",
                "1. Quan sát các chu kỳ bán hiển thị trên màn hình Chi tiết\n2. Chọn chu kỳ bất kỳ (VD: chu kỳ 1 tháng)\n3. Click button \"Mua ngay\"",
                "Hệ thống điều hướng sang màn hình Thanh toán (bước màu xanh dương)\nURL/state phản ánh đúng gói và chu kỳ vừa chọn",
                "Y"
            ),
            (
                "High",
                "Kiểm tra màn hình Checkout load đúng chu kỳ và số tiền đã chọn từ Chi tiết",
                "- Đã chọn gói UltraFast chu kỳ X giá Y tại màn hình Chi tiết\n- Vừa click \"Mua ngay\" → đang chờ Checkout load",
                "1. Quan sát màn hình Checkout sau khi load xong\n2. Kiểm tra thông tin chu kỳ hiển thị\n3. Kiểm tra số tiền hiển thị",
                "Màn hình Checkout hiển thị đúng:\n- Chu kỳ: khớp chu kỳ X đã chọn tại B1\n- Số tiền: khớp giá Y của chu kỳ đó",
                "Y"
            ),
            (
                "High",
                "Kiểm tra Block \"Sản phẩm dịch vụ đã chọn\" hiển thị đúng thông tin",
                "- Đã chọn gói UltraFast cụ thể (tên gói, chu kỳ, số lượng, giá) từ màn hình Chi tiết\n- Đang ở màn hình Checkout",
                "1. Quan sát Block \"Sản phẩm dịch vụ đã chọn\"\n2. Kiểm tra từng thông tin hiển thị trong block",
                "Block hiển thị đầy đủ và chính xác:\n- Tên gói: khớp gói đã chọn\n- Chu kỳ: khớp chu kỳ đã chọn\n- Số lượng: đúng\n- Số tiền: đúng giá của chu kỳ đó",
                "Y"
            ),
        ]
    },
    {
        "name": "Nhóm 2: Block Thông tin cá nhân — Số điện thoại",
        "tcs": [
            (
                "High",
                "Kiểm tra nhập số điện thoại hợp lệ (10 số, bắt đầu bằng 0)",
                "- Đang ở màn hình Checkout UltraFast\n- Trường Số điện thoại đang trống",
                "1. Click vào trường \"Số điện thoại\"\n2. Nhập SĐT hợp lệ: 10 số, bắt đầu bằng 0 (VD: 0901234567)\n3. Quan sát trạng thái của trường",
                "- Textbox chấp nhận giá trị nhập\n- Không hiển thị thông báo lỗi\n- Border trường bình thường (không đỏ)\n- Icon X xuất hiện ở cuối textbox",
                "Y"
            ),
            (
                "High",
                "Kiểm tra báo lỗi khi bỏ trống Số điện thoại và click Thanh toán",
                "- Đang ở màn hình Checkout UltraFast\n- Trường Số điện thoại đang trống\n- Đã chọn PTTT",
                "1. Để trống trường \"Số điện thoại\"\n2. Click button \"Thanh toán\"",
                "- Hệ thống KHÔNG thực hiện thanh toán\n- Hiển thị thông báo lỗi: \"Vui lòng nhập số điện thoại.\"\n- Border đỏ xuất hiện quanh trường Số điện thoại",
                "Y"
            ),
            (
                "Medium",
                "Kiểm tra báo lỗi khi nhập Số điện thoại ít hơn 10 số",
                "- Đang ở màn hình Checkout UltraFast\n- Trường Số điện thoại đang trống\nDữ liệu test: SĐT 9 số (VD: 090123456)",
                "1. Click vào trường \"Số điện thoại\"\n2. Nhập SĐT 9 số (VD: 090123456)\n3. Click ra ngoài trường hoặc click button \"Thanh toán\"",
                "- Hiển thị thông báo lỗi: \"Số điện thoại không hợp lệ.\"\n- Border đỏ xuất hiện quanh textbox",
                "Y"
            ),
            (
                "Medium",
                "Kiểm tra báo lỗi khi nhập Số điện thoại không bắt đầu bằng 0",
                "- Đang ở màn hình Checkout UltraFast\n- Trường Số điện thoại đang trống\nDữ liệu test: SĐT 10 số bắt đầu bằng 1 (VD: 1901234567)",
                "1. Click vào trường \"Số điện thoại\"\n2. Nhập SĐT 10 số không bắt đầu bằng 0 (VD: 1901234567)\n3. Click ra ngoài trường hoặc click \"Thanh toán\"",
                "- Hiển thị thông báo lỗi: \"Số điện thoại không hợp lệ.\"\n- Border đỏ xuất hiện quanh textbox",
                "Y"
            ),
            (
                "Medium",
                "Kiểm tra trường Số điện thoại không cho nhập quá 10 ký tự (boundary)",
                "- Đang ở màn hình Checkout UltraFast\n- Trường Số điện thoại đang trống\nDữ liệu test: chuỗi 11 chữ số (VD: 09012345678)",
                "1. Click vào trường \"Số điện thoại\"\n2. Cố nhập chuỗi 11 chữ số (VD: 09012345678)",
                "- Hệ thống chỉ nhận đúng 10 ký tự đầu (0901234567)\n- Ký tự thứ 11 không được nhập vào trường",
                "Y"
            ),
            (
                "Medium",
                "Kiểm tra icon X xóa nội dung đã nhập trong trường Số điện thoại",
                "- Đang ở màn hình Checkout UltraFast\n- Trường Số điện thoại đang trống",
                "1. Click vào trường \"Số điện thoại\"\n2. Nhập bất kỳ ký tự nào\n3. Quan sát icon X xuất hiện cuối textbox\n4. Click vào icon X",
                "- Sau bước 2: Icon X xuất hiện ở cuối textbox\n- Sau bước 4: Toàn bộ nội dung trong trường SĐT bị xóa sạch; Icon X biến mất",
                "Y"
            ),
        ]
    },
    {
        "name": "Nhóm 3: Block Phương thức thanh toán",
        "tcs": [
            (
                "High",
                "Kiểm tra Block PTTT hiển thị đúng danh sách theo cấu hình QLCS",
                "- Gói UltraFast đã được cấu hình các PTTT Online trên tool QLCS (VD: ATM, Momo, VietQR, Zalopay, Thẻ tín dụng)\n- Đang ở màn hình Checkout UltraFast",
                "1. Quan sát Block \"Phương thức thanh toán\"\n2. Đếm và liệt kê các PTTT hiển thị\n3. So sánh với danh sách PTTT đã cấu hình trên QLCS",
                "- Block PTTT hiển thị đúng và đủ các phương thức thanh toán Online đã cấu hình trên QLCS\n- Số lượng PTTT hiển thị khớp với số lượng đã cấu hình",
                "Y"
            ),
            (
                "High",
                "Kiểm tra không xuất hiện option COD (Thanh toán khi triển khai) trong Block PTTT",
                "- Đang ở màn hình Checkout UltraFast\n- Block PTTT đã load xong",
                "1. Quan sát toàn bộ danh sách trong Block \"Phương thức thanh toán\"\n2. Tìm kiếm option có nội dung \"Thanh toán khi triển khai\" hoặc \"COD\"",
                "- Option \"Thanh toán khi triển khai\" (COD) KHÔNG xuất hiện trong danh sách PTTT\n- Chỉ có các PTTT Online được hiển thị",
                "Y"
            ),
            (
                "Medium",
                "Kiểm tra Block PTTT chỉ hiển thị đúng số lượng PTTT theo cấu hình QLCS",
                "- Gói UltraFast được cấu hình chỉ 2 PTTT Online (VD: ATM và Momo) trên QLCS\n- Đang ở màn hình Checkout UltraFast",
                "1. Quan sát Block \"Phương thức thanh toán\"\n2. Đếm số lượng option PTTT hiển thị",
                "- Block PTTT chỉ hiển thị đúng 2 phương thức đã khai báo (ATM và Momo)\n- Không xuất hiện thêm PTTT nào khác ngoài 2 cái đã cấu hình",
                "Y"
            ),
        ]
    },
    {
        "name": "Nhóm 4: Block Thông tin khách hàng & Thông tin thanh toán",
        "tcs": [
            (
                "Medium",
                "Kiểm tra Block Thông tin khách hàng hiển thị đúng data có sẵn",
                "- Account test đã có thông tin (VD: số điện thoại đã được lưu trong hệ thống)\n- Đang ở màn hình Checkout UltraFast",
                "1. Quan sát Block \"Thông tin khách hàng\" sau khi Checkout load xong\n2. Kiểm tra từng field hiển thị trong block",
                "- Block \"Thông tin khách hàng\" hiển thị đúng các field đã có data trong hệ thống (VD: SĐT đã có → hiển thị SĐT)\n- Block không bị ẩn",
                "Y"
            ),
            (
                "Medium",
                "Kiểm tra Block Thông tin khách hàng hiển thị rỗng khi chưa có data lắp đặt",
                "- Account test chưa có thông tin lắp đặt\n- Đang ở màn hình Checkout UltraFast",
                "1. Quan sát Block \"Thông tin khách hàng\" sau khi Checkout load xong",
                "- Block \"Thông tin khách hàng\" hiển thị trạng thái rỗng (các field không có dữ liệu)\n- Block vẫn hiển thị trên trang (không bị ẩn)",
                "Y"
            ),
            (
                "High",
                "Kiểm tra \"Cần thanh toán\" hiển thị đúng tổng tiền sản phẩm (không có voucher)",
                "- Đã chọn gói UltraFast chu kỳ X giá Y\n- Không áp dụng voucher\n- Đang ở màn hình Checkout UltraFast",
                "1. Quan sát dòng \"Cần thanh toán\" sau khi Checkout load xong\n2. Kiểm tra số tiền hiển thị\n3. So sánh với giá gói đã chọn",
                "- Dòng \"Cần thanh toán\" hiển thị đúng tổng tiền Y = giá gói đã chọn\n- Không có khoản giảm trừ nào",
                "Y"
            ),
        ]
    },
    {
        "name": "Nhóm 5: Button Thanh toán — Validate & Execute",
        "tcs": [
            (
                "High",
                "Kiểm tra button Thanh toán không thực hiện khi còn trường bắt buộc chưa nhập",
                "- Đang ở màn hình Checkout UltraFast\n- Trường Số điện thoại đang để trống\n- Đã chọn một PTTT Online",
                "1. Để trống trường \"Số điện thoại\"\n2. Click button \"Thanh toán\"",
                "- Hệ thống KHÔNG chuyển hướng sang trang 3rd party\n- Hiển thị thông báo lỗi validate cho trường bắt buộc chưa nhập (border đỏ + thông báo lỗi tương ứng)\n- Trang vẫn ở màn hình Checkout",
                "Y"
            ),
            (
                "High",
                "Kiểm tra button Thanh toán báo lỗi khi chính sách QLCS đã không còn active",
                "- Đang ở màn hình Checkout UltraFast\n- Đã nhập SĐT hợp lệ, đã chọn PTTT Online\n- Chính sách gói UltraFast trên QLCS đã bị deactivate (cần setup trên QLCS trước khi test)\nDữ liệu test: gói UltraFast đã bị tắt chính sách trên QLCS",
                "1. Điền đầy đủ thông tin hợp lệ: SĐT, PTTT\n2. Click button \"Thanh toán\"",
                "- Hệ thống KHÔNG chuyển hướng sang trang 3rd party\n- Hiển thị thông báo lỗi (chính sách không còn active / không thể thực hiện thanh toán)\n- Đơn hàng không được tạo",
                "N"
            ),
            (
                "High",
                "Kiểm tra button Thanh toán redirect đúng trang 3rd party khi data hợp lệ",
                "- Đang ở màn hình Checkout UltraFast\n- Đã nhập SĐT hợp lệ: 0901234567\n- Đã chọn PTTT Online (VD: Thẻ ATM)\n- Chính sách gói còn active trên QLCS",
                "1. Nhập SĐT hợp lệ vào trường Số điện thoại (VD: 0901234567)\n2. Chọn PTTT Online (VD: Thẻ ATM)\n3. Click button \"Thanh toán\"",
                "- Hệ thống validate thành công\n- Redirect sang trang thanh toán của nhà cung cấp 3rd party tương ứng với PTTT đã chọn (VD: cổng ATM bank)",
                "Y"
            ),
        ]
    },
    {
        "name": "Nhóm 6: Luồng thanh toán Online (3rd party)",
        "tcs": [
            (
                "High",
                "Kiểm tra màn hình hoàn tất đơn hàng hiển thị trạng thái \"Đã thanh toán\" khi TT thành công",
                "- Đã được redirect sang trang thanh toán 3rd party (VD: cổng ATM)\n- Có thông tin thẻ ATM/Visa hợp lệ để test\nLưu ý: Momo/VietQR/Zalopay cần app mobile để thực hiện",
                "1. Tại trang thanh toán 3rd party, nhập thông tin thẻ hợp lệ\n2. Xác nhận thanh toán thành công trên trang 3rd party\n3. Quan sát màn hình sau khi hệ thống nhận tín hiệu callback",
                "- Hệ thống điều hướng về màn hình \"Hoàn tất đơn hàng\"\n- Trạng thái đơn hàng: \"Đã thanh toán\" (bước hoàn tất màu xanh lá)\n- Không đẩy khoản thu vào Phiếu thi công (PTC)",
                "N"
            ),
            (
                "High",
                "Kiểm tra quay về màn hình Checkout khi hủy thanh toán tại 3rd party",
                "- Đã được redirect sang trang thanh toán 3rd party\n- Chưa thực hiện thanh toán",
                "1. Tại trang 3rd party, click button \"Hủy\" hoặc \"Cancel\"\n   Hoặc: nhấn nút Back trên browser để quay về",
                "- Hệ thống điều hướng về màn hình Checkout UltraFast\n- Đơn hàng chưa được tạo / chưa có giao dịch",
                "N"
            ),
            (
                "Medium",
                "Kiểm tra chỉ trường PTTT có thể thay đổi sau khi back từ 3rd party",
                "- Đã back về màn hình Checkout UltraFast sau khi hủy/back từ trang 3rd party",
                "1. Quan sát trạng thái các trường trên form Checkout\n2. Cố gắng sửa trường Số điện thoại\n3. Cố gắng thay đổi PTTT\n4. Click button \"Thanh toán\" sau khi đổi PTTT",
                "- Trường Số điện thoại và các trường khác (ngoài PTTT) bị disabled, không thể chỉnh sửa\n- Trường PTTT có thể thay đổi (chọn PTTT khác)\n- Sau khi đổi PTTT và click \"Thanh toán\" → hệ thống xử lý bình thường",
                "N"
            ),
        ]
    },
    {
        "name": "Nhóm 7: Navigation & UI",
        "tcs": [
            (
                "Medium",
                "Kiểm tra button \"Quay lại\" điều hướng về màn hình Chi tiết gói",
                "- Đang ở màn hình Checkout UltraFast",
                "1. Click button \"Quay lại\" ở màn hình Checkout",
                "Hệ thống điều hướng về màn hình Chi tiết gói UltraFast (màn hình trước đó)",
                "Y"
            ),
            (
                "Low",
                "Kiểm tra click Logo FPT điều hướng về trang chủ FPT.vn",
                "- Đang ở màn hình Checkout UltraFast\n- Logo FPT hiển thị ở header",
                "1. Click vào Logo FPT ở header của màn hình Checkout",
                "Hệ thống điều hướng về trang chủ FPT.vn",
                "Y"
            ),
            (
                "Low",
                "Kiểm tra click text điều khoản điều hướng đến trang Privacy Policy",
                "- Đang ở màn hình Checkout UltraFast\n- Text \"Bằng việc nhấn vào nút Thanh toán bạn đã đồng ý với các điều khoản của FPT Telecom.\" hiển thị",
                "1. Quan sát vị trí text điều khoản phía dưới button Thanh toán\n2. Click vào phần text điều khoản (link)",
                "Hệ thống điều hướng đến trang: https://fpt.vn/shop/privacy-policy",
                "Y"
            ),
        ]
    },
    {
        "name": "Nhóm 8: Blocked — Chờ confirm / Chưa implement",
        "tcs": [
            (
                "Medium",
                "[BLOCKED] Kiểm tra \"Cần thanh toán\" hiển thị số tiền đã trừ voucher khi có áp dụng voucher",
                "- Đang ở màn hình Checkout UltraFast\n- Voucher đã được áp dụng tại bước thanh toán",
                "[BLOCKED – Tính năng voucher tại bước thanh toán chưa implement (CLARY-DANGKYUF-005). TC sẽ được kích hoạt trong sprint sau khi feature hoàn thiện.]",
                "[BLOCKED – Chờ implement tính năng voucher tại bước thanh toán]",
                "N"
            ),
            (
                "Medium",
                "[BLOCKED] Kiểm tra hành vi khi truy cập màn hình Checkout khi chưa xác thực (nếu yêu cầu đăng nhập)",
                "- Chưa đăng nhập hoặc session đã hết hạn\n- URL màn hình Checkout UltraFast",
                "[BLOCKED – Spec không định nghĩa behavior khi chưa authenticate. Cần BA/Dev confirm: (1) Checkout có yêu cầu đăng nhập không? (2) Nếu có → redirect về trang đăng nhập hay hiển thị lỗi?]",
                "[BLOCKED – Cần confirm: liệu màn hình Checkout có yêu cầu xác thực không và behavior khi chưa login]",
                "N"
            ),
            (
                "Medium",
                "[BLOCKED] Kiểm tra layout màn hình Checkout UltraFast trên thiết bị mobile (≤768px)",
                "- Trình duyệt ở chế độ mobile hoặc thiết bị thực ≤768px\n- Gói UltraFast còn active",
                "[BLOCKED – Spec không mô tả behavior responsive riêng biệt cho mobile. Cần BA/Dev confirm: có yêu cầu responsive hay không? Layout mobile có khác desktop không?]",
                "[BLOCKED – Cần confirm behavior responsive trên mobile ≤768px]",
                "N"
            ),
        ]
    },
]


def make_sheet(wb, sheet_name, func_id, func_name):
    # Remove if exists
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ── Column widths ──────────────────────────────────────────────────────
    widths = {"A":6,"B":14,"C":10,"D":45,"E":45,"F":50,"G":55,"H":8,
              "I":12,"J":15,"K":12,"L":20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Row 1-2: empty spacer ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 6

    # ── Row 3: Function ID ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 20
    ws['C3'] = "Mã chức năng:"
    ws['C3'].font = BOLD_FONT
    ws['D3'] = func_id
    ws['D3'].font = BOLD_FONT

    # ── Row 4: Function Name ───────────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    ws['C4'] = "Tên chức năng:"
    ws['C4'].font = BOLD_FONT
    ws['D4'] = func_name
    ws['D4'].font = BOLD_FONT

    # ── Row 5-6: empty ─────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 6
    ws.row_dimensions[6].height = 6

    # ── Row 7-8: Header ────────────────────────────────────────────────────
    headers_r7 = ["QC/AI","Testcase ID","Mức Độ Ưu Tiên",
                  "Nội Dung Test","Điều Kiện / Dữ Liệu Test",
                  "Các Bước Thực Hiện","Kết Quả Mong Đợi",
                  "Có thể\nTự Động Hóa","Round 1","","",""]
    cols = list("ABCDEFGHIJKL")

    for i, (col, hdr) in enumerate(zip(cols, headers_r7)):
        cell = ws[f"{col}7"]
        cell.value = hdr
        cell.fill  = BLUE_FILL
        cell.font  = WHITE_FONT
        cell.alignment = CTR_ALIGN
        cell.border = BORDER

    # Merge A-H rows 7-8
    for col in list("ABCDEFGH"):
        ws.merge_cells(f"{col}7:{col}8")

    # Round 1 merge I7:L7
    ws.merge_cells("I7:L7")

    # Sub-headers row 8
    sub = {"I":"Kết Quả Thực Hiện","J":"Người Thực Hiện",
           "K":"ID Bugs","L":"Ghi Chú"}
    for col, txt in sub.items():
        cell = ws[f"{col}8"]
        cell.value = txt
        cell.fill  = BLUE_FILL
        cell.font  = WHITE_FONT
        cell.alignment = CTR_ALIGN
        cell.border = BORDER

    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 20

    # ── TC rows ────────────────────────────────────────────────────────────
    row_idx = 9
    tc_seq  = 0

    for group in GROUPS:
        # Group header row
        ws.row_dimensions[row_idx].height = 18
        gh = ws.cell(row=row_idx, column=2)
        gh.value     = group["name"]
        gh.fill      = GREEN_FILL
        gh.font      = BOLD_FONT
        gh.alignment = Alignment(vertical="center", wrap_text=True)
        gh.border    = BORDER
        ws.merge_cells(f"B{row_idx}:L{row_idx}")
        # Fill A with green too
        ws.cell(row=row_idx, column=1).fill   = GREEN_FILL
        ws.cell(row=row_idx, column=1).border = BORDER
        row_idx += 1

        for tc in group["tcs"]:
            priority, title, pre, steps, expected, auto = tc
            tc_seq += 1
            tc_id = f"{func_id}.{tc_seq}"

            ws.row_dimensions[row_idx].height = 80

            data = ["AI", tc_id, priority, title, pre, steps, expected, auto,
                    "", "", "", ""]
            for j, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=j, value=val)
                cell.border    = BORDER
                cell.alignment = Alignment(wrap_text=True, vertical="top",
                                           horizontal="center" if j in (1,2,3,8,9,10,11,12) else "left")
                if j == 1:
                    cell.font = Font(name="Calibri", size=11, color="7030A0", bold=True)
                elif j in (2, 3):
                    cell.font = BOLD_FONT
                else:
                    cell.font = NORM_FONT

            row_idx += 1

    # Freeze panes below header
    ws.freeze_panes = "A9"
    return ws


# ── Main ──────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(TC_FILE)
make_sheet(wb, SHEET_NAME, FUNC_ID, FUNC_NAME)
wb.save(TC_FILE)

# Count TCs
total = sum(len(g["tcs"]) for g in GROUPS)
high  = sum(1 for g in GROUPS for tc in g["tcs"] if tc[0]=="High")
med   = sum(1 for g in GROUPS for tc in g["tcs"] if tc[0]=="Medium")
low   = sum(1 for g in GROUPS for tc in g["tcs"] if tc[0]=="Low")
blocked = sum(1 for g in GROUPS for tc in g["tcs"] if "[BLOCKED]" in tc[1])
auto_y  = sum(1 for g in GROUPS for tc in g["tcs"] if tc[5]=="Y")
auto_n  = sum(1 for g in GROUPS for tc in g["tcs"] if tc[5]=="N")

import io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
print(f"[OK] Sheet '{SHEET_NAME}' da duoc them vao {TC_FILE}")
print(f"   Tong TC: {total} | High:{high} Medium:{med} Low:{low}")
print(f"   Blocked: {blocked} | Auto Y:{auto_y} N:{auto_n}")
