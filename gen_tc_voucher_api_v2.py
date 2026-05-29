"""
Generate TC API Excel for chucnang_Voucher — ECP_API_Documentation_v4
4 endpoints: voucher/list, voucher/content, voucher/apply, voucher/check
Output: ecom-pdh/03_test-cases/api/AI_ISC_ecom-pdh_v1.1_TC_API_v2.0.xlsx
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

# ── Styles ──────────────────────────────────────────────────────────────────

def make_border(sides="all"):
    thin = Side(style="thin")
    none = Side(style=None)
    return Border(
        left=thin if sides in ("all", "left") else none,
        right=thin if sides in ("all", "right") else none,
        top=thin if sides in ("all", "top") else none,
        bottom=thin if sides in ("all", "bottom") else none,
    )

# Color palette
C_HEADER_INFO   = "4472C4"   # Blue — API info header
C_COL_HEADER    = "2F75B6"   # Darker blue — column names
C_GROUP_AUTH    = "C6EFCE"   # Green — auth group
C_GROUP_VALID   = "FFEB9C"   # Yellow — validation
C_GROUP_FLOW    = "DEEBF7"   # Light blue — business flow
C_GROUP_ERR     = "FCE4D6"   # Orange-red — error handling
C_GROUP_SEC     = "E2EFDA"   # Light green — security
C_P_HIGH        = "FF0000"   # Red text
C_P_MEDIUM      = "FF8C00"   # Orange text
C_P_LOW         = "70AD47"   # Green text
C_ROUND_HEADER  = "A9D08E"   # Green — round header

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

BASE_URL = "http://ecp-api-stag.fpt.net/ordering"

STANDARD_STEPS = "1. Gởi request theo điều kiện test\n2. Kiểm tra response trả về"

# ── TC Data definitions ──────────────────────────────────────────────────────

# Each TC: (priority, title, condition, steps, expected)
# Groups: ("GROUP", group_name, color_key)
# color_key: auth | valid | flow | err | sec

APIS = [
    {
        "code": "API_16",
        "name": "Lấy danh sách Voucher khả dụng",
        "method": "POST",
        "endpoint": "/public/v1/voucher/list",
        "sheet": "16_Danh sách Voucher",
        "tcs": [
            ("GROUP", "Authentication / Authorization", "auth"),
            ("High",
             "Lấy danh sách voucher thành công với X-Checkout-Token hợp lệ",
             "Header: X-Checkout-Token = <token hợp lệ>, Accept-Language = vi\nCheckout session đang active và đã chọn phương thức thanh toán",
             "1. Chuẩn bị checkout session có payment method\n2. Gửi POST /public/v1/voucher/list với X-Checkout-Token hợp lệ",
             "HTTP 200\nsuccess = true\ndata[] không rỗng, mỗi item có: voucher_code, description, to_date (format dd/MM/yyyy), voucher_type (General/Individual/Unknown), register_type_id, policy_group_id, apply_type_id, promotion_type_id\nerror = null"),
            ("High",
             "Thiếu header X-Checkout-Token",
             "Không gửi header X-Checkout-Token",
             "1. Gửi POST /public/v1/voucher/list không có header X-Checkout-Token",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\"\nerror.message = \"Phiên đặt hàng không hợp lệ. Vui lòng thử lại.\"\ndata = null"),
            ("High",
             "X-Checkout-Token = empty string",
             "Header: X-Checkout-Token = \"\" (chuỗi rỗng)",
             "1. Gửi POST /public/v1/voucher/list với X-Checkout-Token = \"\"",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\" hoặc \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
            ("High",
             "X-Checkout-Token không hợp lệ (sai format / random string)",
             "Header: X-Checkout-Token = \"invalid_token_xyz_123\"",
             "1. Gửi POST /public/v1/voucher/list với token sai format",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nerror.message = \"Phiên đặt hàng không hợp lệ. Vui lòng thử lại.\"\ndata = null"),
            ("High",
             "X-Checkout-Token hết hạn",
             "Header: X-Checkout-Token = token của session đã hết hạn",
             "1. Lấy token cũ của checkout session đã hết hạn\n2. Gửi POST /public/v1/voucher/list với token đó",
             "HTTP 400 hoặc 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\" hoặc \"CHECKOUT_NOT_FOUND\"\ndata = null"),
            ("Medium",
             "[SECURITY-INFERRED] X-Checkout-Token bị giả mạo (tampered token)",
             "Header: X-Checkout-Token = token hợp lệ nhưng đã chỉnh sửa 1-2 ký tự",
             "1. Lấy token hợp lệ\n2. Thay đổi 1 ký tự ở cuối token\n3. Gửi POST /public/v1/voucher/list với token đã sửa",
             "HTTP 400 hoặc 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nToken bị giả mạo không được chấp nhận"),
            ("Medium",
             "[SECURITY-INFERRED] Client-Id không hợp lệ kết hợp với X-Checkout-Token hợp lệ",
             "Header: X-Checkout-Token = valid token, Client-Id = \"invalid_client_999\"",
             "1. Gửi POST /public/v1/voucher/list với X-Checkout-Token hợp lệ và Client-Id sai",
             "HTTP 400 hoặc 401 hoặc 200 (tuỳ config)\nsuccess = false nếu Client-Id bắt buộc\nGhi chú: xác nhận behavior thực tế từ BA/Dev"),
            ("Medium",
             "[SECURITY-INFERRED] Replay attack — gửi lại X-Checkout-Token của session đã completed",
             "Header: X-Checkout-Token = token của checkout đã hoàn thành/đã submit đơn",
             "1. Hoàn thành 1 checkout session\n2. Lưu lại checkout token\n3. Gửi POST /public/v1/voucher/list với token cũ đó",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\" hoặc \"CHECKOUT_TOKEN_INVALID\"\nToken của session đã kết thúc không được reuse"),

            ("GROUP", "Business Flow — Happy Path", "flow"),
            ("High",
             "Lấy danh sách voucher thành công khi có nhiều voucher khả dụng",
             "X-Checkout-Token hợp lệ, checkout đã chọn payment\nCheckout này có ít nhất 2 voucher khả dụng",
             "1. Chuẩn bị checkout có nhiều voucher\n2. Gửi POST /public/v1/voucher/list",
             "HTTP 200\nsuccess = true\ndata[] có ít nhất 2 items\nMỗi item có đủ 9 fields: voucher_code, description, note, to_date, register_type_id, voucher_type, policy_group_id, apply_type_id, promotion_type_id"),
            ("Medium",
             "Trả về empty array khi không có voucher khả dụng",
             "X-Checkout-Token hợp lệ, checkout hợp lệ\nCheckout này không có voucher nào đủ điều kiện",
             "1. Chuẩn bị checkout không có voucher khả dụng\n2. Gửi POST /public/v1/voucher/list",
             "HTTP 200\nsuccess = true\ndata = [] (empty array)\nmeta.pagination = null"),

            ("GROUP", "Business Flow — Edge Cases", "flow"),
            ("Medium",
             "Checkout không tồn tại trong hệ thống",
             "Header: X-Checkout-Token = token của checkout không tồn tại trong DB",
             "1. Tạo token giả không tham chiếu đến checkout nào\n2. Gửi POST /public/v1/voucher/list",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\"\nerror.message = \"Không tìm thấy phiên đặt hàng.\""),
            ("Medium",
             "Checkout chưa chọn phương thức thanh toán",
             "X-Checkout-Token hợp lệ, checkout tồn tại nhưng chưa chọn payment method",
             "1. Tạo checkout session chưa chọn payment\n2. Gửi POST /public/v1/voucher/list",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_PAYMENT_REQUIRED\"\nerror.message = \"Vui lòng chọn phương thức thanh toán.\""),
            ("Medium",
             "Verify format to_date và voucher_type trong response",
             "X-Checkout-Token hợp lệ, có voucher khả dụng",
             "1. Gửi POST /public/v1/voucher/list thành công\n2. Kiểm tra từng field trong data[]",
             "HTTP 200\ndata[].to_date có format dd/MM/yyyy (VD: 31/12/2026)\ndata[].voucher_type nằm trong enum: Unknown, General, Individual\ndata[].voucher_code không null và không rỗng"),

            ("GROUP", "Error Handling", "err"),
            ("Low",
             "Server lỗi nội bộ",
             "Mô phỏng lỗi server (cần môi trường test đặc biệt)",
             "1. Gửi request đúng format khi server có lỗi nội bộ",
             "HTTP 400 hoặc 500\nsuccess = false\nerror.code = \"BUSINESS_INTERNAL_ERROR\"\nerror.message = \"Đã xảy ra lỗi. Vui lòng thử lại sau.\""),
            ("Low",
             "Accept-Language không hợp lệ → mặc định ngôn ngữ vi",
             "Header: Accept-Language = \"fr\" (không được hỗ trợ), X-Checkout-Token hợp lệ",
             "1. Gửi POST /public/v1/voucher/list với Accept-Language = \"fr\"",
             "HTTP 200 hoặc 400\nNếu thành công: response trả về tiếng Việt (default)\nKhông trả về 500"),
        ]
    },
    {
        "code": "API_17",
        "name": "Lấy nội dung chi tiết Voucher",
        "method": "POST",
        "endpoint": "/public/v1/voucher/content",
        "sheet": "17_Nội dung Voucher",
        "tcs": [
            ("GROUP", "Authentication / Authorization", "auth"),
            ("High",
             "Thiếu header X-Checkout-Token",
             "Request body: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}\nKhông có header X-Checkout-Token",
             "1. Gửi POST /public/v1/voucher/content không có X-Checkout-Token",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\"\nerror.message = \"Phiên đặt hàng không hợp lệ. Vui lòng thử lại.\"\ndata = null"),
            ("High",
             "X-Checkout-Token = empty string",
             "X-Checkout-Token = \"\"\nRequest body: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
             "1. Gửi POST /public/v1/voucher/content với X-Checkout-Token rỗng",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\" hoặc \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
            ("High",
             "X-Checkout-Token không hợp lệ (sai format)",
             "X-Checkout-Token = \"bad_token_xyz\"\nRequest body: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
             "1. Gửi POST /public/v1/voucher/content với token sai",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nerror.message = \"Phiên đặt hàng không hợp lệ. Vui lòng thử lại.\""),
            ("High",
             "X-Checkout-Token hết hạn",
             "X-Checkout-Token = token của session đã hết hạn\nRequest body hợp lệ",
             "1. Lấy token session cũ đã hết hạn\n2. Gửi POST /public/v1/voucher/content",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\" hoặc \"CHECKOUT_NOT_FOUND\"\ndata = null"),
            ("Medium",
             "Checkout không tồn tại",
             "X-Checkout-Token = token không ánh xạ checkout nào\nRequest body hợp lệ",
             "1. Gửi POST /public/v1/voucher/content với token không tồn tại",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\"\nerror.message = \"Không tìm thấy phiên đặt hàng.\""),
            ("Medium",
             "[SECURITY-INFERRED] X-Checkout-Token bị giả mạo",
             "X-Checkout-Token = token hợp lệ đã sửa 1 ký tự\nRequest body hợp lệ",
             "1. Sửa 1 ký tự của valid token\n2. Gửi POST /public/v1/voucher/content",
             "HTTP 400 hoặc 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nToken giả mạo bị từ chối"),
            ("Medium",
             "[SECURITY-INFERRED] Client-Id sai + X-Checkout-Token hợp lệ",
             "Client-Id = \"tampered_client\", X-Checkout-Token hợp lệ\nRequest body hợp lệ",
             "1. Gửi POST /public/v1/voucher/content với Client-Id không hợp lệ",
             "HTTP 400 hoặc 401 hoặc 200 (tuỳ config)\nXác nhận behavior thực tế"),
            ("Medium",
             "[SECURITY-INFERRED] Replay token của session đã kết thúc",
             "X-Checkout-Token = token của checkout đã hoàn thành đơn hàng\nRequest body hợp lệ",
             "1. Hoàn thành 1 checkout\n2. Gửi POST /public/v1/voucher/content với token cũ",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\" hoặc \"CHECKOUT_TOKEN_INVALID\""),

            ("GROUP", "Validation — Required Fields", "valid"),
            ("High",
             "Thiếu field voucher_code trong request body",
             "X-Checkout-Token hợp lệ\nRequest body: {} (không có voucher_code)",
             "1. Gửi POST /public/v1/voucher/content với body rỗng {}",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\nerror.message = \"Mã voucher là bắt buộc.\"\ndata = null"),
            ("High",
             "voucher_code = null",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": null}",
             "1. Gửi POST /public/v1/voucher/content với voucher_code = null",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\nerror.message = \"Mã voucher là bắt buộc.\"\ndata = null"),

            ("GROUP", "Validation — Format / Boundary", "valid"),
            ("Medium",
             "voucher_code = empty string (vi phạm Min=1)",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"\"}",
             "1. Gửi POST /public/v1/voucher/content với voucher_code = \"\"",
             "HTTP 400\nsuccess = false\nLỗi validate: voucher_code không được rỗng\nerror.code = \"VOUCHER_CODE_REQUIRED_400\" hoặc validation error"),
            ("Medium",
             "voucher_code = 1 ký tự (giá trị biên Min=1)",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"A\"}",
             "1. Gửi POST /public/v1/voucher/content với voucher_code = \"A\"",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\" (không tồn tại)\nKhông bị lỗi format, chỉ lỗi nghiệp vụ"),
            ("Medium",
             "[SECURITY-INFERRED] voucher_code chứa ký tự injection (SQL/NoSQL injection attempt)",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"' OR 1=1 --\"}",
             "1. Gửi POST /public/v1/voucher/content với voucher_code chứa SQL injection",
             "HTTP 400\nsuccess = false\nServer không bị ảnh hưởng bởi injection\nKhông trả về data nhạy cảm\nerror.code = \"VOUCHER_INVALID\" hoặc validation error"),

            ("GROUP", "Business Flow — Happy Path", "flow"),
            ("High",
             "Lấy nội dung chi tiết voucher thành công",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}\nvoucher_code tồn tại và có nội dung hiển thị",
             "1. Chuẩn bị voucher code có content\n2. Gửi POST /public/v1/voucher/content",
             "HTTP 200\nsuccess = true\ndata.voucher_code = \"CA21060100KTHIETBIKHOFG039\"\ndata.content1 có giá trị (non-null)\ndata.content2~content6 có giá trị hoặc null\nerror = null"),
            ("Medium",
             "Voucher không có nội dung hiển thị → data = null",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"<mã voucher không có content>\"}\nvoucher_code tồn tại nhưng chưa có dữ liệu content",
             "1. Chuẩn bị voucher không có content\n2. Gửi POST /public/v1/voucher/content",
             "HTTP 200\nsuccess = true\ndata = null\nerror = null\nKhông trả về 404 hay 400"),

            ("GROUP", "Business Flow — Edge Cases", "flow"),
            ("Medium",
             "voucher_code không tồn tại trong hệ thống",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"INVALID_CODE_99999\"}",
             "1. Gửi POST /public/v1/voucher/content với mã không tồn tại",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\"\nerror.message = \"Mã voucher không hợp lệ hoặc đã hết hạn.\""),
            ("Medium",
             "voucher_code đã hết hạn",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"<mã voucher hết hạn>\"}",
             "1. Chuẩn bị mã voucher đã hết hạn\n2. Gửi POST /public/v1/voucher/content",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\"\nerror.message = \"Mã voucher không hợp lệ hoặc đã hết hạn.\""),
            ("Medium",
             "Verify data.voucher_code trong response khớp với request",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}",
             "1. Gửi POST /public/v1/voucher/content thành công\n2. So sánh data.voucher_code với voucher_code đã gửi",
             "HTTP 200\nsuccess = true\ndata.voucher_code == request.voucher_code\nmeta.request_id, meta.trace_id, meta.timestamp có giá trị hợp lệ"),
        ]
    },
    {
        "code": "API_18",
        "name": "Áp dụng eVouchers cho Checkout",
        "method": "POST",
        "endpoint": "/public/v1/voucher/apply",
        "sheet": "18_Áp dụng Voucher",
        "tcs": [
            ("GROUP", "Authentication / Authorization", "auth"),
            ("High",
             "Thiếu header X-Checkout-Token",
             "Request body: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}]}\nKhông có X-Checkout-Token",
             "1. Gửi POST /public/v1/voucher/apply không có X-Checkout-Token",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\"\nerror.message = \"Phiên đặt hàng không hợp lệ. Vui lòng thử lại.\"\ndata = null"),
            ("High",
             "X-Checkout-Token = empty string",
             "X-Checkout-Token = \"\"\nRequest body hợp lệ",
             "1. Gửi POST /public/v1/voucher/apply với X-Checkout-Token rỗng",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\" hoặc \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
            ("High",
             "X-Checkout-Token không hợp lệ (sai format)",
             "X-Checkout-Token = \"bad_token_xyz\"\nRequest body hợp lệ",
             "1. Gửi POST /public/v1/voucher/apply với token sai format",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nerror.message = \"Phiên đặt hàng không hợp lệ. Vui lòng thử lại.\""),
            ("High",
             "X-Checkout-Token hết hạn",
             "X-Checkout-Token = token hết hạn\nRequest body hợp lệ",
             "1. Lấy token đã hết hạn\n2. Gửi POST /public/v1/voucher/apply",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\" hoặc \"CHECKOUT_NOT_FOUND\"\ndata = null"),
            ("Medium",
             "Checkout không tồn tại",
             "X-Checkout-Token = token không ánh xạ checkout nào\nRequest body hợp lệ",
             "1. Gửi POST /public/v1/voucher/apply với token không tồn tại",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\"\nerror.message = \"Không tìm thấy phiên đặt hàng.\""),
            ("Medium",
             "[SECURITY-INFERRED] X-Checkout-Token bị giả mạo",
             "X-Checkout-Token = valid token đã sửa 1 ký tự\nRequest body hợp lệ",
             "1. Chỉnh sửa 1 ký tự của valid token\n2. Gửi POST /public/v1/voucher/apply",
             "HTTP 400 hoặc 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\nToken giả mạo bị từ chối, không apply được voucher"),
            ("Medium",
             "[SECURITY-INFERRED] Client-Id không hợp lệ + X-Checkout-Token hợp lệ",
             "Client-Id = \"tampered_id\", X-Checkout-Token hợp lệ\nRequest body hợp lệ",
             "1. Gửi POST /public/v1/voucher/apply với Client-Id sai",
             "HTTP 400 hoặc 401 hoặc 200\nXác nhận behavior khi Client-Id sai"),
            ("Medium",
             "[SECURITY-INFERRED] Replay attack — gửi lại request apply đã xử lý thành công",
             "X-Checkout-Token hợp lệ\nGửi lại request apply đã thực thi thành công trước đó",
             "1. Apply voucher thành công lần đầu\n2. Gửi lại đúng request đó lần 2",
             "HTTP 200 hoặc 400\nNếu idempotent: trả về cùng kết quả\nKhông tạo duplicate apply\nerror.retryable = false nếu lỗi"),

            ("GROUP", "Validation — Required Fields", "valid"),
            ("High",
             "vouchers[].voucher_code thiếu trong item",
             "X-Checkout-Token hợp lệ\nRequest body: {\"vouchers\": [{\"voucher_type\": \"General\"}]} (thiếu voucher_code)",
             "1. Gửi POST /public/v1/voucher/apply với item không có voucher_code",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\nerror.message = \"Mã voucher là bắt buộc.\"\ndata = null"),
            ("High",
             "vouchers[].voucher_code = null",
             "X-Checkout-Token hợp lệ\nRequest body: {\"vouchers\": [{\"voucher_code\": null, \"voucher_type\": \"General\"}]}",
             "1. Gửi POST /public/v1/voucher/apply với voucher_code = null",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\ndata = null"),

            ("GROUP", "Validation — Format / Type / Boundary", "valid"),
            ("Medium",
             "vouchers[].voucher_code = empty string (vi phạm Min=1)",
             "X-Checkout-Token hợp lệ\nRequest body: {\"vouchers\": [{\"voucher_code\": \"\"}]}",
             "1. Gửi POST /public/v1/voucher/apply với voucher_code rỗng",
             "HTTP 400\nsuccess = false\nLỗi validate voucher_code\nerror.code chứa \"VOUCHER_CODE_REQUIRED\" hoặc validation error"),
            ("Medium",
             "vouchers[].voucher_type = giá trị không trong enum",
             "X-Checkout-Token hợp lệ\nRequest body: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"Premium\"}]}",
             "1. Gửi POST /public/v1/voucher/apply với voucher_type không hợp lệ",
             "HTTP 400\nsuccess = false\nLỗi validate voucher_type\nVoucher_type chỉ chấp nhận: Unknown, General, Individual"),
            ("Medium",
             "vouchers[].voucher_code = 1 ký tự (boundary Min=1)",
             "X-Checkout-Token hợp lệ\nRequest body: {\"vouchers\": [{\"voucher_code\": \"A\"}]}",
             "1. Gửi POST /public/v1/voucher/apply với voucher_code = \"A\"",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\" (không tồn tại)\nKhông bị lỗi format"),

            ("GROUP", "Business Flow — Happy Path", "flow"),
            ("High",
             "Áp dụng 1 voucher General thành công",
             "X-Checkout-Token hợp lệ, checkout đã chọn payment\nRequest body: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}]}",
             "1. Chuẩn bị checkout active với payment method\n2. Gửi POST /public/v1/voucher/apply",
             "HTTP 200\nsuccess = true\ndata[] có 1 item\ndata[0].voucher_code = \"CA21060100KTHIETBIKHOFG039\"\ndata[0].discount_value = 0\ndata[0].discount_ex_vat_value = 0\ndata[0].original_discount_value > 0\ndata[0].applies[] không rỗng\nerror = null"),
            ("High",
             "Áp dụng 2 voucher (General + Individual) thành công",
             "X-Checkout-Token hợp lệ\nRequest body: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}, {\"voucher_code\": \"CA23039F50GOLD\", \"voucher_type\": \"Individual\"}]}",
             "1. Chuẩn bị 2 voucher hợp lệ (1 General, 1 Individual)\n2. Gửi POST /public/v1/voucher/apply",
             "HTTP 200\nsuccess = true\ndata[] có đúng 2 items\nMỗi item có voucher_code tương ứng\ndata[0].discount_value = 0, data[1].discount_value = 0\nCả 2 items có applies[] không rỗng"),
            ("High",
             "Gỡ bỏ tất cả voucher bằng cách gửi vouchers = []",
             "X-Checkout-Token hợp lệ, checkout đang có voucher đã apply\nRequest body: {\"vouchers\": []}",
             "1. Apply voucher trước\n2. Gửi POST /public/v1/voucher/apply với vouchers = []",
             "HTTP 200\nsuccess = true\ndata = [] (empty array)\nerror = null\nTất cả voucher đã bị gỡ bỏ"),
            ("Medium",
             "Gỡ bỏ tất cả voucher khi vouchers = null",
             "X-Checkout-Token hợp lệ\nRequest body: {\"vouchers\": null}",
             "1. Gửi POST /public/v1/voucher/apply với vouchers = null",
             "HTTP 200, success = true, data = []\nHOẶC HTTP 400 (tuỳ validation)\nXác nhận behavior thực tế — xem CLA-3"),

            ("GROUP", "Business Flow — Edge Cases", "flow"),
            ("High",
             "Trùng mã voucher trong cùng một request",
             "X-Checkout-Token hợp lệ\nRequest body: {\"vouchers\": [{\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}, {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"}]}",
             "1. Gửi POST /public/v1/voucher/apply với 2 item có cùng voucher_code",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_DUPLICATE_400\"\nerror.message = \"Danh sách mã voucher bị trùng lặp. Vui lòng kiểm tra lại.\"\ndata = null"),
            ("High",
             "Voucher không hợp lệ hoặc đã hết hạn",
             "X-Checkout-Token hợp lệ\nRequest body: {\"vouchers\": [{\"voucher_code\": \"EXPIRED_VOUCHER_001\"}]}",
             "1. Chuẩn bị voucher code hết hạn\n2. Gửi POST /public/v1/voucher/apply",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\"\nerror.message = \"Mã voucher không hợp lệ hoặc đã hết hạn.\"\ndata = null"),
            ("High",
             "Rà soát voucher thất bại — điều kiện áp dụng không thỏa",
             "X-Checkout-Token hợp lệ\nRequest body: {\"vouchers\": [{\"voucher_code\": \"<voucher không đủ điều kiện cho checkout này>\"}]}",
             "1. Chuẩn bị voucher không đủ điều kiện áp dụng cho checkout\n2. Gửi POST /public/v1/voucher/apply",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_RECHECK_FAILED\"\nerror.message = \"Rà soát voucher thất bại. Vui lòng kiểm tra lại điều kiện áp dụng.\"\ndata = null"),
            ("Medium",
             "Checkout chưa chọn phương thức thanh toán",
             "X-Checkout-Token hợp lệ, checkout chưa chọn payment\nRequest body hợp lệ",
             "1. Chuẩn bị checkout chưa chọn payment method\n2. Gửi POST /public/v1/voucher/apply",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_PAYMENT_REQUIRED\"\nerror.message = \"Vui lòng chọn phương thức thanh toán.\""),
            ("Medium",
             "Verify business rule: discount_value và discount_ex_vat_value = 0 sau apply",
             "X-Checkout-Token hợp lệ\nVoucher hợp lệ được áp dụng thành công",
             "1. Gửi POST /public/v1/voucher/apply thành công\n2. Kiểm tra các trường discount trong response",
             "HTTP 200\nsuccess = true\ndata[].discount_value = 0\ndata[].discount_ex_vat_value = 0\ndata[].original_discount_value > 0 (giá trị gốc của voucher)\nGhi chú: discount_value = 0 vì tính sau bước calculate"),

            ("GROUP", "Error Handling", "err"),
            ("Medium",
             "Verify cấu trúc applies[] sau apply thành công",
             "X-Checkout-Token hợp lệ\nVoucher hợp lệ được apply thành công",
             "1. Gửi POST /public/v1/voucher/apply thành công\n2. Kiểm tra data[].applies[]",
             "HTTP 200\ndata[].applies[] không rỗng\nMỗi applies[] item có: service_id (int), sub_service_type_id (int), sub_service_id (int), service_code (int), discount_ex_vat (number), discount (number), dismonth (number), is_deduct_order (0 hoặc 1), original_discount_value (>0), original_discount_ex_vat (>0)"),
            ("Low",
             "Server lỗi nội bộ khi apply voucher",
             "Mô phỏng lỗi server (môi trường test đặc biệt)",
             "1. Gửi POST /public/v1/voucher/apply khi server có lỗi nội bộ",
             "HTTP 400 hoặc 500\nsuccess = false\nerror.code = \"BUSINESS_INTERNAL_ERROR\"\nerror.message = \"Đã xảy ra lỗi. Vui lòng thử lại sau.\"\nerror.retryable không null"),
        ]
    },
    {
        "code": "API_19",
        "name": "Kiểm tra tính hợp lệ Voucher",
        "method": "POST",
        "endpoint": "/public/v1/voucher/check",
        "sheet": "19_Kiểm tra Voucher",
        "tcs": [
            ("GROUP", "Authentication / Authorization", "auth"),
            ("High",
             "Thiếu header X-Checkout-Token",
             "Request body: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}\nKhông có X-Checkout-Token",
             "1. Gửi POST /public/v1/voucher/check không có X-Checkout-Token",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\"\nerror.message = \"Phiên đặt hàng không hợp lệ. Vui lòng thử lại.\""),
            ("High",
             "X-Checkout-Token = empty string",
             "X-Checkout-Token = \"\"\nRequest body hợp lệ",
             "1. Gửi POST /public/v1/voucher/check với X-Checkout-Token rỗng",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_REQUIRED\" hoặc \"CHECKOUT_TOKEN_INVALID\"\ndata = null hoặc data.is_valid = false"),
            ("High",
             "X-Checkout-Token không hợp lệ (sai format)",
             "X-Checkout-Token = \"bad_token_xyz\"\nRequest body hợp lệ",
             "1. Gửi POST /public/v1/voucher/check với token sai",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\"\ndata = null"),
            ("High",
             "X-Checkout-Token hết hạn",
             "X-Checkout-Token = token đã hết hạn\nRequest body hợp lệ",
             "1. Dùng token hết hạn\n2. Gửi POST /public/v1/voucher/check",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\" hoặc \"CHECKOUT_NOT_FOUND\"\ndata = null"),
            ("Medium",
             "Checkout không tồn tại",
             "X-Checkout-Token = token không ánh xạ checkout nào\nRequest body hợp lệ",
             "1. Gửi POST /public/v1/voucher/check với token không tồn tại",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\"\nerror.message = \"Không tìm thấy phiên đặt hàng.\""),
            ("Medium",
             "[SECURITY-INFERRED] X-Checkout-Token bị giả mạo",
             "X-Checkout-Token = valid token đã sửa 1 ký tự\nRequest body hợp lệ",
             "1. Sửa 1 ký tự của valid token\n2. Gửi POST /public/v1/voucher/check",
             "HTTP 400 hoặc 401\nsuccess = false\nerror.code = \"CHECKOUT_TOKEN_INVALID\""),
            ("Medium",
             "[SECURITY-INFERRED] Client-Id không hợp lệ + X-Checkout-Token hợp lệ",
             "Client-Id = \"tampered_id\", X-Checkout-Token hợp lệ\nRequest body hợp lệ",
             "1. Gửi POST /public/v1/voucher/check với Client-Id sai",
             "HTTP 400 hoặc 401 hoặc 200\nXác nhận behavior khi Client-Id không hợp lệ"),
            ("Medium",
             "[SECURITY-INFERRED] Replay attack — gửi lại check request với token cũ",
             "X-Checkout-Token = token hợp lệ của session đã kết thúc",
             "1. Gửi POST /public/v1/voucher/check với token của session đã kết thúc",
             "HTTP 400\nsuccess = false\nerror.code = \"CHECKOUT_NOT_FOUND\" hoặc \"CHECKOUT_TOKEN_INVALID\"\nToken replay bị từ chối"),

            ("GROUP", "Validation — Required Fields", "valid"),
            ("High",
             "Thiếu field voucher_code trong request body",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_type\": \"General\"} (không có voucher_code)",
             "1. Gửi POST /public/v1/voucher/check không có voucher_code",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\nerror.message = \"Mã voucher là bắt buộc.\"\ndata = null"),
            ("High",
             "voucher_code = null",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": null, \"voucher_type\": \"General\"}",
             "1. Gửi POST /public/v1/voucher/check với voucher_code = null",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_CODE_REQUIRED_400\"\ndata = null"),

            ("GROUP", "Validation — Format / Boundary", "valid"),
            ("Medium",
             "voucher_code = empty string (vi phạm Min=1)",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"\"}",
             "1. Gửi POST /public/v1/voucher/check với voucher_code = \"\"",
             "HTTP 400\nsuccess = false\nLỗi validate: voucher_code không được rỗng"),
            ("Medium",
             "voucher_code = 1 ký tự (giá trị biên Min=1)",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"A\"}",
             "1. Gửi POST /public/v1/voucher/check với voucher_code = \"A\"",
             "HTTP 400\nsuccess = false\nerror.code = \"VOUCHER_INVALID\" (không tồn tại)\nKhông bị lỗi format, qua được validation Min=1"),
            ("Medium",
             "voucher_type = giá trị không trong enum (strict validation)",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"Premium\"}",
             "1. Gửi POST /public/v1/voucher/check với voucher_type không hợp lệ",
             "HTTP 400\nsuccess = false\nLỗi validate voucher_type\nHOẶC HTTP 200 nếu server bỏ qua giá trị không hợp lệ (cần xác nhận)"),

            ("GROUP", "Business Flow — Happy Path", "flow"),
            ("High",
             "Check voucher hợp lệ, voucher_type = General",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"General\"}\nvoucher_code hợp lệ và còn hiệu lực",
             "1. Chuẩn bị voucher General hợp lệ\n2. Gửi POST /public/v1/voucher/check",
             "HTTP 200\nsuccess = true\ndata.is_valid = true\ndata.message = \"Voucher is valid.\"\nerror = null"),
            ("High",
             "Check voucher hợp lệ, voucher_type = Individual",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"<mã Individual hợp lệ>\", \"voucher_type\": \"Individual\"}\nvoucher_code Individual hợp lệ",
             "1. Chuẩn bị voucher Individual hợp lệ\n2. Gửi POST /public/v1/voucher/check",
             "HTTP 200\nsuccess = true\ndata.is_valid = true\ndata.message = \"Voucher is valid.\" hoặc message tương đương\nerror = null"),
            ("Medium",
             "Check voucher không gửi voucher_type (optional field)",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\"} (không có voucher_type)",
             "1. Gửi POST /public/v1/voucher/check chỉ với voucher_code",
             "HTTP 200\nsuccess = true\ndata.is_valid = true hoặc false (tuỳ voucher)\nKhông bị lỗi do thiếu voucher_type"),
            ("Medium",
             "Check voucher với voucher_type = Unknown",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"CA21060100KTHIETBIKHOFG039\", \"voucher_type\": \"Unknown\"}",
             "1. Gửi POST /public/v1/voucher/check với voucher_type = \"Unknown\"",
             "HTTP 200\nsuccess = true\ndata.is_valid = true hoặc false\nserver không reject Unknown type"),

            ("GROUP", "Business Flow — Edge Cases", "flow"),
            ("High",
             "Check voucher không hợp lệ (pending CLA-2 — xem Clarification)",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"INVALID_VOUCHER_CODE\", \"voucher_type\": \"General\"}\nvoucher_code không tồn tại",
             "1. Gửi POST /public/v1/voucher/check với voucher không hợp lệ\n\n⚠️ Pending CLA-2: Chưa rõ server trả về 400 VOUCHER_INVALID hay 200 + is_valid=false",
             "PENDING CLA-2 — Cần BA xác nhận:\nOption A: HTTP 400, success=false, error.code=\"VOUCHER_INVALID\"\nOption B: HTTP 200, success=true, data.is_valid=false, data.message có thông báo lỗi"),
            ("Medium",
             "Check voucher hết hạn",
             "X-Checkout-Token hợp lệ\nRequest body: {\"voucher_code\": \"<mã voucher hết hạn>\"}",
             "1. Chuẩn bị voucher đã hết hạn\n2. Gửi POST /public/v1/voucher/check",
             "HTTP 400 VOUCHER_INVALID\nHOẶC HTTP 200 + data.is_valid=false\n⚠️ Pending CLA-2 — cần xác nhận behavior"),

            ("GROUP", "Error Handling", "err"),
            ("Low",
             "Server lỗi nội bộ",
             "Mô phỏng lỗi server (môi trường đặc biệt)",
             "1. Gửi POST /public/v1/voucher/check khi server lỗi",
             "HTTP 400 hoặc 500\nsuccess = false\nerror.code = \"BUSINESS_INTERNAL_ERROR\"\nerror.message = \"Đã xảy ra lỗi. Vui lòng thử lại sau.\""),
        ]
    },
]

# ── Excel Generation ─────────────────────────────────────────────────────────

GROUP_COLORS = {
    "auth":  "C6EFCE",   # green
    "valid": "FFEB9C",   # yellow
    "flow":  "DEEBF7",   # light blue
    "err":   "FCE4D6",   # orange/red
    "sec":   "E2EFDA",   # light green
}

def build_sheet(ws, api_info):
    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8    # QC/AI
    ws.column_dimensions["B"].width = 12   # TC ID
    ws.column_dimensions["C"].width = 10   # Priority
    ws.column_dimensions["D"].width = 45   # Test Title
    ws.column_dimensions["E"].width = 45   # Điều kiện / Dữ liệu Test
    ws.column_dimensions["F"].width = 45   # Các Bước Thực Hiện
    ws.column_dimensions["G"].width = 55   # Expected Response
    ws.column_dimensions["H"].width = 14   # Round 1 — Result
    ws.column_dimensions["I"].width = 18   # Round 1 — Executed By
    ws.column_dimensions["J"].width = 14   # Round 1 — Bug ID
    ws.column_dimensions["K"].width = 22   # Round 1 — Remark

    # ── Row 1 — Sheet title ───────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    c = ws.cell(1, 1, "TEST CASE — API")
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    c.fill = fill(C_HEADER_INFO)
    c.alignment = align("left", "center")
    ws.merge_cells("A1:K1")

    # ── Rows 2-3 — Info block ─────────────────────────────────────────────────
    labels = ["Mã API", "Tên API", "Phương thức", "Base URL + Endpoint"]
    values = [
        api_info["code"],
        api_info["name"],
        api_info["method"],
        BASE_URL + api_info["endpoint"],
    ]
    for i, (lbl, val) in enumerate(zip(labels, values)):
        r = i + 2
        ws.row_dimensions[r].height = 18
        lc = ws.cell(r, 1, lbl)
        lc.font = font(bold=True, color="FFFFFF")
        lc.fill = fill(C_COL_HEADER)
        lc.alignment = align("center", "center")
        ws.merge_cells(f"A{r}:B{r}")

        vc = ws.cell(r, 3, val)
        vc.font = font(bold=True if i == 0 else False)
        vc.alignment = align("left", "center")
        ws.merge_cells(f"C{r}:K{r}")
    # Store code in a standalone cell outside the merged range (used for reference)
    # The code is already in row 2 col 3 (value cell covers C2:J2)

    # ── Row 6 — Column headers ────────────────────────────────────────────────
    ws.row_dimensions[6].height = 36
    headers = ["QC/AI", "Testcase ID", "Priority", "Nội dung Test (Test Title)",
               "Điều Kiện/ Dữ Liệu Test", "Các Bước Thực Hiện",
               "Kết quả mong đợi (Expected Response)",
               "Kết quả thực hiện", "Người thực hiện", "Bug ID", "Ghi chú"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(6, col_idx, h)
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill(C_COL_HEADER)
        c.alignment = align("center", "center")
        c.border = make_border()

    # ── Row 7 — Round header ──────────────────────────────────────────────────
    ws.row_dimensions[7].height = 20
    rc = ws.cell(7, 8, "Round 1")
    rc.font = font(bold=True)
    rc.fill = fill(C_ROUND_HEADER)
    rc.alignment = align("center", "center")
    rc.border = make_border()
    ws.merge_cells("H7:K7")
    for col_idx in range(1, 8):
        c = ws.cell(7, col_idx)
        c.fill = fill(C_COL_HEADER)
        c.border = make_border()

    # ── TCs from row 8 ────────────────────────────────────────────────────────
    row = 8
    tc_num = 0
    current_group_color = "FFFFFF"

    for item in api_info["tcs"]:
        if item[0] == "GROUP":
            _, group_name, color_key = item
            current_group_color = GROUP_COLORS.get(color_key, "EEEEEE")
            ws.row_dimensions[row].height = 20
            gc = ws.cell(row, 1, group_name)
            gc.font = font(bold=True)
            gc.fill = fill(current_group_color)
            gc.alignment = align("center", "center")
            gc.border = make_border()
            ws.merge_cells(f"A{row}:J{row}")
            row += 1
            continue

        priority, title, condition, steps, expected = item
        tc_num += 1
        tc_id = f"{api_info['code']}.{tc_num}"

        ws.row_dimensions[row].height = 90

        # Col A — QC/AI
        ca = ws.cell(row, 1, "AI")
        ca.font = font()
        ca.fill = fill("F2F2F2")
        ca.alignment = align("center", "center")
        ca.border = make_border()

        # Col B — TC ID
        cb = ws.cell(row, 2, tc_id)
        cb.font = font(bold=True)
        cb.fill = fill("EBF3FB")
        cb.alignment = align("center", "center")
        cb.border = make_border()

        # Col C — Priority
        p_colors = {"High": C_P_HIGH, "Medium": C_P_MEDIUM, "Low": C_P_LOW}
        cc = ws.cell(row, 3, priority)
        cc.font = font(bold=True, color=p_colors.get(priority, "000000"))
        cc.alignment = align("center", "center")
        cc.border = make_border()

        # Col D — Test Title
        cd = ws.cell(row, 4, title)
        cd.font = font()
        cd.alignment = align("left", "top")
        cd.border = make_border()

        # Col E — Điều Kiện / Dữ Liệu Test
        ce = ws.cell(row, 5, condition)
        ce.font = font()
        ce.alignment = align("left", "top")
        ce.border = make_border()

        # Col F — Các Bước Thực Hiện (chuẩn cho toàn bộ API TC)
        cf = ws.cell(row, 6, STANDARD_STEPS)
        cf.font = font()
        cf.alignment = align("left", "top")
        cf.border = make_border()

        # Col G — Expected Response
        cg = ws.cell(row, 7, expected)
        cg.font = font()
        cg.alignment = align("left", "top")
        cg.border = make_border()

        # Cols H-K — Round 1
        for col_idx in range(8, 12):
            rc = ws.cell(row, col_idx, "")
            rc.border = make_border()

        row += 1

    # Return last row for reference
    return row


def main():
    out_dir = r"E:\AI\Ecom\ecom-pdh\03_test-cases\api"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "AI_ISC_ecom-pdh_v1.1_TC_API_v2.0.xlsx")

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    total_tcs = 0
    for api_info in APIS:
        ws = wb.create_sheet(api_info["sheet"])
        last_row = build_sheet(ws, api_info)
        tc_count = sum(1 for item in api_info["tcs"] if item[0] != "GROUP")
        total_tcs += tc_count
        print(f"  {api_info['code']}: {tc_count} TCs")

    wb.save(out_path)
    print(f"Done: {total_tcs} TCs")
    print(f"File: {out_path}")


if __name__ == "__main__":
    main()
